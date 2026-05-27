#!/usr/bin/env python3
"""
Build per-team Excel tabs from a Throw Down schedule CSV (group + playoff summary).

Usage:
  python3 schedule_data/build_team_schedules.py
  python3 schedule_data/build_team_schedules.py --csv path/to/schedule.csv --out path/to/out.xlsx
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
DEFAULT_CSV = HERE / "schedule_may_27.csv"
DEFAULT_OUT = HERE / "The Throw Down_ 5th Edition Team Schedules - May 27.xlsx"
STREAM_COURT = "Court 1"

FILL_PLAY = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_REF = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_OFF = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

FONT_TITLE = Font(bold=True, size=16)
FONT_SUBTITLE = Font(size=12)
FONT_TABLE_HEADER = Font(bold=True, color="FFFFFF")
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")


@dataclass(frozen=True)
class GameRow:
    date: str
    court: str
    phase: str
    group: str
    round_num: int
    home: str
    away: str
    referees: str

    @classmethod
    def from_dict(cls, row: dict[str, str]) -> GameRow:
        return cls(
            date=(row.get("Date") or "").strip(),
            court=(row.get("Court") or "").strip(),
            phase=(row.get("Phase") or "").strip(),
            group=(row.get("Group") or "").strip(),
            round_num=int((row.get("Round") or "0").strip()),
            home=_clean_team(row.get("Home Team", "")),
            away=_clean_team(row.get("Away Team", "")),
            referees=_clean_team(row.get("Referees", "")),
        )


def _clean_team(s: str) -> str:
    t = (s or "").strip()
    if t.lower() == "undefined":
        return ""
    return t


def load_games(path: Path) -> list[GameRow]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [GameRow.from_dict(dict(r)) for r in reader]


def group_phase_teams(games: list[GameRow]) -> list[str]:
    names: set[str] = set()
    for g in games:
        if g.phase != "Group Phase":
            continue
        if g.home:
            names.add(g.home)
        if g.away:
            names.add(g.away)
    return sorted(names)


def games_by_round(games: list[GameRow], phase: str) -> dict[int, list[GameRow]]:
    out: dict[int, list[GameRow]] = defaultdict(list)
    for g in games:
        if g.phase == phase:
            out[g.round_num].append(g)
    for r in out:
        out[r].sort(key=lambda x: (x.court, x.date))
    return dict(out)


def excel_sheet_title(name: str, used: set[str]) -> str:
    bad = '[]:*?/\\'
    s = "".join("_" if c in bad else c for c in name)
    if len(s) <= 31:
        base = s
    else:
        base = s[:31]
    candidate = base
    n = 2
    while candidate in used:
        suffix = f" ({n})"
        candidate = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        n += 1
    used.add(candidate)
    return candidate


def group_phase_round_stats(
    team: str, gp_by_round: dict[int, list[GameRow]]
) -> tuple[int, int, int, int]:
    """Returns (home_rounds, away_rounds, ref_rounds, off_rounds) for group phase rounds 1–10."""
    gp_rounds = sorted(k for k in gp_by_round if 1 <= k <= 10)
    home_c = away_c = ref_c = off_c = 0
    for rnd in gp_rounds:
        games = gp_by_round[rnd]
        if not games:
            off_c += 1
            continue
        if next((g for g in games if g.home == team), None):
            home_c += 1
        elif next((g for g in games if g.away == team), None):
            away_c += 1
        elif next((g for g in games if g.referees == team), None):
            ref_c += 1
        else:
            off_c += 1
    return home_c, away_c, ref_c, off_c


def group_phase_stream_play_rounds(
    team: str, gp_by_round: dict[int, list[GameRow]]
) -> list[int]:
    """Rounds (1–10) where the team is home or away on the stream court (Court 1)."""
    rounds: list[int] = []
    for rnd in sorted(k for k in gp_by_round if 1 <= k <= 10):
        for g in gp_by_round[rnd]:
            if g.court != STREAM_COURT:
                continue
            if g.home == team or g.away == team:
                rounds.append(rnd)
                break
    return rounds


def write_summary_sheet(
    ws,
    teams: list[str],
    gp_by_round: dict[int, list[GameRow]],
) -> None:
    ws["A1"] = "Group phase — team summary"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:H1")

    ws["A2"] = "Throw Down — group rounds 1–10"
    ws["A2"].font = FONT_SUBTITLE
    ws.merge_cells("A2:H2")

    headers = [
        "Team",
        "Home",
        "Away",
        "Ref",
        "Off",
        "Total rounds",
        "Playing total",
        "Court 1 (stream) — playing rounds",
    ]
    hr = 4
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = FONT_TABLE_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = hr + 1
    for team in teams:
        h, a, r, o = group_phase_round_stats(team, gp_by_round)
        total = h + a + r + o
        play = h + a
        c1_rounds = group_phase_stream_play_rounds(team, gp_by_round)
        c1_str = ", ".join(str(x) for x in c1_rounds) if c1_rounds else "—"
        values = [team, h, a, r, o, total, play, c1_str]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.alignment = ALIGN_WRAP
            if col in (2, 3):
                c.fill = FILL_PLAY
            elif col == 4:
                c.fill = FILL_REF
            elif col == 5:
                c.fill = FILL_OFF
            elif col == 8:
                c.fill = FILL_PLAY
        row_idx += 1

    widths = [28, 10, 10, 10, 10, 14, 14, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_team_sheet(
    ws,
    team: str,
    gp_by_round: dict[int, list[GameRow]],
    playoff_summary: list[tuple[int, str, str, str]],
) -> None:
    ws["A1"] = team
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:G1")

    ws["A2"] = "Throw Down — schedule export"
    ws["A2"].font = FONT_SUBTITLE
    ws.merge_cells("A2:G2")

    headers = ["Round", "Time", "Court", "Status", "Home Team", "Away Team", "Ref Team"]
    hr = 4
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = FONT_TABLE_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = hr + 1
    gp_rounds = sorted(k for k in gp_by_round if 1 <= k <= 10)

    for rnd in gp_rounds:
        games = gp_by_round[rnd]
        if not games:
            continue
        round_time = games[0].date

        playing_home = next((g for g in games if g.home == team), None)
        playing_away = next((g for g in games if g.away == team), None)
        reffing = next((g for g in games if g.referees == team), None)

        if playing_home:
            status = "PLAYING (Home)"
            g = playing_home
            fill = FILL_PLAY
            home_v, away_v, ref_v = g.home, g.away, g.referees
            court_v = g.court
        elif playing_away:
            status = "PLAYING (Away)"
            g = playing_away
            fill = FILL_PLAY
            home_v, away_v, ref_v = g.home, g.away, g.referees
            court_v = g.court
        elif reffing:
            status = "REFFING"
            g = reffing
            fill = FILL_REF
            home_v, away_v, ref_v = g.home, g.away, g.referees
            court_v = g.court
        else:
            status = "OFF"
            fill = FILL_OFF
            home_v, away_v, ref_v = "—", "—", "—"
            court_v = "—"

        values = [rnd, round_time, court_v, status, home_v, away_v, ref_v]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = fill
            c.alignment = ALIGN_WRAP
        row_idx += 1

    n_play = sum(
        1
        for rnd in gp_rounds
        if next((g for g in gp_by_round[rnd] if g.home == team or g.away == team), None)
    )
    n_ref = sum(
        1
        for rnd in gp_rounds
        if next((g for g in gp_by_round[rnd] if g.referees == team), None)
    )
    n_off = len(gp_rounds) - n_play - n_ref
    row_idx += 1
    summary = (
        f"Group phase summary — Playing: {n_play} rounds • "
        f"Reffing: {n_ref} rounds • Off: {n_off} rounds"
    )
    csum = ws.cell(row=row_idx, column=1, value=summary)
    csum.font = Font(italic=True, size=11)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Playoffs — Bracket TBD Based on Seeding").font = Font(
        bold=True, size=12
    )
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
    row_idx += 1

    ph = ["Round", "Time", "Phase / matches", "Courts"]
    for col, h in enumerate(ph, start=1):
        cell = ws.cell(row=row_idx, column=col, value=h)
        cell.font = FONT_TABLE_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row_idx += 1

    for rnd, tme, desc, courts in playoff_summary:
        for col, val in enumerate([rnd, tme, desc, courts], start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.alignment = ALIGN_WRAP
        row_idx += 1

    widths = [10, 22, 12, 18, 22, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_playoff_summary(playoff_by_round: dict[int, list[GameRow]]) -> list[tuple[int, str, str, str]]:
    summary: list[tuple[int, str, str, str]] = []
    for rnd in sorted(playoff_by_round):
        games = playoff_by_round[rnd]
        if not games:
            continue
        tme = games[0].date
        groups = [g.group for g in games]
        unique_groups = list(dict.fromkeys(groups))
        desc = " | ".join(unique_groups) if unique_groups else "Playoff block"
        courts_nums = []
        for g in games:
            c = g.court.replace("Court ", "").strip()
            if c.isdigit():
                courts_nums.append(int(c))
        if courts_nums:
            lo, hi = min(courts_nums), max(courts_nums)
            courts_str = f"Courts {lo}–{hi}" if lo != hi else f"Court {lo}"
        else:
            courts_str = ", ".join(g.court for g in games)
        summary.append((rnd, tme, desc, courts_str))
    return summary


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--csv", type=Path, default=DEFAULT_CSV, help="Input schedule CSV")
    p.add_argument("--out", type=Path, default=DEFAULT_OUT, help="Output xlsx path")
    args = p.parse_args()

    games = load_games(args.csv)
    teams = group_phase_teams(games)
    gp_by_round = games_by_round(games, "Group Phase")
    playoff_by_round = games_by_round(games, "Playoff")
    playoff_summary = build_playoff_summary(playoff_by_round)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_sum = wb.create_sheet("Summary", 0)
    write_summary_sheet(ws_sum, teams, gp_by_round)

    used_titles: set[str] = set()
    for team in teams:
        title = excel_sheet_title(team, used_titles)
        ws = wb.create_sheet(title)
        write_team_sheet(ws, team, gp_by_round, playoff_summary)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Wrote summary + {len(teams)} team sheets to {args.out}")


if __name__ == "__main__":
    main()
