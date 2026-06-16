#!/usr/bin/env python3
"""
Generate print-ready Throw Down round-robin scoresheets from schedule CSV.

Layout matches the "Court Assignment Printouts" tab in the 4th Edition schedule
workbook: Georgia fonts, 6-column cards, 4 courts across, 2 rounds per page.

Usage:
  python3 schedule_data/build_throwdown_scoresheets.py --tabs-only
  python3 schedule_data/build_throwdown_scoresheets.py --merge-into schedule_data/The\\ Throw\\ Down_\\ 5th\\ Edition\\ Team\\ Schedules_may_27.xlsx
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parent
DEFAULT_CSV = REPO_ROOT / "throwdown_5_schedule.csv"
DEFAULT_OUT = HERE / "Throw Down Scoresheets (Print).xlsx"
DEFAULT_TABS_OUT = HERE / "Throw Down Scoresheets Tabs.xlsx"
DEFAULT_SCHEDULE_XLSX = HERE / "The Throw Down_ 5th Edition Team Schedules_may_27.xlsx"

PRINTOUT_SHEET = "Court Assignment Printouts"
LEGACY_TAB_PREFIX = "Scoresheets Court "

GROUP_PHASE = "group phase"
MAX_GROUP_ROUND = 10

CARD_WIDTH = 6
CARD_HEIGHT = 17
COURT_STRIDE = 8  # 6 card columns + 2 gap columns

THICK = Side(style="thick", color="000000")
MEDIUM = Side(style="medium", color="000000")
THIN = Side(style="thin", color="000000")

FONT_COURT = Font(name="Georgia", size=25, bold=True)
FONT_GEORGIA_BOLD = Font(name="Georgia", bold=True)
FONT_GEORGIA_TEAM = Font(name="Georgia", size=14, bold=True)
FONT_GEORGIA_SCORE = Font(name="Georgia", bold=True, underline="single")
FONT_GEORGIA = Font(name="Georgia")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

TIME_FORMAT = "h:mm AM/PM"


@dataclass(frozen=True)
class GameRow:
    date: str
    court: str
    phase: str
    round_num: int
    home: str
    away: str
    referees: str


@dataclass(frozen=True)
class ScoresheetCard:
    court: int
    court_label: str
    round_num: int
    time: time
    time_display: str
    ref: str
    home: str
    away: str
    home_where_next: str
    away_where_next: str
    ref_where_next: str


def _clean_team(s: str) -> str:
    t = (s or "").strip()
    if t.lower() == "undefined":
        return ""
    return t


def court_number(court: str) -> int:
    m = re.search(r"(\d+)", court or "")
    return int(m.group(1)) if m else 0


def parse_game_time(date_str: str) -> tuple[time, str]:
    s = date_str.strip()
    for fmt in ("%Y-%m-%d %I:%M %p", "%Y-%m-%d %H:%M"):
        try:
            dt = datetime.strptime(s, fmt)
            hour = dt.hour % 12 or 12
            display = f"{hour}:{dt.minute:02d} {dt.strftime('%p')}"
            return dt.time().replace(second=0, microsecond=0), display
        except ValueError:
            continue
    m = re.search(r"(\d{1,2}):(\d{2})\s*(am|pm)", s, re.I)
    if m:
        hour = int(m.group(1)) % 12 or 12
        minute = int(m.group(2))
        ampm = m.group(3).upper()
        display = f"{hour}:{minute:02d} {ampm}"
        hour24 = hour
        if ampm == "PM" and hour != 12:
            hour24 += 12
        if ampm == "AM" and hour == 12:
            hour24 = 0
        return time(hour24, minute), display
    return time(9, 0), s


def load_games(path: Path) -> list[GameRow]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [
            GameRow(
                date=(row.get("Date") or "").strip(),
                court=(row.get("Court") or "").strip(),
                phase=(row.get("Phase") or "").strip(),
                round_num=int((row.get("Round") or "0").strip() or 0),
                home=_clean_team(row.get("Home Team", "")),
                away=_clean_team(row.get("Away Team", "")),
                referees=_clean_team(row.get("Referees", "")),
            )
            for row in reader
        ]


def group_phase_games(games: list[GameRow]) -> list[GameRow]:
    return [g for g in games if g.phase.strip().lower() == GROUP_PHASE]


def next_assignment(team: str, round_num: int, group_games: list[GameRow]) -> str:
    if round_num > MAX_GROUP_ROUND:
        return "Bracket"
    for g in group_games:
        if g.round_num != round_num:
            continue
        n = court_number(g.court)
        if g.home == team or g.away == team:
            return f"Court {n}"
        if g.referees == team:
            return f"Ref Court {n}"
    return "OFF"


def build_cards(games: list[GameRow]) -> list[ScoresheetCard]:
    group = group_phase_games(games)
    cards: list[ScoresheetCard] = []
    for g in group:
        if not g.home or not g.away or not g.round_num:
            continue
        nxt = g.round_num + 1
        if g.round_num >= MAX_GROUP_ROUND:
            home_next = away_next = ref_next = "Bracket"
        else:
            home_next = next_assignment(g.home, nxt, group)
            away_next = next_assignment(g.away, nxt, group)
            ref_next = next_assignment(g.referees, nxt, group) if g.referees else "OFF"
        game_time, display = parse_game_time(g.date)
        cards.append(
            ScoresheetCard(
                court=court_number(g.court),
                court_label=g.court,
                round_num=g.round_num,
                time=game_time,
                time_display=display,
                ref=g.referees,
                home=g.home,
                away=g.away,
                home_where_next=home_next,
                away_where_next=away_next,
                ref_where_next=ref_next,
            )
        )
    cards.sort(key=lambda c: (c.court, c.round_num))
    return cards


def _border(left=THIN, top=THIN, right=THIN, bottom=THIN) -> Border:
    return Border(left=left, top=top, right=right, bottom=bottom)


def _set_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value,
    *,
    font: Font = FONT_GEORGIA,
    alignment: Alignment = ALIGN_LEFT,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def _merge(
    ws: Worksheet,
    r1: int,
    c1: int,
    r2: int,
    c2: int,
    value=None,
    *,
    font: Font = FONT_GEORGIA,
    alignment: Alignment = ALIGN_LEFT,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    if value is not None:
        _set_cell(ws, r1, c1, value, font=font, alignment=alignment, border=border, number_format=number_format)


def _apply_card_borders(ws: Worksheet, top_row: int, left_col: int) -> None:
    """Apply border pattern matching the 4th Edition printout card."""
    c0 = left_col
    c1 = left_col + CARD_WIDTH - 1

    def cell(r: int, c: int):
        return ws.cell(row=top_row + r, column=c)

    # Header thick box (rows 0-1)
    for r in (0, 1):
        for c in range(c0, c1 + 1):
            left = THICK if c == c0 else None
            right = THICK if c == c1 else None
            top = THICK if r == 0 else None
            bottom = THICK if r == 1 else None
            cell(r, c).border = _border(
                left or THIN,
                top or THIN,
                right or THIN,
                bottom or THIN,
            )

    medium_box = _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM)
    thin_box = _border(THIN, THIN, THIN, THIN)

    for r in range(3, 12):
        for c in range(c0, c1 + 1):
            cell(r, c).border = medium_box

    for r in range(12, 14):
        for c in range(c0, c1 + 1):
            cell(r, c).border = medium_box

    for r in range(14, 16):
        for c in range(c0, c1 + 1):
            cell(r, c).border = thin_box


def write_card(ws: Worksheet, top_row: int, left_col: int, card: ScoresheetCard) -> None:
    """Write one 6×17 scoresheet card at the given position."""
    c0 = left_col
    c_mid = left_col + 3
    c1 = left_col + CARD_WIDTH - 1
    half_left = (c0, c_mid - 1)
    half_right = (c_mid, c1)

    # Rows 1-2: Court header
    _merge(
        ws,
        top_row,
        c0,
        top_row + 1,
        c1,
        card.court_label,
        font=FONT_COURT,
        alignment=ALIGN_CENTER,
        border=_border(THICK, THICK, THICK, THICK),
    )

    r_round = top_row + 2
    _set_cell(ws, r_round, c0 + 1, "Round ", font=FONT_GEORGIA_BOLD, alignment=ALIGN_RIGHT)
    _set_cell(
        ws,
        r_round,
        c0 + 2,
        card.round_num,
        font=FONT_GEORGIA_BOLD,
        alignment=ALIGN_CENTER,
    )
    _set_cell(
        ws,
        r_round,
        c0 + 3,
        card.time,
        font=FONT_GEORGIA_BOLD,
        alignment=ALIGN_CENTER,
        number_format=TIME_FORMAT,
    )

    r_refs = top_row + 3
    _merge(
        ws,
        r_refs,
        half_left[0],
        r_refs,
        half_left[1],
        f"Ref: {card.ref}",
        font=FONT_GEORGIA_BOLD,
        alignment=ALIGN_LEFT,
    )
    _merge(
        ws,
        r_refs,
        half_right[0],
        r_refs,
        half_right[1],
        f"Next: {card.ref_where_next}",
        font=FONT_GEORGIA_BOLD,
        alignment=ALIGN_LEFT,
    )

    r_teams = top_row + 4
    _merge(ws, r_teams, half_left[0], r_teams, half_left[1], card.home, font=FONT_GEORGIA_TEAM, alignment=ALIGN_CENTER)
    _merge(ws, r_teams, half_right[0], r_teams, half_right[1], card.away, font=FONT_GEORGIA_TEAM, alignment=ALIGN_CENTER)

    r_score = top_row + 5
    _merge(ws, r_score, half_left[0], top_row + 8, half_left[1], "Score", font=FONT_GEORGIA_SCORE, alignment=ALIGN_CENTER)
    _merge(ws, r_score, half_right[0], top_row + 8, half_right[1], "Score", font=FONT_GEORGIA_SCORE, alignment=ALIGN_CENTER)

    r_next = top_row + 9
    _merge(
        ws,
        r_next,
        half_left[0],
        r_next,
        half_left[1],
        f"Where to next:  {card.home_where_next}",
        font=FONT_GEORGIA,
        alignment=ALIGN_LEFT,
    )
    _merge(
        ws,
        r_next,
        half_right[0],
        r_next,
        half_right[1],
        f"Where to next: {card.away_where_next}",
        font=FONT_GEORGIA,
        alignment=ALIGN_LEFT,
    )

    _merge(ws, top_row + 10, c0, top_row + 10, c1, "Comments:", font=FONT_GEORGIA_BOLD, alignment=ALIGN_LEFT)
    _merge(ws, top_row + 11, c0, top_row + 11, c1, "Penalties or cards:", font=FONT_GEORGIA_BOLD, alignment=ALIGN_LEFT)

    r_sig = top_row + 12
    _merge(ws, r_sig, half_left[0], top_row + 13, half_left[1], "Signatures", font=FONT_GEORGIA_SCORE, alignment=ALIGN_CENTER)
    _merge(ws, r_sig, half_right[0], top_row + 13, half_right[1], "Signatures", font=FONT_GEORGIA_SCORE, alignment=ALIGN_CENTER)

    r_player = top_row + 14
    _merge(ws, r_player, half_left[0], top_row + 15, half_left[1], "Player: ", font=FONT_GEORGIA, alignment=ALIGN_LEFT)
    _merge(ws, r_player, half_right[0], top_row + 15, half_right[1], "Player: ", font=FONT_GEORGIA, alignment=ALIGN_LEFT)

    _apply_card_borders(ws, top_row, left_col)


def court_column(court_num: int) -> int:
    return 1 + (court_num - 1) * COURT_STRIDE


def page_start_row(page_index: int) -> int:
    """First card row for each printed page (0-based page index)."""
    if page_index == 0:
        return 1
    return 36 + (page_index - 1) * 35


def card_top_row(round_num: int) -> int:
    page_index = (round_num - 1) // 2
    slot = (round_num - 1) % 2
    return page_start_row(page_index) + slot * CARD_HEIGHT


def configure_printout_columns(ws: Worksheet) -> None:
    for court_idx in range(4):
        base = 1 + court_idx * COURT_STRIDE
        for offset in range(CARD_WIDTH):
            ws.column_dimensions[get_column_letter(base + offset)].width = 13.0
        gap_col = base + CARD_WIDTH
        if gap_col <= 26:
            ws.column_dimensions[get_column_letter(gap_col)].width = 8.88


def write_printout_sheet(ws: Worksheet, cards: list[ScoresheetCard]) -> None:
    ws.title = PRINTOUT_SHEET
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    configure_printout_columns(ws)

    cards_by_court: dict[int, list[ScoresheetCard]] = {}
    for card in cards:
        cards_by_court.setdefault(card.court, []).append(card)

    max_row = 1
    page_count = (MAX_GROUP_ROUND + 1) // 2

    for court_num in sorted(cards_by_court):
        left_col = court_column(court_num)
        for card in cards_by_court[court_num]:
            top_row = card_top_row(card.round_num)
            write_card(ws, top_row, left_col, card)
            max_row = max(max_row, top_row + CARD_HEIGHT - 1)

    for page_index in range(page_count - 1):
        ws.row_breaks.append(Break(id=34 + page_index * 35))

    last_col = get_column_letter(1 + 4 * COURT_STRIDE - 2)
    ws.print_area = f"A1:{last_col}{max_row}"


def is_scoresheets_tab(name: str) -> bool:
    return name == PRINTOUT_SHEET or name.startswith(LEGACY_TAB_PREFIX)


def remove_scoresheets_tabs(wb: Workbook) -> int:
    removed = 0
    for name in list(wb.sheetnames):
        if is_scoresheets_tab(name):
            wb.remove(wb[name])
            removed += 1
    return removed


def build_workbook(cards: list[ScoresheetCard]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title=PRINTOUT_SHEET)
    write_printout_sheet(ws, cards)
    return wb


def merge_into_schedule(
    schedule_path: Path,
    cards: list[ScoresheetCard],
    out_path: Path,
) -> tuple[int, str]:
    wb = load_workbook(schedule_path)
    removed = remove_scoresheets_tabs(wb)
    ws = wb.create_sheet(title=PRINTOUT_SHEET)
    write_printout_sheet(ws, cards)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return removed, PRINTOUT_SHEET


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV, help="Input schedule CSV")
    parser.add_argument("--out", type=Path, default=None, help="Output xlsx path")
    parser.add_argument(
        "--merge-into",
        type=Path,
        default=None,
        help="Existing schedule workbook — append Court Assignment Printouts tab",
    )
    parser.add_argument(
        "--tabs-only",
        action="store_true",
        help="Write only the Court Assignment Printouts tab",
    )
    args = parser.parse_args()

    games = load_games(args.csv)
    cards = build_cards(games)
    if not cards:
        raise SystemExit("No group-phase games found in CSV.")

    if args.merge_into:
        out_path = args.out or args.merge_into
        removed, tab_name = merge_into_schedule(args.merge_into, cards, out_path)
        print(
            f"Merged {len(cards)} scoresheets into {out_path}\n"
            f"  Replaced {removed} old scoresheet tab(s)\n"
            f"  Added tab: {tab_name}"
        )
        return

    out_path = args.out or (DEFAULT_TABS_OUT if args.tabs_only else DEFAULT_OUT)
    wb = build_workbook(cards)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(
        f"Wrote {len(cards)} scoresheets to {out_path}\n"
        f"  Tab: {PRINTOUT_SHEET} (4 courts across, 2 rounds per page)\n"
        f"  Import into your schedule workbook, or run with --merge-into"
    )


if __name__ == "__main__":
    main()
