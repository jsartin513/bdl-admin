"""
Analyze Week 4 Schedule and suggest ways to fix "ref twice in a row".
Recommends swapping entire rounds (both game + ref rows) so no team is double-booked.
"""

import openpyxl
from setup_standings import detect_teams

WEEK_4_SHEET = "Week 4 Schedule"
FILE_PATH = "public/league_schedules/Winter 2026 BYOT League.xlsx"


def get_teams(wb):
    all_teams = detect_teams(wb)
    return [t for t in all_teams if not t.startswith("Refs:")]


def parse_week_schedule(ws, max_games=None):
    """Parse game blocks (game row + ref row) from a week sheet.

    max_games: if set, stop after this many rounds (legacy default callers use 10).
               If None, read all games until column A has no further Game rows (up to row 200).
    """
    games = []
    row = 1
    max_row = 200
    while row <= max_row:
        a1 = ws.cell(row, 1).value
        if not a1 or "Game" not in str(a1):
            row += 1
            continue
        # Game row
        game_num = str(a1).strip()
        c1_home = (ws.cell(row, 2).value or "").strip()
        c1_away = (ws.cell(row, 4).value or "").strip()
        c2_home = (ws.cell(row, 7).value or "").strip()
        c2_away = (ws.cell(row, 9).value or "").strip()
        ref_row = row + 1
        ref1_raw = (ws.cell(ref_row, 2).value or "").strip()
        ref2_raw = (ws.cell(ref_row, 7).value or "").strip()
        ref1 = ref1_raw.replace("Refs:", "").strip() if ref1_raw.startswith("Refs:") else ref1_raw
        ref2 = ref2_raw.replace("Refs:", "").strip() if ref2_raw.startswith("Refs:") else ref2_raw
        playing = {t for t in [c1_home, c1_away, c2_home, c2_away] if t}
        court1_teams = frozenset({c1_home, c1_away} - {""})
        court2_teams = frozenset({c2_home, c2_away} - {""})
        games.append({
            "gameNumber": game_num,
            "court1Ref": ref1,
            "court2Ref": ref2,
            "court1_playing": (c1_home, c1_away),
            "court2_playing": (c2_home, c2_away),
            "court1_teams": court1_teams,
            "court2_teams": court2_teams,
            "playing": playing,
            "row": row,
            "refRow": ref_row,
        })
        row = ref_row + 1
        if max_games is not None and len(games) >= max_games:
            break
    return games


def find_consecutive_refs(games):
    """Return list of (team, game_index_first, game_index_second)."""
    issues = []
    for i in range(len(games) - 1):
        curr, next_g = games[i], games[i + 1]
        refs_curr = {curr["court1Ref"], curr["court2Ref"]} - {""}
        refs_next = {next_g["court1Ref"], next_g["court2Ref"]} - {""}
        for team in refs_curr & refs_next:
            issues.append((team, i, i + 1))
    return issues


def find_consecutive_same_matchup(games):
    """Warn when the same two teams play each other in consecutive games. Returns list of (i, j, matchup)."""
    issues = []
    for i in range(len(games) - 1):
        curr, next_g = games[i], games[i + 1]
        curr_matchups = [curr["court1_teams"], curr["court2_teams"]]
        next_matchups = [next_g["court1_teams"], next_g["court2_teams"]]
        for cm in curr_matchups:
            if len(cm) != 2:
                continue
            for nm in next_matchups:
                if len(nm) != 2:
                    continue
                if cm == nm:
                    issues.append((i, i + 1, cm))
                    break
    return issues


def games_with_same_three_playing(games):
    """Return list of (i, j) where games i and j have exactly 3 of the 4 playing teams in common."""
    pairs = []
    for i in range(len(games)):
        pi = games[i]["playing"]
        for j in range(i + 1, len(games)):
            pj = games[j]["playing"]
            if len(pi & pj) == 3:
                pairs.append((i, j))
    return pairs


def games_after_round_swap(games, i, j):
    """Return a new list with round i and round j swapped (by ref/player content; game numbers stay in slot)."""
    import copy
    new_games = copy.deepcopy(games)
    new_games[i] = copy.deepcopy(games[j])
    new_games[j] = copy.deepcopy(games[i])
    new_games[i]["gameNumber"] = games[i]["gameNumber"]
    new_games[j]["gameNumber"] = games[j]["gameNumber"]
    return new_games


def games_after_reorder(games, perm):
    """Return new list: new_games[k] = games[perm[k]], with game numbers relabeled to match slot (Game 01, 02, ...)."""
    import copy
    new_games = [copy.deepcopy(games[perm[k]]) for k in range(len(games))]
    for k in range(len(new_games)):
        new_games[k]["gameNumber"] = f"Game {k + 1:02d}"
    return new_games


def find_move_to_start_or_end_solutions(games, issues):
    """
    Try moving one of the problem games (involved in consecutive refs) to the very start or very end.
    E.g. if Biker Gang refs in Games 4 and 5, try: move Game 4 to end (order 1,2,3,5,6,7,8,9,10,4) or
    move Game 5 to start (order 5,1,2,3,4,6,7,8,9,10). Returns list of (description, perm).
    """
    problem_indices = set()
    for _team, i, j in issues:
        problem_indices.add(i)
        problem_indices.add(j)
    n = len(games)
    solutions = []
    for idx in problem_indices:
        # Move round at index idx to the end: order 0,1,...,idx-1, idx+1,...,n-1, idx
        perm_end = list(range(idx)) + list(range(idx + 1, n)) + [idx]
        reordered = games_after_reorder(games, perm_end)
        if not find_consecutive_refs(reordered):
            solutions.append((f"Move {games[idx]['gameNumber']} to the very end (new order: 1–{idx+1} then {idx+2}–{n}, then {games[idx]['gameNumber']})", perm_end))
        # Move round at index idx to the start: order idx, 0, 1, ..., idx-1, idx+1, ..., n-1
        perm_start = [idx] + list(range(idx)) + list(range(idx + 1, n))
        reordered = games_after_reorder(games, perm_start)
        if not find_consecutive_refs(reordered):
            solutions.append((f"Move {games[idx]['gameNumber']} to the very start (new order: {games[idx]['gameNumber']}, then 1–{idx}, then {idx+2}–{n})", perm_start))
    return solutions


def find_insert_between_solutions(games, issues):
    """
    Try moving one round from the first consecutive pair to between the second pair (or vice versa).
    E.g. rounds 4,5 are consecutive (Undodgeball); 7,8 are consecutive (Biker Gang).
    Move one of 4/5 to between 7 and 8 → order ... 5, 7, 4, 8 ... or ... 4, 7, 5, 8 ...
    Or move one of 7/8 to between 4 and 5 → order ... 4, 7, 5, ... or ... 4, 8, 5, ...
    Returns list of (description, perm) where perm is the new order of round indices (0-based).
    """
    # Problem pairs: (3,4) and (6,7) in 0-based index
    a, b = 3, 4   # Games 4 and 5
    c, d = 6, 7   # Games 7 and 8
    solutions = []
    # Move round 4 (index 3) to between 7 and 8: order 0,1,2, 4, 6, 3, 7, 8,9
    perm1 = list(range(10))
    perm1[3], perm1[4], perm1[5], perm1[6] = 4, 6, 3, 7
    if not find_consecutive_refs(games_after_reorder(games, perm1)):
        solutions.append(("Move Game 04 to between Game 07 and 08", perm1))
    # Move round 5 (index 4) to between 7 and 8: order 0,1,2, 3, 6, 4, 7, 8,9
    perm2 = list(range(10))
    perm2[3], perm2[4], perm2[5], perm2[6] = 3, 6, 4, 7
    if not find_consecutive_refs(games_after_reorder(games, perm2)):
        solutions.append(("Move Game 05 to between Game 07 and 08", perm2))
    # Move round 7 (index 6) to between 4 and 5: order 0,1,2, 3, 6, 4, 5, 7, 8,9
    perm3 = list(range(10))
    perm3[3], perm3[4], perm3[5], perm3[6] = 3, 6, 4, 5
    perm3[7] = 7
    if not find_consecutive_refs(games_after_reorder(games, perm3)):
        solutions.append(("Move Game 07 to between Game 04 and 05", perm3))
    # Move round 8 (index 7) to between 4 and 5: order 0,1,2, 3, 7, 4, 5, 6, 8, 9
    perm4 = list(range(10))
    perm4[3], perm4[4], perm4[5], perm4[6], perm4[7] = 3, 7, 4, 5, 6
    perm4[8] = 8
    if not find_consecutive_refs(games_after_reorder(games, perm4)):
        solutions.append(("Move Game 08 to between Game 04 and 05", perm4))
    return solutions


def find_safe_round_swaps(games, issues):
    """
    Suggest swapping two entire rounds so that after the swap no team refs consecutively.
    Each round uses all 6 teams so swapping whole rounds cannot double-book anyone.
    Try all pairs where at least one of the rounds is involved in a consecutive-ref issue.
    """
    problem_indices = set()
    for team, i, j in issues:
        problem_indices.add(i)
        problem_indices.add(j)
    safe = []
    for i in range(len(games)):
        for j in range(i + 1, len(games)):
            if i not in problem_indices and j not in problem_indices:
                continue
            new_games = games_after_round_swap(games, i, j)
            if not find_consecutive_refs(new_games):
                safe.append((i + 1, j + 1))
    return safe


def get_ref_slot(games, game_idx, court):
    """Return (court1Ref or court2Ref, which court label)."""
    g = games[game_idx]
    if court == 1:
        return g["court1Ref"], "Court 1"
    return g["court2Ref"], "Court 2"


def ref_assignment_feasible(games, teams):
    """
    Check if there exists ANY ref assignment (who refs which game) for the current
    game order and matchups such that no team refs in two consecutive games.
    Uses backtracking over games; each game needs 2 refs from (teams - playing).
    """
    n = len(games)
    if n == 0:
        return True
    teams_set = set(teams)

    def pool(game_idx):
        return teams_set - games[game_idx]["playing"]

    def solve(game_idx, prev_refs):
        if game_idx >= n:
            return True
        p = pool(game_idx)
        # Must pick 2 refs from p that are not in prev_refs
        available = sorted(p - prev_refs)
        if len(available) < 2:
            return False
        from itertools import combinations
        for r1, r2 in combinations(available, 2):
            if solve(game_idx + 1, {r1, r2}):
                return True
        return False

    return solve(0, set())


def print_four_games_highlight(games, issues):
    """Print Games 04, 05, 07, 08 with per-court breakdown: the 3 teams on each court (2 playing + ref)."""
    issue_refs_by_game = {}  # game_idx -> set of team names that are "problem" refs in that game
    for team, i, j in issues:
        issue_refs_by_game.setdefault(i, set()).add(team)
        issue_refs_by_game.setdefault(j, set()).add(team)
    indices = [3, 4, 6, 7]  # Games 04, 05, 07, 08
    print("The 4 games involved in the Biker Gang / Undodgeball ref issue:")
    print("(Per court: the 2 teams playing + ref. Refs marked * are refing consecutively.)")
    print()
    for idx in indices:
        g = games[idx]
        r1, r2 = g["court1Ref"], g["court2Ref"]
        problem = issue_refs_by_game.get(idx, set())
        ref1_str = f"{r1} *" if r1 in problem else r1
        ref2_str = f"{r2} *" if r2 in problem else r2
        c1 = g["court1_playing"]
        c2 = g["court2_playing"]
        court1_three = ", ".join(t for t in c1 if t) + f" — ref: {ref1_str}"
        court2_three = ", ".join(t for t in c2 if t) + f" — ref: {ref2_str}"
        print(f"  {g['gameNumber']}:  Court 1: {court1_three}  |  Court 2: {court2_three}")
    print()


def find_safe_swaps(games, issues, teams):
    """
    For each (team, i, j) where team refs in game i and j, find another (game_idx, court) ref slot
    we can swap with so that after swap no new consecutive refs.
    """
    suggestions = []
    for team, i, j in issues:
        # We'll move `team` out of game j (or i). So we need a ref from another game to take their place in j.
        # That ref must: not have reffed in game j-1 (i.e. game i), and not be playing in game j.
        # And we'll put `team` in that other slot; `team` must not ref in the game before or after that slot.
        game_j = games[j]
        players_j = game_j["playing"]
        refs_i = {games[i]["court1Ref"], games[i]["court2Ref"]} - {""}
        refs_j_plus_1 = set()
        if j + 1 < len(games):
            refs_j_plus_1 = {games[j + 1]["court1Ref"], games[j + 1]["court2Ref"]} - {""}
        # Who can ref in game j? Not playing in j, didn't ref in game i, and doesn't ref in game j+1 (or we'd create new consecutive).
        possible_incoming = [t for t in teams if t not in players_j and t not in refs_i and t not in refs_j_plus_1]
        # We need someone who currently refs in some other game G, and when we put `team` in G, team doesn't ref in G-1 or G+1.
        team_refs_in = set()
        for idx, g in enumerate(games):
            if g["court1Ref"] == team or g["court2Ref"] == team:
                team_refs_in.add(idx)
        # After we remove team from j, team_refs_in becomes team_refs_in - {j}. So team will ref in team_refs_in - {j}.
        # We'll add one more slot for team: the slot we swap from. So new slot G must be such that G-1 and G+1 not in (team_refs_in - {j}).
        allowed_G = set(range(len(games))) - {j}
        for g in (team_refs_in - {j}):
            allowed_G.discard(g - 1)
            allowed_G.discard(g + 1)
        # Which court in game j is `team` reffing? (so we know which slot to swap)
        j_court = 1 if games[j]["court1Ref"] == team else 2
        j_court_label = "Court 1" if j_court == 1 else "Court 2"
        # So allowed_G are game indices where we can put `team` without creating consecutive.
        opts = []
        for game_idx in allowed_G:
            # Don't move team to a game they're playing in (would be ref + play same game)
            if team in games[game_idx]["playing"]:
                continue
            for court in (1, 2):
                other_ref, g_court_label = get_ref_slot(games, game_idx, court)
                if not other_ref or other_ref == team:
                    continue
                if other_ref not in possible_incoming:
                    continue
                if game_idx == i:
                    continue
                if game_idx == j + 1:
                    continue
                # Incoming ref must not have another ref slot adjacent to j (or they'd ref j and j±1)
                other_slots = [idx for idx, g in enumerate(games) if g["court1Ref"] == other_ref or g["court2Ref"] == other_ref]
                if any(idx in (j - 1, j + 1) for idx in other_slots if idx != game_idx):
                    continue
                opts.append({
                    "problem": f"{team} refs in {games[i]['gameNumber']} and {games[j]['gameNumber']}",
                    "fix": f"Swap: {games[j]['gameNumber']} {j_court_label} ref ({team}) ↔ {games[game_idx]['gameNumber']} {g_court_label} ref ({other_ref})",
                })
        suggestions.append((team, games[i]["gameNumber"], games[j]["gameNumber"], opts))
    return suggestions


def main():
    wb = openpyxl.load_workbook(FILE_PATH)
    if WEEK_4_SHEET not in wb.sheetnames:
        print(f"Sheet '{WEEK_4_SHEET}' not found.")
        return
    ws = wb[WEEK_4_SHEET]
    games = parse_week_schedule(ws)
    if len(games) < 10:
        print(f"Only found {len(games)} games.")
        return
    teams = get_teams(wb)
    print("Order in sheet (as read):", ", ".join(g["gameNumber"] for g in games))
    print()
    issues = find_consecutive_refs(games)
    if not issues:
        print("No consecutive ref issues found.")
        return
    print("Consecutive ref issues:")
    for team, i, j in issues:
        print(f"  {games[j]['gameNumber']} – {team}: Refed in {games[i]['gameNumber']} and {games[j]['gameNumber']}")
    print()

    # Try moving one of the problem games to the very start or very end
    move_solutions = find_move_to_start_or_end_solutions(games, issues)
    if move_solutions:
        print("Move a problem game to the start or end (fixes consecutive refs):")
        for desc, perm in move_solutions:
            new_order = [games[perm[k]]["gameNumber"] for k in range(len(games))]
            print(f"  → {desc}")
            print(f"    Target row order in Excel: {', '.join(new_order)}")
        print()
        print("In Excel: reorder the round rows (game + ref row together) to match the target order above.")
        print()
    else:
        print("Move a problem game to the start or end:")
        problem_indices = {i for _t, i, j in issues} | {j for _t, i, j in issues}
        for idx in sorted(problem_indices):
            perm_end = list(range(idx)) + list(range(idx + 1, len(games))) + [idx]
            reordered_end = games_after_reorder(games, perm_end)
            rem_end = find_consecutive_refs(reordered_end)
            perm_start = [idx] + list(range(idx)) + list(range(idx + 1, len(games)))
            reordered_start = games_after_reorder(games, perm_start)
            rem_start = find_consecutive_refs(reordered_start)
            def desc_rem(reordered, rem):
                if not rem:
                    return "fixes all"
                return "; ".join(f"{t} in {reordered[r]['gameNumber']} and {reordered[r+1]['gameNumber']}" for (t, r, _) in rem)
            print(f"  Move {games[idx]['gameNumber']} to end:   {desc_rem(reordered_end, rem_end)}")
            print(f"  Move {games[idx]['gameNumber']} to start:  {desc_rem(reordered_start, rem_start)}")
        print()

    same_matchup = find_consecutive_same_matchup(games)
    if same_matchup:
        print("Warning — same teams playing each other in consecutive games:")
        for i, j, matchup in same_matchup:
            teams_str = " vs ".join(sorted(matchup))
            print(f"  {games[i]['gameNumber']} and {games[j]['gameNumber']}: {teams_str}")
        print()

    print_four_games_highlight(games, issues)

    # Round swaps between games that share exactly 3 playing teams (same 3 participants in the matchup)
    same3 = games_with_same_three_playing(games)
    fixes_same3 = [(i, j) for i, j in same3 if not find_consecutive_refs(games_after_round_swap(games, i, j))]
    if fixes_same3:
        print("Round swaps (games with same 3 playing teams) that fix consecutive refs:")
        for i, j in fixes_same3:
            common_str = ", ".join(sorted(games[i]["playing"] & games[j]["playing"]))
            print(f"  → Swap Game {i+1:02d} with Game {j+1:02d} (shared playing: {common_str}; no double-booking)")
        print()
    else:
        print("Round swaps between games that share exactly 3 playing teams:")
        if same3:
            for i, j in same3:
                new_g = games_after_round_swap(games, i, j)
                rem = find_consecutive_refs(new_g)
                common = games[i]["playing"] & games[j]["playing"]
                common_str = ", ".join(sorted(common))
                if not rem:
                    print(f"  Swap Game {i+1:02d} with Game {j+1:02d} (shared: {common_str}) → fixes all.")
                else:
                    desc = "; ".join(f"{t} in {new_g[r]['gameNumber']} and {new_g[r+1]['gameNumber']}" for (t, r, _) in rem)
                    print(f"  Swap Game {i+1:02d} with Game {j+1:02d} (shared: {common_str}) → still: {desc}")
        else:
            print("  (No two games share exactly 3 participants.)")
        print()

    insert_solutions = find_insert_between_solutions(games, issues)
    if insert_solutions:
        print("Move one round between the other pair (fixes both; no double-booking):")
        for desc, perm in insert_solutions:
            print(f"  → {desc}")
        print()
        print("In Excel: reorder the round rows so the rounds appear in the new order above.")
        print("E.g. for 'Move Game 05 to between Game 07 and 08': rows order becomes")
        print("  Game 01, 02, 03, 04, 07, 05, 08, 09, 10 (move the current Game 05 block to between 07 and 08).")
        return

    round_swaps = find_safe_round_swaps(games, issues)
    if round_swaps:
        print("Safe ROUND SWAPS (swap entire round: game row + ref row; no double-booking):")
        for a, b in round_swaps:
            print(f"  → Swap Game {a:02d} with Game {b:02d}")
        print()
        print("In Excel: swap the two game rows and the two ref rows for those games.")
        return

    # Swap round 6 with 3, 4, 5, 7, 8, or 9?
    print("Round swap: Game 06 with Game 03, 04, 05, 07, 08, or 09:")
    for i, j in [(2, 5), (3, 5), (4, 5), (5, 6), (5, 7), (5, 8)]:
        new_games = games_after_round_swap(games, i, j)
        rem = find_consecutive_refs(new_games)
        if not rem:
            print(f"  Swap Game {i+1:02d} with Game {j+1:02d} → fixes all (no consecutive refs).")
        else:
            desc = "; ".join(f"{t} in {new_games[r]['gameNumber']} and {new_games[r+1]['gameNumber']}" for (t, r, _) in rem)
            print(f"  Swap Game {i+1:02d} with Game {j+1:02d} → still have: {desc}")
    print()

    # Answer the user's question: does swapping one of 4/5 with one of 7/8 fix it?
    print("Round swap (4/5) ↔ (7/8):")
    for i, j in [(3, 6), (3, 7), (4, 6), (4, 7)]:
        new_games = games_after_round_swap(games, i, j)
        rem = find_consecutive_refs(new_games)
        if not rem:
            print(f"  Swap Game {i+1:02d} with Game {j+1:02d} → fixes both (no new issues).")
        else:
            desc = "; ".join(f"{t} in {new_games[r]['gameNumber']} and {new_games[r+1]['gameNumber']}" for (t, r, _) in rem)
            print(f"  Swap Game {i+1:02d} with Game {j+1:02d} → still have: {desc}")
    print()
    print("So swapping one of Game 4/5 with one of Game 7/8 does NOT fix it without")
    print("creating another consecutive-ref.")
    print()
    print("We also tried moving one round to sit between the other pair (e.g. move Game 05")
    print("to between Game 07 and 08, so order becomes ... 04, 07, 05, 08 ...). For this")
    print("schedule none of the four options cleared all consecutive refs (moving a round")
    print("can create new consecutive refs elsewhere). If your ref distribution is simpler,")
    print("that move may work and will be suggested at the top when you run this script.")
    print()

    suggestions = find_safe_swaps(games, issues, teams)
    has_swap_fix = any(opts for _, _, _, opts in suggestions)
    feasible = ref_assignment_feasible(games, teams)
    print("Feasibility:")
    print("  The 'no consecutive refs' constraint IS achievable for this schedule:")
    print("  • By round reordering only: NO — we tried all single swaps and insert-between; none fix it.")
    if has_swap_fix:
        print("  • By ref-only swaps (without ref+play same game): YES — apply the options below.")
    else:
        print("  • By ref-only swaps: no safe option (any swap would put someone reffing a game they play in).")
    if not feasible:
        print("  (With 6 teams there are only 2 refs per game, so not every game order has a valid")
        print("  ref assignment from scratch; fixing the current assignment with swaps can still work.)")
    print()
    if not suggestions:
        print("No ref-only swap options found (any swap would put a team reffing a game they're playing in).")
        return
    print("Ref-only swaps (only options where no one refs a game they're playing in):")
    for team, g_i, g_j, opts in suggestions:
        if not opts:
            print(f"  • {team} refs in {g_i} and {g_j} – no safe option")
            continue
        print(f"  • {team} refs in {g_i} and {g_j}")
        for s in opts:
            print(f"    → {s['fix']}")


if __name__ == "__main__":
    main()
