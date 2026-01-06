"""
Generate week sheets from the Schedule Generator sheet.
Reads the matchup matrix and creates game schedules for each week.
"""

import openpyxl
import re
from itertools import combinations

def get_teams_from_schedule_generator(ws):
    """Extract team names from Schedule Generator sheet."""
    teams = []
    # Teams are typically in row 2, starting from column 2
    for col in range(2, 10):
        cell_value = ws.cell(2, col).value
        if cell_value and isinstance(cell_value, str) and cell_value.strip():
            teams.append(cell_value.strip())
    return teams

def get_matchup_matrix(ws, teams):
    """Read the matchup matrix from Schedule Generator."""
    matrix = {}
    
    # Find team rows (typically starting at row 3)
    team_row_map = {}
    for row in range(3, 20):
        team_name = ws.cell(row, 1).value
        if team_name and team_name in teams:
            team_row_map[team_name] = row
    
    # Read matchups
    for team in teams:
        matrix[team] = {}
        if team in team_row_map:
            row = team_row_map[team]
            for i, opponent in enumerate(teams, start=2):
                if opponent != team:
                    value = ws.cell(row, i).value
                    # Convert to int if it's a number
                    if isinstance(value, (int, float)):
                        matrix[team][opponent] = int(value)
                    elif isinstance(value, str) and value.strip().isdigit():
                        matrix[team][opponent] = int(value.strip())
                    else:
                        matrix[team][opponent] = 0
    
    return matrix

def generate_round_robin_schedule(teams, games_per_matchup=2):
    """Generate a round-robin schedule where each team plays each other team N times."""
    schedule = []
    game_num = 1
    
    # Generate all possible matchups
    for team1, team2 in combinations(teams, 2):
        for _ in range(games_per_matchup):
            schedule.append({
                'game': game_num,
                'team1': team1,
                'team2': team2
            })
            game_num += 1
    
    return schedule

def distribute_games_into_weeks(schedule, num_weeks):
    """Distribute games evenly across weeks."""
    games_per_week = len(schedule) // num_weeks
    if len(schedule) % num_weeks != 0:
        games_per_week += 1
    
    weeks = [[] for _ in range(num_weeks)]
    
    for i, game in enumerate(schedule):
        week_index = i // games_per_week
        if week_index >= num_weeks:
            week_index = num_weeks - 1
        weeks[week_index].append(game)
    
    return weeks

def create_week_sheet(wb, week_name, games, teams):
    """Create or update a week sheet with the game schedule."""
    if week_name in wb.sheetnames:
        ws = wb[week_name]
        # Clear existing games (keep headers and win/loss section)
        for row in range(2, 30):
            for col in range(1, 6):
                if row < 25:  # Don't clear win/loss section
                    ws.cell(row, col).value = None
    else:
        ws = wb.create_sheet(week_name)
        # Add header
        ws.cell(1, 2).value = 'Court 1'
    
    # Add games
    row = 2
    for game in games:
        # Game number
        ws.cell(row, 1).value = f'Game {game["game"]:02d}'
        # Team 1 (home)
        ws.cell(row, 2).value = game['team1']
        # Team 2 (away)
        ws.cell(row, 4).value = game['team2']
        # Empty row for ref
        row += 1
        ws.cell(row, 2).value = 'Ref'
        row += 1
    
    return ws

def main(file_path, num_weeks=6):
    """Main function to generate schedule from Schedule Generator."""
    print(f"Loading workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    
    if 'Schedule Generator' not in wb.sheetnames:
        print("ERROR: Schedule Generator sheet not found!")
        return
    
    ws = wb['Schedule Generator']
    
    print("\n=== Reading Schedule Generator ===")
    teams = get_teams_from_schedule_generator(ws)
    print(f"Found {len(teams)} teams: {', '.join(teams)}")
    
    if not teams:
        print("ERROR: No teams found in Schedule Generator!")
        return
    
    matrix = get_matchup_matrix(ws, teams)
    print("\nMatchup matrix:")
    for team in teams:
        print(f"  {team}: {matrix[team]}")
    
    # Determine games per matchup (typically 2)
    games_per_matchup = 2
    if teams:
        # Check first matchup to see how many games
        first_team = teams[0]
        if teams[1] in matrix[first_team]:
            games_per_matchup = matrix[first_team][teams[1]]
    
    print(f"\n=== Generating Schedule ===")
    print(f"Games per matchup: {games_per_matchup}")
    print(f"Number of weeks: {num_weeks}")
    
    schedule = generate_round_robin_schedule(teams, games_per_matchup)
    print(f"Total games: {len(schedule)}")
    
    weeks = distribute_games_into_weeks(schedule, num_weeks)
    
    print(f"\n=== Creating Week Sheets ===")
    week_names = []
    for i, week_games in enumerate(weeks, start=1):
        # Try to find existing week sheet name pattern
        week_name = None
        for sheet_name in wb.sheetnames:
            if re.match(rf'Week\s+{i}\s*\(', sheet_name, re.IGNORECASE):
                week_name = sheet_name
                break
        
        if not week_name:
            # Create new name - you might want to customize this
            week_name = f'Week {i}'
        
        print(f"  {week_name}: {len(week_games)} games")
        create_week_sheet(wb, week_name, week_games, teams)
        week_names.append(week_name)
    
    print(f"\n=== Saving Workbook ===")
    wb.save(file_path)
    print(f"✅ Schedule generated! Created/updated {len(week_names)} week sheets.")
    
    print(f"\nNext steps:")
    print(f"  1. Review the generated week sheets")
    print(f"  2. Run setup_standings.py to add win/loss formulas")
    print(f"  3. Manually adjust game order if needed")

if __name__ == '__main__':
    import sys
    import os
    
    # Check if arguments provided (non-interactive mode)
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        num_weeks = int(sys.argv[2]) if len(sys.argv) > 2 else 6
    else:
        # Interactive mode
        print("📅 Schedule Generator")
        print("=" * 50)
        print()
        
        # Find existing spreadsheets
        default_dir = 'public/league_schedules'
        existing_files = []
        if os.path.exists(default_dir):
            existing_files = [f for f in os.listdir(default_dir) if f.endswith('.xlsx')]
        
        if existing_files:
            print("Existing league files:")
            for i, f in enumerate(existing_files, 1):
                print(f"  {i}. {f}")
            print()
        
        file_path = input("Spreadsheet file path: ").strip()
        if not file_path:
            if existing_files:
                file_path = os.path.join(default_dir, existing_files[0])
                print(f"Using: {file_path}")
            else:
                print("❌ No file specified")
                sys.exit(1)
        
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            sys.exit(1)
        
        num_weeks_input = input("Number of weeks (default: 6): ").strip()
        num_weeks = int(num_weeks_input) if num_weeks_input.isdigit() else 6
        
        print()
        print(f"📋 Summary:")
        print(f"  File: {file_path}")
        print(f"  Weeks: {num_weeks}")
        
        confirm = input("\nGenerate schedule? (Y/n): ").strip().lower()
        if confirm and confirm != 'y':
            print("Cancelled.")
            sys.exit(0)
        
        print()
    
    main(file_path, num_weeks)

