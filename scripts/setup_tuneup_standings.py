#!/usr/bin/env python3
"""
Wire BDL Tune Up Tournament Schedule_Bracket.xlsx:

- Score entry cells on Team Schedule (cols D / F)
- Matchup W/L/D + standings (3 pts win, 1 pt draw; seed by Points then H2H, PF fallback)
- 5-team single-elim bracket seeded from standings
- Cleanup leftover Throw Down junk / #REF! formulas

Usage:
  python3 scripts/setup_tuneup_standings.py
  python3 scripts/setup_tuneup_standings.py --workbook "BDL Tune Up Tournament Schedule_Bracket.xlsx"
  python3 scripts/setup_tuneup_standings.py --smoke-test
  python3 scripts/setup_tuneup_standings.py --clear-scores
"""

from __future__ import annotations

import argparse
import re
from datetime import time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "BDL Tune Up Tournament Schedule_Bracket.xlsx"

RR_BLOCKS = [2, 6, 10, 14, 18]
COURT_TEAM_COLS = (3, 5)  # C, E
COURT_SCORE_COLS = (4, 6)  # D, F

THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SEED_FILL = PatternFill("solid", fgColor="D6EAF8")
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
PLAY_FILL = PatternFill("solid", fgColor="FFF2CC")
FINAL_FILL = PatternFill("solid", fgColor="C6EFCE")


def col_letter(col: int) -> str:
    return get_column_letter(col)


def cell_addr(col: int, row: int) -> str:
    return f"{col_letter(col)}{row}"


def read_teams_from_rosters(ws) -> list[str]:
    teams: list[str] = []
    for col in range(3, 13, 2):
        name = ws.cell(1, col).value
        if isinstance(name, str) and name.strip():
            teams.append(name.strip())
    if len(teams) != 5:
        raise ValueError(f"Expected 5 teams on Team Rosters, found {teams!r}")
    return teams


def discover_matchups(ws_schedule) -> list[dict]:
    matchups: list[dict] = []
    for start in RR_BLOCKS:
        for team_col, score_col in zip(COURT_TEAM_COLS, COURT_SCORE_COLS):
            a = ws_schedule.cell(start, team_col).value
            b = ws_schedule.cell(start + 1, team_col).value
            if not (isinstance(a, str) and isinstance(b, str)):
                continue
            a, b = a.strip(), b.strip()
            if not a or not b or a == b:
                continue
            matchups.append(
                {
                    "team_a": a,
                    "team_b": b,
                    "score_a": cell_addr(score_col, start),
                    "score_b": cell_addr(score_col, start + 1),
                }
            )
    if len(matchups) != 10:
        raise ValueError(f"Expected 10 RR matchups, found {len(matchups)}: {matchups}")
    return matchups


def games_for_team(team: str, matchups: list[dict]) -> list[tuple[str, str]]:
    games: list[tuple[str, str]] = []
    for m in matchups:
        if m["team_a"] == team:
            games.append((m["score_a"], m["score_b"]))
        elif m["team_b"] == team:
            games.append((m["score_b"], m["score_a"]))
    if len(games) != 4:
        raise ValueError(f"{team} should play 4 games, found {len(games)}: {games}")
    return games


def h2h_cells(team: str, opponent: str, matchups: list[dict]) -> tuple[str, str] | None:
    for m in matchups:
        if {m["team_a"], m["team_b"]} == {team, opponent}:
            if m["team_a"] == team:
                return m["score_a"], m["score_b"]
            return m["score_b"], m["score_a"]
    return None


def sum_if_parts(games: list[tuple[str, str]], kind: str) -> str:
    parts: list[str] = []
    for own, opp in games:
        both = f"AND(ISNUMBER('Team Schedule'!{own}),ISNUMBER('Team Schedule'!{opp}))"
        own_ref = f"'Team Schedule'!{own}"
        opp_ref = f"'Team Schedule'!{opp}"
        if kind == "W":
            parts.append(f"IF({both},IF({own_ref}>{opp_ref},1,0),0)")
        elif kind == "L":
            parts.append(f"IF({both},IF({own_ref}<{opp_ref},1,0),0)")
        elif kind == "D":
            parts.append(f"IF({both},IF({own_ref}={opp_ref},1,0),0)")
        elif kind == "PF":
            parts.append(f"IF(ISNUMBER({own_ref}),{own_ref},0)")
        elif kind == "PA":
            parts.append(f"IF({both},{opp_ref},0)")
        else:
            raise ValueError(kind)
    return "=" + "+".join(parts)


def unmerge_overlapping(ws, min_col: int, min_row: int, max_col: int, max_row: int) -> None:
    to_remove = []
    for merged in list(ws.merged_cells.ranges):
        if not (
            merged.max_col < min_col
            or merged.min_col > max_col
            or merged.max_row < min_row
            or merged.min_row > max_row
        ):
            to_remove.append(str(merged))
    for rng in to_remove:
        ws.unmerge_cells(rng)


def unmerge_all(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))


def clear_sheet_values(ws, max_row: int = 60, max_col: int = 20) -> None:
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()
            cell.alignment = Alignment()


def label_schedule_scores(ws) -> None:
    unmerge_overlapping(ws, 3, 1, 6, 1)
    ws["C1"] = "Court 1"
    ws["D1"] = "Score"
    ws["E1"] = "Court 2"
    ws["F1"] = "Score"
    for addr in ("C1", "D1", "E1", "F1"):
        ws[addr].font = Font(bold=True)
    for start in RR_BLOCKS:
        for score_col in COURT_SCORE_COLS:
            for row in (start, start + 1):
                cell = ws.cell(row, score_col)
                if cell.value is None or cell.value == "":
                    cell.value = None
                cell.alignment = Alignment(horizontal="center")


def clear_schedule_junk(ws) -> None:
    unmerge_overlapping(ws, 8, 1, 12, 40)
    unmerge_overlapping(ws, 17, 1, 20, 40)
    for row in range(1, 40):
        for col in list(range(8, 13)) + list(range(17, 21)):
            ws.cell(row, col).value = None


def wire_playoff_labels(ws) -> None:
    unmerge_overlapping(ws, 13, 1, 19, 30)

    ws["M2"] = "Playoffs"
    ws["M2"].font = SECTION_FONT
    for col, h in ((17, "Score"), (18, "Score"), (19, "Winner")):
        ws.cell(4, col, h).font = Font(bold=True)

    ws["M5"] = "Play-In Game"
    if ws["N5"].value is None:
        ws["N5"] = time(12, 35)

    ws["M8"] = "Break (5 min)"
    ws["M9"] = "Semi Final 1"
    if ws["N9"].value is None:
        ws["N9"] = time(13, 20)

    ws["M12"] = "Break (5 min)"
    ws["M13"] = "Semi Final 2"
    if ws["N13"].value is None:
        ws["N13"] = time(14, 5)

    ws["M16"] = "Break (5 min)"
    ws["M17"] = "Championship"
    if ws["N17"].value is None:
        ws["N17"] = time(14, 50)

    ws["M20"] = "Awards"
    if ws["N20"].value is None:
        ws["N20"] = time(15, 30)


def rebuild_standings(ws, teams: list[str], matchups: list[dict]) -> None:
    unmerge_all(ws)
    clear_sheet_values(ws, max_row=40, max_col=20)

    ws["A1"] = "Pool Standings"
    ws["A1"].font = TITLE_FONT

    headers = [
        "Team",
        "Wins",
        "Losses",
        "Draws",
        "Points",
        "Points For",
        "Points Against",
        "Differential",
        "Sort Key",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(2, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center")

    for i, team in enumerate(teams):
        row = 3 + i
        games = games_for_team(team, matchups)
        ws.cell(row, 1, team).border = THIN
        ws.cell(row, 2).value = sum_if_parts(games, "W")
        ws.cell(row, 3).value = sum_if_parts(games, "L")
        ws.cell(row, 4).value = sum_if_parts(games, "D")
        ws.cell(row, 5).value = f"=B{row}*3+D{row}"
        ws.cell(row, 6).value = sum_if_parts(games, "PF")
        ws.cell(row, 7).value = sum_if_parts(games, "PA")
        ws.cell(row, 8).value = f"=F{row}-G{row}"
        for col in range(2, 9):
            ws.cell(row, col).border = THIN
            ws.cell(row, col).alignment = Alignment(horizontal="center")

    for i, team in enumerate(teams):
        row = 3 + i
        h2h_parts: list[str] = []
        for j, opp in enumerate(teams):
            if i == j:
                continue
            opp_row = 3 + j
            cells = h2h_cells(team, opp, matchups)
            if not cells:
                continue
            own_c, opp_c = cells
            both = (
                f"AND(ISNUMBER('Team Schedule'!{own_c}),"
                f"ISNUMBER('Team Schedule'!{opp_c}))"
            )
            tied = f"$E{row}=$E{opp_row}"
            won = f"'Team Schedule'!{own_c}>'Team Schedule'!{opp_c}"
            h2h_parts.append(f"IF(AND({tied},{both},{won}),0.01,0)")
        h2h_expr = "+".join(h2h_parts) if h2h_parts else "0"
        ws.cell(row, 9).value = f"=E{row}+({h2h_expr})+F{row}/10000"
        ws.cell(row, 9).border = THIN
        ws.cell(row, 9).number_format = "0.0000"

    ws.column_dimensions["I"].hidden = True

    ws["A10"] = "Seeding (after pool play)"
    ws["A10"].font = SECTION_FONT

    seed_headers = ["Seed", "Team", "Wins", "Losses", "Draws", "Points", "PF", "PA", "Diff"]
    for col, h in enumerate(seed_headers, start=1):
        cell = ws.cell(11, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN

    raw_lo, raw_hi = 3, 7
    key_rng = f"$I${raw_lo}:$I${raw_hi}"
    for seed_n in range(1, 6):
        row = 11 + seed_n
        large = f"LARGE({key_rng},{seed_n})"
        ws.cell(row, 1, seed_n).border = THIN
        ws.cell(row, 1).fill = SEED_FILL
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 2).value = f"=INDEX($A${raw_lo}:$A${raw_hi},MATCH({large},{key_rng},0))"
        ws.cell(row, 3).value = f"=INDEX($B${raw_lo}:$B${raw_hi},MATCH({large},{key_rng},0))"
        ws.cell(row, 4).value = f"=INDEX($C${raw_lo}:$C${raw_hi},MATCH({large},{key_rng},0))"
        ws.cell(row, 5).value = f"=INDEX($D${raw_lo}:$D${raw_hi},MATCH({large},{key_rng},0))"
        ws.cell(row, 6).value = f"=INDEX($E${raw_lo}:$E${raw_hi},MATCH({large},{key_rng},0))"
        ws.cell(row, 7).value = f"=INDEX($F${raw_lo}:$F${raw_hi},MATCH({large},{key_rng},0))"
        ws.cell(row, 8).value = f"=INDEX($G${raw_lo}:$G${raw_hi},MATCH({large},{key_rng},0))"
        ws.cell(row, 9).value = f"=INDEX($H${raw_lo}:$H${raw_hi},MATCH({large},{key_rng},0))"
        for col in range(2, 10):
            ws.cell(row, col).border = THIN
            ws.cell(row, col).alignment = Alignment(horizontal="center")
        ws.cell(row, 2).alignment = Alignment(horizontal="left")

    for col, width in {
        "A": 20,
        "B": 20,
        "C": 8,
        "D": 8,
        "E": 8,
        "F": 8,
        "G": 12,
        "H": 14,
        "I": 12,
    }.items():
        ws.column_dimensions[col].width = width

    ws["A18"] = (
        "Points = 3×Wins + Draws. Seeding: Points, then head-to-head, then Points For."
    )
    ws["A18"].font = Font(italic=True, size=10, color="666666")


def seed_team_formula(seed_n: int) -> str:
    return f"='Team Standings'!B{11 + seed_n}"


def box(ws, row: int, col: int, value=None, fill=None):
    cell = ws.cell(row, col, value)
    cell.border = THIN
    if fill:
        cell.fill = fill
    return cell


def rebuild_bracket(ws) -> None:
    unmerge_all(ws)
    clear_sheet_values(ws, max_row=60, max_col=20)

    ws["A1"] = "BDL Tune Up — 5-Team Single Elimination"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")

    ws["A2"] = (
        "Seeds from pool standings (Points, then H2H, then PF). "
        "Enter playoff scores on this sheet; winners advance."
    )
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:G2")

    ws["A4"] = "Play-In"
    ws["A4"].font = SECTION_FONT
    ws["A5"] = "Seed 4"
    ws["A6"] = "Seed 5"
    box(ws, 5, 2, seed_team_formula(4), PLAY_FILL)
    box(ws, 6, 2, seed_team_formula(5), PLAY_FILL)
    box(ws, 5, 3)
    box(ws, 6, 3)
    ws["C4"] = "Score"
    ws["C4"].font = Font(bold=True)
    ws["B8"] = "Winner →"
    ws["B8"].font = Font(bold=True)
    box(
        ws,
        8,
        3,
        '=IF(AND(ISNUMBER(C5),ISNUMBER(C6),C5<>C6),IF(C5>C6,B5,B6),"")',
        PLAY_FILL,
    )

    ws["E4"] = "Semifinals"
    ws["E4"].font = SECTION_FONT
    ws["E5"] = "Semi 1"
    ws["D6"] = "Seed 1"
    ws["D7"] = "Play-in W"
    box(ws, 6, 5, seed_team_formula(1), PLAY_FILL)
    box(ws, 7, 5, '=IF(C8="","",C8)', PLAY_FILL)
    box(ws, 6, 6)
    box(ws, 7, 6)
    ws["F5"] = "Score"
    ws["F5"].font = Font(bold=True)
    box(
        ws,
        8,
        5,
        '=IF(AND(ISNUMBER(F6),ISNUMBER(F7),F6<>F7),IF(F6>F7,E6,E7),"")',
        PLAY_FILL,
    )
    ws["D8"] = "Winner"

    ws["E10"] = "Semi 2"
    ws["D11"] = "Seed 2"
    ws["D12"] = "Seed 3"
    box(ws, 11, 5, seed_team_formula(2), PLAY_FILL)
    box(ws, 12, 5, seed_team_formula(3), PLAY_FILL)
    box(ws, 11, 6)
    box(ws, 12, 6)
    box(
        ws,
        13,
        5,
        '=IF(AND(ISNUMBER(F11),ISNUMBER(F12),F11<>F12),IF(F11>F12,E11,E12),"")',
        PLAY_FILL,
    )
    ws["D13"] = "Winner"

    ws["H4"] = "Championship"
    ws["H4"].font = SECTION_FONT
    ws["G6"] = "Semi 1 W"
    ws["G7"] = "Semi 2 W"
    box(ws, 6, 8, '=IF(E8="","",E8)', FINAL_FILL)
    box(ws, 7, 8, '=IF(E13="","",E13)', FINAL_FILL)
    box(ws, 6, 9)
    box(ws, 7, 9)
    ws["I5"] = "Score"
    ws["I5"].font = Font(bold=True)
    box(
        ws,
        9,
        8,
        '=IF(AND(ISNUMBER(I6),ISNUMBER(I7),I6<>I7),IF(I6>I7,H6,H7),"")',
        FINAL_FILL,
    )
    ws["G9"] = "Champion"
    ws["G9"].font = Font(bold=True)
    box(ws, 10, 8, '=IF(H9="","",IF(H9=H6,H7,H6))')
    ws["G10"] = "Runner-up"

    for col, width in {
        "A": 12,
        "B": 22,
        "C": 10,
        "D": 12,
        "E": 22,
        "F": 10,
        "G": 12,
        "H": 22,
        "I": 10,
    }.items():
        ws.column_dimensions[col].width = width


def link_schedule_playoffs_to_bracket(ws_schedule) -> None:
    ws_schedule["O5"] = "='Bracket'!B5"
    ws_schedule["P5"] = "='Bracket'!B6"
    ws_schedule["Q5"] = "='Bracket'!C5"
    ws_schedule["R5"] = "='Bracket'!C6"
    ws_schedule["S5"] = "='Bracket'!C8"

    ws_schedule["O9"] = "='Bracket'!E6"
    ws_schedule["P9"] = "='Bracket'!E7"
    ws_schedule["Q9"] = "='Bracket'!F6"
    ws_schedule["R9"] = "='Bracket'!F7"
    ws_schedule["S9"] = "='Bracket'!E8"

    ws_schedule["O13"] = "='Bracket'!E11"
    ws_schedule["P13"] = "='Bracket'!E12"
    ws_schedule["Q13"] = "='Bracket'!F11"
    ws_schedule["R13"] = "='Bracket'!F12"
    ws_schedule["S13"] = "='Bracket'!E13"

    ws_schedule["O17"] = "='Bracket'!H6"
    ws_schedule["P17"] = "='Bracket'!H7"
    ws_schedule["Q17"] = "='Bracket'!I6"
    ws_schedule["R17"] = "='Bracket'!I7"
    ws_schedule["S17"] = "='Bracket'!H9"

    ws_schedule["O20"] = "='Bracket'!H9"
    ws_schedule["P20"] = "='Bracket'!H10"


def rebuild_final_standings(ws) -> None:
    unmerge_all(ws)
    clear_sheet_values(ws, max_row=20, max_col=10)

    ws["A1"] = "Final Standings"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Place"
    ws["B2"] = "Team"
    ws["A2"].font = HEADER_FONT
    ws["B2"].font = HEADER_FONT
    ws["A2"].fill = HEADER_FILL
    ws["B2"].fill = HEADER_FILL

    ws["A3"] = 1
    ws["B3"] = "='Bracket'!H9"
    ws["A4"] = 2
    ws["B4"] = "='Bracket'!H10"
    ws["A5"] = 3
    ws["B5"] = (
        '=IF(\'Bracket\'!E8="","",'
        'IF(\'Bracket\'!E8=\'Bracket\'!E6,\'Bracket\'!E7,\'Bracket\'!E6))'
    )
    ws["A6"] = 4
    ws["B6"] = (
        '=IF(\'Bracket\'!E13="","",'
        'IF(\'Bracket\'!E13=\'Bracket\'!E11,\'Bracket\'!E12,\'Bracket\'!E11))'
    )
    ws["A7"] = 5
    ws["B7"] = (
        '=IF(\'Bracket\'!C8="","",'
        'IF(\'Bracket\'!C8=\'Bracket\'!B5,\'Bracket\'!B6,\'Bracket\'!B5))'
    )

    for r in range(3, 8):
        ws.cell(r, 1).border = THIN
        ws.cell(r, 2).border = THIN
        ws.cell(r, 1).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 24
    ws["A9"] = "Note: Places 3 and 4 are the semifinal losers (no consolation game)."
    ws["A9"].font = Font(italic=True, size=10, color="666666")


def clean_court_printouts(ws) -> int:
    cleared = 0
    for row in ws.iter_rows(min_row=1, max_row=200, max_col=40):
        for cell in row:
            val = cell.value
            if not isinstance(val, str) or not val.startswith("="):
                continue
            if "#REF!" in val:
                cell.value = None
                cleared += 1
                continue
            if re.search(r"'Team Schedule'![A-Z](2[1-9]|[3-9]\d|\d{3,})", val):
                cell.value = None
                cleared += 1
                continue
            if re.search(r"Team Schedule![A-Z](2[1-9]|[3-9]\d)", val):
                cell.value = None
                cleared += 1
    return cleared


def apply(workbook_path: Path) -> dict:
    wb = load_workbook(workbook_path)
    required = [
        "Team Rosters",
        "Team Schedule",
        "Team Standings",
        "Bracket",
        "Final Standings",
        "Court Assignment Printouts",
    ]
    for name in required:
        if name not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {name}")

    teams = read_teams_from_rosters(wb["Team Rosters"])
    matchups = discover_matchups(wb["Team Schedule"])
    scheduled = {m["team_a"] for m in matchups} | {m["team_b"] for m in matchups}
    if set(teams) != scheduled:
        raise ValueError(f"Roster/schedule team mismatch: {teams} vs {scheduled}")

    label_schedule_scores(wb["Team Schedule"])
    clear_schedule_junk(wb["Team Schedule"])
    wire_playoff_labels(wb["Team Schedule"])
    rebuild_standings(wb["Team Standings"], teams, matchups)
    rebuild_bracket(wb["Bracket"])
    link_schedule_playoffs_to_bracket(wb["Team Schedule"])
    rebuild_final_standings(wb["Final Standings"])
    cleared = clean_court_printouts(wb["Court Assignment Printouts"])

    wb.save(workbook_path)

    return {
        "teams": teams,
        "matchups": len(matchups),
        "game_map": {t: games_for_team(t, matchups) for t in teams},
        "printout_cleared": cleared,
        "path": str(workbook_path),
    }


def clear_rr_scores(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Team Schedule"]
    for start in RR_BLOCKS:
        for score_col in COURT_SCORE_COLS:
            ws.cell(start, score_col).value = None
            ws.cell(start + 1, score_col).value = None
    wb.save(workbook_path)


def smoke_test(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    teams = read_teams_from_rosters(wb["Team Rosters"])
    matchups = discover_matchups(wb["Team Schedule"])
    ws = wb["Team Schedule"]

    for m in matchups:
        ws[m["score_a"]] = None
        ws[m["score_b"]] = None

    def set_score(team_a, team_b, sa, sb):
        for m in matchups:
            if {m["team_a"], m["team_b"]} == {team_a, team_b}:
                if m["team_a"] == team_a:
                    ws[m["score_a"]] = sa
                    ws[m["score_b"]] = sb
                else:
                    ws[m["score_a"]] = sb
                    ws[m["score_b"]] = sa
                return
        raise ValueError(f"No matchup {team_a} vs {team_b}")

    set_score("High Octane", "Velocity Syndicate", 2, 2)
    set_score("Hyperdrive", "Carbon Fiber", 3, 1)
    set_score("Full Throttle", "Velocity Syndicate", 3, 0)
    set_score("High Octane", "Hyperdrive", 2, 1)
    set_score("Full Throttle", "Carbon Fiber", 3, 1)
    set_score("Velocity Syndicate", "Hyperdrive", 2, 2)
    set_score("Full Throttle", "Hyperdrive", 3, 0)
    set_score("Carbon Fiber", "High Octane", 1, 3)
    set_score("Full Throttle", "High Octane", 2, 1)
    set_score("Carbon Fiber", "Velocity Syndicate", 2, 3)

    stats = {
        t: {"W": 0, "L": 0, "D": 0, "PF": 0, "PA": 0, "H2H": {}} for t in teams
    }
    for m in matchups:
        a, b = m["team_a"], m["team_b"]
        sa = ws[m["score_a"]].value
        sb = ws[m["score_b"]].value
        assert sa is not None and sb is not None
        stats[a]["PF"] += sa
        stats[a]["PA"] += sb
        stats[b]["PF"] += sb
        stats[b]["PA"] += sa
        if sa > sb:
            stats[a]["W"] += 1
            stats[b]["L"] += 1
            stats[a]["H2H"][b] = 1
            stats[b]["H2H"][a] = -1
        elif sa < sb:
            stats[b]["W"] += 1
            stats[a]["L"] += 1
            stats[b]["H2H"][a] = 1
            stats[a]["H2H"][b] = -1
        else:
            stats[a]["D"] += 1
            stats[b]["D"] += 1
            stats[a]["H2H"][b] = 0
            stats[b]["H2H"][a] = 0

    for t, s in stats.items():
        s["Pts"] = s["W"] * 3 + s["D"]

    def sort_key(t: str):
        s = stats[t]
        bonus = 0.0
        for opp in teams:
            if opp == t:
                continue
            if stats[opp]["Pts"] == s["Pts"] and s["H2H"].get(opp, 0) == 1:
                bonus += 0.01
        return (s["Pts"] + bonus + s["PF"] / 10000, t)

    ranked = sorted(teams, key=sort_key, reverse=True)

    print("Smoke test expected standings (Points → H2H → PF):")
    for i, t in enumerate(ranked, 1):
        s = stats[t]
        print(
            f"  Seed {i}: {t:20s}  W{s['W']}-L{s['L']}-D{s['D']}  "
            f"Pts={s['Pts']}  PF={s['PF']} PA={s['PA']}"
        )

    ws_st = wb["Team Standings"]
    ft_row = None
    for row in range(3, 8):
        if ws_st.cell(row, 1).value == "Full Throttle":
            ft_row = row
            break
    assert ft_row is not None
    wins_f = ws_st.cell(ft_row, 2).value
    assert isinstance(wins_f, str) and wins_f.startswith("=")
    assert "Team Schedule" in wins_f
    assert stats["Full Throttle"]["W"] == 4
    assert stats["Full Throttle"]["Pts"] == 12
    assert ranked[0] == "Full Throttle"

    assert wb["Bracket"]["B5"].value == "='Team Standings'!B15"
    assert wb["Bracket"]["E6"].value == "='Team Standings'!B12"

    # Verify formulas encode expected W/L/D for Full Throttle via python recompute
    # matching the Excel IF structure for one team
    ft_games = games_for_team("Full Throttle", matchups)
    w = l = d = 0
    for own, opp in ft_games:
        sa = ws[own].value
        sb = ws[opp].value
        if sa > sb:
            w += 1
        elif sa < sb:
            l += 1
        else:
            d += 1
    assert (w, l, d) == (4, 0, 0)

    wb.save(workbook_path)
    print(f"Sample scores written to {workbook_path}")
    print("Open in Excel/Sheets to confirm calculated standings match above.")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--smoke-test", action="store_true")
    parser.add_argument("--clear-scores", action="store_true")
    args = parser.parse_args()

    path = args.workbook
    if not path.is_file():
        raise SystemExit(f"Workbook not found: {path}")

    result = apply(path)
    print(f"Updated {result['path']}")
    print(f"Teams: {', '.join(result['teams'])}")
    print(f"RR matchups: {result['matchups']}")
    print(f"Court printout formulas cleared: {result['printout_cleared']}")
    print("Game map (own score cell, opp score cell):")
    for team, games in result["game_map"].items():
        print(f"  {team}: {games}")

    if args.smoke_test:
        smoke_test(path)

    if args.clear_scores:
        clear_rr_scores(path)
        print("Cleared RR score cells.")


if __name__ == "__main__":
    main()
