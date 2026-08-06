#!/usr/bin/env python3
"""
Generate a 5-team, 2-court team-ref round-robin schedule.

Each team plays every other team exactly once (10 games total → 5 rounds).
Each round: 2 games on Court 1 + Court 2; the bye team refs both courts.
Balance targets per team:
  - 2 home games
  - 2 away games
  - 2 refs (one bye round covering both courts counts as 2)
"""

from __future__ import annotations

import argparse
from collections import Counter
from itertools import combinations
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
import sys

sys.path.insert(0, str(ROOT))

from league_schedule_format import TEAM_REF, write_format_to_teams_sheet  # noqa: E402

NUM_TEAMS = 5
NUM_ROUNDS = 5
GAMES_TOTAL = 10
HOME_PER_TEAM = 2
AWAY_PER_TEAM = 2
REFS_PER_TEAM = 2


def _is_home(i: int, j: int) -> bool:
    """Regular tournament orientation: i is home vs j iff (j - i) mod 5 ∈ {1, 2}."""
    return ((j - i) % NUM_TEAMS) in (1, 2)


def _circle_pairings(n: int = NUM_TEAMS) -> list[tuple[int, list[tuple[int, int]]]]:
    """Standard circle method for odd n: round k has bye when 2i ≡ k (mod n).

    Returns list of (bye_index, [(i, j), ...]) for k = 0 .. n-1.
    """
    out: list[tuple[int, list[tuple[int, int]]]] = []
    for k in range(n):
        # bye: unique i with 2i ≡ k (mod n); multiply by modular inverse of 2
        bye = (k * ((n + 1) // 2)) % n
        pairs: list[tuple[int, int]] = []
        seen: set[int] = {bye}
        for i in range(n):
            if i in seen:
                continue
            j = (k - i) % n
            if j in seen or j == i:
                continue
            pairs.append((i, j))
            seen.add(i)
            seen.add(j)
        out.append((bye, pairs))
    return out


def build_rounds(teams: list[str]) -> list[dict]:
    """Build 5 rounds with balanced home/away and bye-team double-ref."""
    if len(teams) != NUM_TEAMS:
        raise ValueError(f"Expected exactly {NUM_TEAMS} teams, got {len(teams)}")

    def ordered(i: int, j: int) -> tuple[str, str]:
        if _is_home(i, j):
            return teams[i], teams[j]
        return teams[j], teams[i]

    rounds: list[dict] = []
    for round_num, (bye_idx, pairs) in enumerate(_circle_pairings(), start=1):
        if len(pairs) != 2:
            raise RuntimeError(f"round {round_num}: expected 2 pairs, got {pairs}")
        (a, b), (c, d) = pairs
        home1, away1 = ordered(a, b)
        home2, away2 = ordered(c, d)
        ref = teams[bye_idx]
        rounds.append(
            {
                "round": round_num,
                "court1": {"home": home1, "away": away1, "ref": ref},
                "court2": {"home": home2, "away": away2, "ref": ref},
                "bye": ref,
            }
        )
    return rounds


def validate_rounds(teams: list[str], rounds: list[dict]) -> list[str]:
    errors: list[str] = []
    if len(rounds) != NUM_ROUNDS:
        errors.append(f"expected {NUM_ROUNDS} rounds, got {len(rounds)}")

    home = Counter()
    away = Counter()
    refs = Counter()
    played_pairs: set[tuple[str, str]] = set()

    for rnd in rounds:
        c1, c2 = rnd["court1"], rnd["court2"]
        playing = {c1["home"], c1["away"], c2["home"], c2["away"]}
        if len(playing) != 4:
            errors.append(f"round {rnd['round']}: expected 4 distinct playing teams, got {playing}")
        bye = rnd["bye"]
        if bye in playing:
            errors.append(f"round {rnd['round']}: bye team {bye} is also playing")
        if set(teams) - playing != {bye}:
            errors.append(f"round {rnd['round']}: bye mismatch")

        for court in (c1, c2):
            home[court["home"]] += 1
            away[court["away"]] += 1
            refs[court["ref"]] += 1
            pair = tuple(sorted((court["home"], court["away"])))
            if pair in played_pairs:
                errors.append(f"duplicate matchup {pair}")
            played_pairs.add(pair)
            if court["ref"] != bye:
                errors.append(
                    f"round {rnd['round']}: court ref {court['ref']} != bye {bye}"
                )

    expected_pairs = {tuple(sorted(p)) for p in combinations(teams, 2)}
    if played_pairs != expected_pairs:
        missing = expected_pairs - played_pairs
        extra = played_pairs - expected_pairs
        if missing:
            errors.append(f"missing matchups: {sorted(missing)}")
        if extra:
            errors.append(f"unexpected matchups: {sorted(extra)}")

    for team in teams:
        if home[team] != HOME_PER_TEAM:
            errors.append(f"{team}: home={home[team]} (want {HOME_PER_TEAM})")
        if away[team] != AWAY_PER_TEAM:
            errors.append(f"{team}: away={away[team]} (want {AWAY_PER_TEAM})")
        if refs[team] != REFS_PER_TEAM:
            errors.append(f"{team}: refs={refs[team]} (want {REFS_PER_TEAM})")

    return errors


def _team_pattern(team: str) -> str:
    if " " in team or team.endswith("*"):
        return f'"{team}*"' if not team.endswith("*") else f'"{team}"'
    return f'"{team}"'


def write_workbook(
    path: str | Path,
    teams: list[str],
    rounds: list[dict] | None = None,
    league_name: str = "Five Team Round Robin",
    week_name: str = "Week 1",
) -> None:
    path = Path(path)
    rounds = rounds or build_rounds(teams)
    errors = validate_rounds(teams, rounds)
    if errors:
        raise ValueError("Invalid schedule:\n  - " + "\n  - ".join(errors))

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws_teams = wb.create_sheet("Teams")
    ws_teams.cell(1, 1).value = "Team Names"
    for col, team in enumerate(teams, start=3):
        ws_teams.cell(1, col).value = team
    write_format_to_teams_sheet(ws_teams, TEAM_REF)
    ws_teams.cell(3, 1).value = "League"
    ws_teams.cell(3, 2).value = league_name

    ws_gen = wb.create_sheet("Schedule Generator")
    for col, team in enumerate(teams, start=2):
        ws_gen.cell(2, col).value = team
    for row_idx, team in enumerate(teams, start=3):
        ws_gen.cell(row_idx, 1).value = team
        for col_idx, opponent in enumerate(teams, start=2):
            if team == opponent:
                ws_gen.cell(row_idx, col_idx).value = "-"
            else:
                ws_gen.cell(row_idx, col_idx).value = 1

    ws_standings = wb.create_sheet("League Standings")
    ws_standings.cell(1, 1).value = "LEAGUE STANDINGS"
    ws_standings.cell(2, 1).value = "Team Name"
    ws_standings.cell(2, 2).value = "Points For"
    ws_standings.cell(2, 3).value = "Points Against"
    ws_standings.cell(2, 4).value = "Point Differential"
    ws_standings.cell(11, 1).value = "Week #"
    ws_standings.cell(11, 2).value = 0
    ws_standings.cell(16, 1).value = "Teams"
    ws_standings.cell(16, 2).value = "Wins"
    ws_standings.cell(16, 5).value = "Losses"

    ws = wb.create_sheet(week_name)
    ws.cell(1, 2).value = "Court 1"
    ws.cell(1, 7).value = "Court 2"

    row = 2
    for rnd in rounds:
        c1, c2 = rnd["court1"], rnd["court2"]
        ws.cell(row, 1).value = f"Game {rnd['round']:02d}"
        ws.cell(row, 2).value = c1["home"]
        ws.cell(row, 4).value = c1["away"]
        ws.cell(row, 7).value = c2["home"]
        ws.cell(row, 9).value = c2["away"]
        row += 1
        ws.cell(row, 2).value = f"Refs: {c1['ref']}"
        ws.cell(row, 7).value = f"Refs: {c2['ref']}"
        row += 1

    # Balance summary
    row += 1
    ws.cell(row, 1).value = "Balance check (home / away / refs)"
    row += 1
    ws.cell(row, 1).value = "Team"
    ws.cell(row, 2).value = "Home"
    ws.cell(row, 3).value = "Away"
    ws.cell(row, 4).value = "Refs"
    row += 1

    home = Counter()
    away = Counter()
    refs = Counter()
    for rnd in rounds:
        for court in (rnd["court1"], rnd["court2"]):
            home[court["home"]] += 1
            away[court["away"]] += 1
            refs[court["ref"]] += 1
    for team in teams:
        ws.cell(row, 1).value = team
        ws.cell(row, 2).value = home[team]
        ws.cell(row, 3).value = away[team]
        ws.cell(row, 4).value = refs[team]
        row += 1

    # Dual-court win/loss formulas (BYOT-style)
    row += 1
    wl_header = row
    ws.cell(wl_header, 1).value = "Team Wins/Losses This Week"
    ws.cell(wl_header + 1, 1).value = "Team Name"
    ws.cell(wl_header + 1, 2).value = "Wins"
    ws.cell(wl_header + 1, 3).value = "Losses"
    first_data = wl_header + 2
    for i, team in enumerate(teams):
        r = first_data + i
        pat = _team_pattern(team)
        ws.cell(r, 1).value = team
        ws.cell(r, 2).value = (
            f"=SUMIFS(C:C,B:B,{pat})+SUMIFS(E:E,D:D,{pat})+"
            f"SUMIFS(H:H,G:G,{pat})+SUMIFS(J:J,I:I,{pat})"
        )
        ws.cell(r, 3).value = (
            f"=SUMIFS(E:E,B:B,{pat})+SUMIFS(C:C,D:D,{pat})+"
            f"SUMIFS(J:J,G:G,{pat})+SUMIFS(H:H,I:I,{pat})"
        )

    for i, team in enumerate(teams):
        r = 17 + i
        data_row = first_data + i
        ws_standings.cell(r, 1).value = team
        ws_standings.cell(r, 2).value = f"='{week_name}'!B{data_row}"
        ws_standings.cell(r, 3).value = f"=ROUND(B{r}+{r}/10000,8)"
        ws_standings.cell(r, 5).value = f"='{week_name}'!C{data_row}"

    c_lo, c_hi = 17, 16 + len(teams)
    c_rng = f"C{c_lo}:C{c_hi}"
    for rank in range(1, len(teams) + 1):
        display_row = 2 + rank
        large_k = f"ROUND(LARGE({c_rng},{rank}),8)"
        ws_standings.cell(display_row, 1).value = (
            f"=INDEX(A{c_lo}:A{c_hi},MATCH({large_k},{c_rng},0))"
        )
        ws_standings.cell(display_row, 2).value = (
            f"=INDEX(B{c_lo}:B{c_hi},MATCH({large_k},{c_rng},0))"
        )
        ws_standings.cell(display_row, 3).value = (
            f"=INDEX(E{c_lo}:E{c_hi},MATCH({large_k},{c_rng},0))"
        )
        ws_standings.cell(display_row, 4).value = f"=B{display_row}-C{display_row}"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def format_schedule_text(teams: list[str], rounds: list[dict]) -> str:
    lines = [
        "5-Team Round Robin (each plays each other once)",
        "2 courts — bye team refs both courts (counts as 2 refs)",
        "",
    ]
    for rnd in rounds:
        c1, c2 = rnd["court1"], rnd["court2"]
        lines.append(f"Round {rnd['round']}  (Refs: {rnd['bye']} — both courts)")
        lines.append(f"  Court 1: {c1['home']} (H) vs {c1['away']} (A)")
        lines.append(f"  Court 2: {c2['home']} (H) vs {c2['away']} (A)")
        lines.append("")

    home = Counter()
    away = Counter()
    refs = Counter()
    for rnd in rounds:
        for court in (rnd["court1"], rnd["court2"]):
            home[court["home"]] += 1
            away[court["away"]] += 1
            refs[court["ref"]] += 1

    lines.append("Per-team balance:")
    lines.append(f"{'Team':<22} {'Home':>4} {'Away':>4} {'Refs':>4}")
    for team in teams:
        lines.append(f"{team:<22} {home[team]:>4} {away[team]:>4} {refs[team]:>4}")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate 5-team 2-court team-ref round-robin schedule"
    )
    parser.add_argument(
        "teams",
        nargs="*",
        help="Exactly five team names (default: Team 1 … Team 5)",
    )
    parser.add_argument("--output", "-o", type=str, help="Write .xlsx workbook")
    parser.add_argument("--league-name", default="Five Team Round Robin")
    parser.add_argument("--print", action="store_true", dest="do_print", help="Print schedule")
    parser.add_argument("--validate", action="store_true", help="Validate only")
    args = parser.parse_args()

    teams = list(args.teams) if args.teams else [f"Team {i}" for i in range(1, 6)]
    if len(teams) != NUM_TEAMS:
        print(f"ERROR: need exactly {NUM_TEAMS} teams, got {len(teams)}")
        return 1

    rounds = build_rounds(teams)
    errors = validate_rounds(teams, rounds)
    if errors:
        print("Validation FAILED:")
        for err in errors:
            print(f"  - {err}")
        return 1

    print("Validation passed.")
    if args.do_print or not args.output:
        print()
        print(format_schedule_text(teams, rounds))

    if args.output:
        out = Path(args.output)
        if not out.is_absolute():
            out = ROOT / out
        write_workbook(out, teams, rounds, league_name=args.league_name)
        print(f"Wrote: {out}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
