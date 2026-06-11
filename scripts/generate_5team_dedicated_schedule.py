#!/usr/bin/env python3
"""
Generate a 6-week, 5-team dedicated-ref league schedule.

Per night: 25 games, 10 per team (5 home / 5 away), each pairing 2 or 3 games.
Per season: 15 games per pairing across 6 weeks.
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter
import random
from itertools import combinations, permutations
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from league_schedule_format import DEDICATED_REF, write_format_to_teams_sheet

GAMES_PER_WEEK = 25
GAMES_PER_TEAM_PER_WEEK = 10
HOME_PER_TEAM_PER_WEEK = 5
SEASON_GAMES_PER_PAIRING = 15
NUM_WEEKS_DEFAULT = 6


def pair_key(a: int, b: int) -> tuple[int, int]:
    return (a, b) if a < b else (b, a)


def all_pairings(n: int = 5) -> list[tuple[int, int]]:
    return [pair_key(i, j) for i in range(n) for j in range(i + 1, n)]


def cycle_edges(perm: tuple[int, ...]) -> frozenset[tuple[int, int]]:
    edges = set()
    for k in range(len(perm)):
        edges.add(pair_key(perm[k], perm[(k + 1) % len(perm)]))
    return frozenset(edges)


def find_heavy_sets(n: int = 5, num_weeks: int = 6) -> list[frozenset[tuple[int, int]]]:
    """Pick num_weeks distinct 5-cycles so each pairing is heavy exactly num_weeks/2 times."""
    pairings = all_pairings(n)
    target = num_weeks // 2  # each pairing is heavy on 3 of 6 weeks

    cycles: list[frozenset[tuple[int, int]]] = []
    seen: set[frozenset[tuple[int, int]]] = set()
    for perm in permutations(range(n)):
        edges = cycle_edges(perm)
        if len(edges) == 5 and edges not in seen:
            seen.add(edges)
            cycles.append(edges)

    pair_counts: Counter[tuple[int, int]] = Counter()
    chosen: list[frozenset[tuple[int, int]]] = []

    def backtrack(week: int) -> bool:
        if week == num_weeks:
            return all(pair_counts[p] == target for p in pairings)

        for cycle in cycles:
            if any(pair_counts[edge] + 1 > target for edge in cycle):
                continue

            chosen.append(cycle)
            for edge in cycle:
                pair_counts[edge] += 1

            if backtrack(week + 1):
                return True

            chosen.pop()
            for edge in cycle:
                pair_counts[edge] -= 1

        return False

    if not backtrack(0):
        raise RuntimeError("Could not construct heavy-set rotation for 5-team dedicated schedule")

    return chosen


HEAVY_SETS = find_heavy_sets()


def assign_home_away(
    teams: list[str],
    heavy_set: frozenset[tuple[int, int]],
    week_index: int,
) -> list[tuple[str, str]]:
    """Build 25 (home, away) games for one week with balanced home/away counts."""
    idx = {i: teams[i] for i in range(len(teams))}
    games: list[tuple[str, str]] = []
    home_counts = Counter()
    away_counts = Counter()

    for pairing in all_pairings(len(teams)):
        count = 3 if pairing in heavy_set else 2
        a, b = pairing
        team_a, team_b = idx[a], idx[b]

        pairing_games: list[tuple[str, str]] = []
        start_home_a = (week_index + a + b) % 2 == 0
        for game_i in range(count):
            if count == 3:
                if game_i == 0:
                    home_is_a = start_home_a
                elif game_i == 1:
                    home_is_a = not start_home_a
                else:
                    home_is_a = start_home_a
            else:
                home_is_a = game_i == 0 if start_home_a else game_i == 1

            if home_is_a:
                pairing_games.append((team_a, team_b))
            else:
                pairing_games.append((team_b, team_a))

        for home, away in pairing_games:
            home_counts[home] += 1
            away_counts[away] += 1
            games.append((home, away))

    # Fix home/away imbalance by swapping within same pairings
    for _ in range(20):
        fixed = False
        for team in teams:
            if home_counts[team] > HOME_PER_TEAM_PER_WEEK:
                for i, (home, away) in enumerate(games):
                    if home == team:
                        games[i] = (away, home)
                        home_counts[home] -= 1
                        away_counts[away] -= 1
                        home_counts[away] += 1
                        away_counts[home] += 1
                        fixed = True
                        break
            elif away_counts[team] > HOME_PER_TEAM_PER_WEEK:
                for i, (home, away) in enumerate(games):
                    if away == team:
                        games[i] = (away, home)
                        home_counts[home] -= 1
                        away_counts[away] -= 1
                        home_counts[away] += 1
                        away_counts[home] += 1
                        fixed = True
                        break
        if not fixed:
            break

    return games


def team_idle_gaps(ordered: list[tuple[str, str]], team: str) -> list[int]:
    """Idle slots before, between, and after a team's games in the ordered list."""
    positions = [i for i, (home, away) in enumerate(ordered) if team in (home, away)]
    if not positions:
        return []
    n = len(ordered)
    gaps = [positions[0]]
    for i in range(1, len(positions)):
        gaps.append(positions[i] - positions[i - 1] - 1)
    gaps.append(n - 1 - positions[-1])
    return gaps


def schedule_evenness_score(ordered: list[tuple[str, str]], teams: list[str]) -> float:
    """Lower is better: penalize long breaks and uneven spacing across teams."""
    max_gap = 0
    variance_penalty = 0.0
    for team in teams:
        gaps = team_idle_gaps(ordered, team)
        if not gaps:
            continue
        max_gap = max(max_gap, max(gaps))
        mean_gap = sum(gaps) / len(gaps)
        variance_penalty += sum((gap - mean_gap) ** 2 for gap in gaps)

    adj_penalty = count_adjacent_duplicate_pairings_from_tuples(teams, ordered) * 250
    return max_gap * 1000 + variance_penalty + adj_penalty


def count_adjacent_duplicate_pairings_from_tuples(
    teams: list[str], games: list[tuple[str, str]]
) -> int:
    count = 0
    for i in range(1, len(games)):
        prev = pair_key(teams.index(games[i - 1][0]), teams.index(games[i - 1][1]))
        cur = pair_key(teams.index(games[i][0]), teams.index(games[i][1]))
        if prev == cur:
            count += 1
    return count


def order_games(games: list[tuple[str, str]], teams: list[str]) -> list[tuple[str, str]]:
    """Build an order that keeps idle time between games even for every team."""
    remaining = games[:]
    ordered: list[tuple[str, str]] = []
    last_pos: dict[str, int] = {t: -1 for t in teams}
    games_remaining_per_team = Counter()
    for home, away in games:
        games_remaining_per_team[home] += 1
        games_remaining_per_team[away] += 1

    total_slots = len(games)

    while remaining:
        best_idx = -1
        best_score = float("-inf")
        current_len = len(ordered)
        slots_left = total_slots - current_len

        wait_times = {
            t: current_len if last_pos[t] == -1 else current_len - last_pos[t] for t in teams
        }

        for idx, (home, away) in enumerate(remaining):
            w1, w2 = wait_times[home], wait_times[away]
            urgency = w1 + w2

            # Penalize teams that still have many games left but have not played recently.
            for team, wait in ((home, w1), (away, w2)):
                games_left = games_remaining_per_team[team] - 1
                if games_left > 0 and slots_left > 0:
                    ideal_gap = slots_left / games_left
                    if wait > ideal_gap:
                        urgency += (wait - ideal_gap) * 400

            adj_penalty = 0
            if ordered:
                prev_home, prev_away = ordered[-1]
                prev_pair = pair_key(
                    teams.index(prev_home),
                    teams.index(prev_away),
                )
                cur_pair = pair_key(teams.index(home), teams.index(away))
                if prev_pair == cur_pair:
                    adj_penalty = 5000

            score = urgency - adj_penalty

            if score > best_score:
                best_score = score
                best_idx = idx

        pick = remaining.pop(best_idx)
        ordered.append(pick)
        pos = len(ordered) - 1
        last_pos[pick[0]] = pos
        last_pos[pick[1]] = pos
        games_remaining_per_team[pick[0]] -= 1
        games_remaining_per_team[pick[1]] -= 1

    ordered = _fix_adjacent_same_pairing(ordered, teams)
    return optimize_game_order_evenness(ordered, teams)


def optimize_game_order_evenness(
    ordered: list[tuple[str, str]], teams: list[str]
) -> list[tuple[str, str]]:
    """Hill-climb and random-swap search to even out breaks between games."""
    best = list(ordered)
    best_score = schedule_evenness_score(best, teams)
    n = len(best)

    improved = True
    while improved:
        improved = False
        for i in range(n):
            for j in range(i + 1, n):
                candidate = best[:]
                candidate[i], candidate[j] = candidate[j], candidate[i]
                score = schedule_evenness_score(candidate, teams)
                if score < best_score:
                    best = candidate
                    best_score = score
                    improved = True

    rng = random.Random(42)
    for _ in range(30000):
        i, j = rng.sample(range(n), 2)
        candidate = best[:]
        candidate[i], candidate[j] = candidate[j], candidate[i]
        score = schedule_evenness_score(candidate, teams)
        if score < best_score:
            best = candidate
            best_score = score

    return _fix_adjacent_same_pairing(best, teams)


def _fix_adjacent_same_pairing(
    games: list[tuple[str, str]], teams: list[str]
) -> list[tuple[str, str]]:
    """Swap games to avoid back-to-back identical matchups when possible."""
    result = games[:]

    def same_pair(a: tuple[str, str], b: tuple[str, str]) -> bool:
        return pair_key(teams.index(a[0]), teams.index(a[1])) == pair_key(
            teams.index(b[0]), teams.index(b[1])
        )

    for i in range(1, len(result)):
        if not same_pair(result[i - 1], result[i]):
            continue
        swapped = False
        for j in range(i - 2, -1, -1):
            if same_pair(result[j], result[i]):
                continue
            if j > 0 and same_pair(result[j - 1], result[i]):
                continue
            if i + 1 < len(result) and same_pair(result[i], result[i + 1]):
                continue
            result[j], result[i] = result[i], result[j]
            swapped = True
            break
        if not swapped and i + 1 < len(result):
            result[i], result[i + 1] = result[i + 1], result[i]

    return result


def build_week_games(teams: list[str], week_index: int) -> list[dict[str, str]]:
    if len(teams) != 5:
        raise ValueError("5-team dedicated schedule requires exactly 5 teams")

    heavy = HEAVY_SETS[week_index % len(HEAVY_SETS)]
    raw = assign_home_away(teams, heavy, week_index)
    ordered = order_games(raw, teams)

    return [
        {"team1": home, "team2": away, "game": i + 1}
        for i, (home, away) in enumerate(ordered)
    ]


def build_season(teams: list[str], num_weeks: int = NUM_WEEKS_DEFAULT) -> list[list[dict[str, str]]]:
    return [build_week_games(teams, w) for w in range(num_weeks)]


def validate_week(teams: list[str], games: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    if len(games) != GAMES_PER_WEEK:
        errors.append(f"expected {GAMES_PER_WEEK} games, got {len(games)}")

    team_games = Counter()
    home_games = Counter()
    away_games = Counter()
    pairing_counts: Counter[tuple[str, str]] = Counter()

    for game in games:
        home, away = game["team1"], game["team2"]
        team_games[home] += 1
        team_games[away] += 1
        home_games[home] += 1
        away_games[away] += 1
        pairing_counts[pair_key(teams.index(home), teams.index(away))] += 1

    for team in teams:
        if team_games[team] != GAMES_PER_TEAM_PER_WEEK:
            errors.append(f"{team}: expected {GAMES_PER_TEAM_PER_WEEK} games, got {team_games[team]}")
        if home_games[team] != HOME_PER_TEAM_PER_WEEK:
            errors.append(f"{team}: expected {HOME_PER_TEAM_PER_WEEK} home games, got {home_games[team]}")
        if away_games[team] != HOME_PER_TEAM_PER_WEEK:
            errors.append(f"{team}: expected {HOME_PER_TEAM_PER_WEEK} away games, got {away_games[team]}")

    for pairing in combinations(range(len(teams)), 2):
        count = pairing_counts[pairing]
        if count not in (2, 3):
            names = (teams[pairing[0]], teams[pairing[1]])
            errors.append(f"pairing {names}: expected 2 or 3 games, got {count}")

    return errors


def count_adjacent_duplicate_pairings(teams: list[str], games: list[dict[str, str]]) -> int:
    tuples = [(g["team1"], g["team2"]) for g in games]
    return count_adjacent_duplicate_pairings_from_tuples(teams, tuples)


def validate_season(teams: list[str], weeks: list[list[dict[str, str]]]) -> list[str]:
    errors: list[str] = []
    season_pairings: Counter[tuple[int, int]] = Counter()

    for week_idx, week in enumerate(weeks):
        errors.extend(f"week {week_idx + 1}: {e}" for e in validate_week(teams, week))
        for game in week:
            pairing = pair_key(teams.index(game["team1"]), teams.index(game["team2"]))
            season_pairings[pairing] += 1

    for pairing in all_pairings(len(teams)):
        count = season_pairings[pairing]
        if count != SEASON_GAMES_PER_PAIRING:
            errors.append(
                f"season pairing {pairing}: expected {SEASON_GAMES_PER_PAIRING}, got {count}"
            )

    return errors


def week_sheet_names(num_weeks: int) -> list[str]:
    return [f"Week {week_num}" for week_num in range(1, num_weeks + 1)]


def write_workbook(
    path: str | Path,
    teams: list[str],
    weeks: list[list[dict[str, str]]],
    league_name: str = "Five Team Dedicated Ref League",
) -> None:
    path = Path(path)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws_teams = wb.create_sheet("Teams")
    ws_teams.cell(1, 1).value = "Team Names"
    for col, team in enumerate(teams, start=3):
        ws_teams.cell(1, col).value = team
    write_format_to_teams_sheet(ws_teams, DEDICATED_REF)

    ws_gen = wb.create_sheet("Schedule Generator")
    ws_gen.cell(2, 1).value = ""
    for col, team in enumerate(teams, start=2):
        ws_gen.cell(2, col).value = team
    for row_idx, team in enumerate(teams, start=3):
        ws_gen.cell(row_idx, 1).value = team
        for col_idx, opponent in enumerate(teams, start=2):
            if team == opponent:
                ws_gen.cell(row_idx, col_idx).value = "-"
            else:
                ws_gen.cell(row_idx, col_idx).value = SEASON_GAMES_PER_PAIRING

    ws_standings = wb.create_sheet("League Standings")
    ws_standings.cell(1, 1).value = "LEAGUE STANDINGS"
    ws_standings.cell(2, 1).value = "Team Name"
    ws_standings.cell(2, 2).value = "Points For"
    ws_standings.cell(2, 3).value = "Points Against"
    ws_standings.cell(2, 4).value = "Point Differential"
    ws_standings.cell(11, 1).value = "Week #"
    ws_standings.cell(11, 2).value = 0
    ws_standings.cell(16, 1).value = "Teams"
    ws_standings.cell(16, 2).value = "Wins"
    ws_standings.cell(16, 5).value = "Losses"

    week_names = week_sheet_names(len(weeks))
    for week_name, week_games in zip(week_names, weeks):
        ws = wb.create_sheet(week_name)
        ws.cell(1, 2).value = "Court 1"
        row = 2
        for game in week_games:
            ws.cell(row, 1).value = f"Game {game['game']:02d}"
            ws.cell(row, 2).value = game["team1"]
            ws.cell(row, 4).value = game["team2"]
            row += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def run_validation(teams: list[str] | None = None, num_weeks: int = NUM_WEEKS_DEFAULT) -> int:
    teams = teams or [f"Team {i}" for i in range(1, 6)]
    weeks = build_season(teams, num_weeks)
    errors = validate_season(teams, weeks)

    print(f"Validated {num_weeks}-week season for {', '.join(teams)}")
    print(f"Heavy-set weeks: {len(HEAVY_SETS)} rotations")
    if errors:
        print(f"FAILED with {len(errors)} error(s):")
        for err in errors:
            print(f"  - {err}")
        return 1

    print("All checks passed.")
    for week_idx, week in enumerate(weeks, start=1):
        adj = count_adjacent_duplicate_pairings(teams, week)
        adj_note = f", {adj} adjacent duplicate pairing(s)" if adj else ""
        print(f"  Week {week_idx}: {len(week)} games{adj_note}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate 5-team dedicated-ref league schedule")
    parser.add_argument("--validate", action="store_true", help="Run validation only")
    parser.add_argument(
        "--output",
        type=str,
        help="Write workbook to this path (implies --teams if omitted)",
    )
    parser.add_argument(
        "--teams",
        nargs=5,
        metavar="TEAM",
        help="Exactly five team names",
    )
    parser.add_argument("--weeks", type=int, default=NUM_WEEKS_DEFAULT)
    args = parser.parse_args()

    teams = list(args.teams) if args.teams else [f"Team {i}" for i in range(1, 6)]

    if args.validate:
        return run_validation(teams, args.weeks)

    if args.output:
        weeks = build_season(teams, args.weeks)
        errors = validate_season(teams, weeks)
        if errors:
            print("Validation failed; not writing workbook:")
            for err in errors:
                print(f"  - {err}")
            return 1
        write_workbook(args.output, teams, weeks)
        print(f"Wrote {args.output}")
        return 0

    return run_validation(teams, args.weeks)


if __name__ == "__main__":
    raise SystemExit(main())
