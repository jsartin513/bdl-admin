#!/usr/bin/env python3
"""
Generate a 6-week, 8-team / 2-court team-ref league schedule.

Per night: 40 games over 20 rounds (2 courts), 10 play / 5 ref / 5 off per team,
5 home / 5 away, each pairing 1 or 2 games (12 heavy pairings, 16 light).

Per season: each pairing plays 8 or 9 games (12 pairings × 8, 16 × 9).
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from itertools import combinations
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from league_schedule_format import TEAM_REF, write_format_to_teams_sheet

# Match Six Team League (Drive) dual-court team-ref styling.
_THIN = Border(
    left=Side(style="thin", color="FF000000"),
    right=Side(style="thin", color="FF000000"),
    top=Side(style="thin", color="FF000000"),
    bottom=Side(style="thin", color="FF000000"),
)
_FILL_WHITE = PatternFill("solid", fgColor="FFFFFFFF")
_FILL_REF = PatternFill("solid", fgColor="FFFEF8E3")
_FILL_GAP = PatternFill("solid", fgColor="FFD9D9D9")
_FILL_STANDINGS_HDR = PatternFill("solid", fgColor="FFD9D9D9")
_FONT_GAME = Font(name="Arial", size=11)
_FONT_TEAM = Font(name="Arial", size=14, bold=True)
_FONT_REF = Font(name="Arial", size=11)
_FONT_STD = Font(name="Arial", size=11)
_ALIGN_GAME = Alignment(horizontal="center", vertical="top")
_ALIGN_SCORE = Alignment(horizontal="right", vertical="bottom")
_ALIGN_STANDINGS_HDR = Alignment(horizontal="center", vertical="top")

# Six Team League week-sheet column widths (A–K); L widened for Off labels.
_WEEK_COL_WIDTHS = {
    "A": 7.88,
    "B": 17.63,
    "C": 4.63,
    "D": 17.63,
    "E": 5.25,
    "F": 12.63,
    "G": 16.38,
    "H": 5.75,
    "I": 17.63,
    "J": 5.88,
    "K": 13.0,
    "L": 28.0,
}


GAMES_PER_WEEK = 40
ROUNDS_PER_WEEK = 20
GAMES_PER_TEAM_PER_WEEK = 10
HOME_PER_TEAM_PER_WEEK = 5
REFS_PER_TEAM_PER_WEEK = 5
OFF_PER_TEAM_PER_WEEK = 5
NUM_TEAMS = 8
NUM_WEEKS_DEFAULT = 6
SEASON_GAMES_PER_PAIRING = (8, 9)  # evenest split

# Precomputed 3-regular heavy sets: each of 28 pairings is heavy on 2 or 3 weeks
# (16 pairings heavy 3× → 9 season games; 12 heavy 2× → 8 season games).
HEAVY_SETS: list[frozenset[tuple[int, int]]] = [
    frozenset(
        [
            (0, 1),
            (0, 2),
            (0, 3),
            (1, 2),
            (1, 3),
            (2, 3),
            (4, 5),
            (4, 6),
            (4, 7),
            (5, 6),
            (5, 7),
            (6, 7),
        ]
    ),
    frozenset(
        [
            (0, 4),
            (0, 5),
            (0, 6),
            (1, 4),
            (1, 5),
            (1, 7),
            (2, 4),
            (2, 6),
            (2, 7),
            (3, 5),
            (3, 6),
            (3, 7),
        ]
    ),
    frozenset(
        [
            (0, 1),
            (0, 2),
            (0, 7),
            (1, 2),
            (1, 6),
            (2, 5),
            (3, 4),
            (3, 5),
            (3, 6),
            (4, 5),
            (4, 7),
            (6, 7),
        ]
    ),
    frozenset(
        [
            (0, 3),
            (0, 4),
            (0, 5),
            (1, 3),
            (1, 4),
            (1, 6),
            (2, 4),
            (2, 6),
            (2, 7),
            (3, 7),
            (5, 6),
            (5, 7),
        ]
    ),
    frozenset(
        [
            (0, 1),
            (0, 6),
            (0, 7),
            (1, 5),
            (1, 7),
            (2, 3),
            (2, 4),
            (2, 5),
            (3, 4),
            (3, 5),
            (4, 6),
            (6, 7),
        ]
    ),
    frozenset(
        [
            (0, 2),
            (0, 3),
            (0, 4),
            (1, 2),
            (1, 3),
            (1, 5),
            (2, 6),
            (3, 7),
            (4, 6),
            (4, 7),
            (5, 6),
            (5, 7),
        ]
    ),
]


def pair_key(a: int, b: int) -> tuple[int, int]:
    return (a, b) if a < b else (b, a)


def all_pairings(n: int = NUM_TEAMS) -> list[tuple[int, int]]:
    return [pair_key(i, j) for i in range(n) for j in range(i + 1, n)]


def assign_home_away(
    teams: list[str],
    heavy_set: frozenset[tuple[int, int]],
    week_index: int,
) -> list[tuple[str, str]]:
    """Build 40 (home, away) games with 5 home / 5 away per team."""
    idx = {i: teams[i] for i in range(len(teams))}
    games: list[tuple[str, str]] = []
    home_counts: Counter[str] = Counter()
    away_counts: Counter[str] = Counter()

    for pairing in all_pairings(len(teams)):
        count = 2 if pairing in heavy_set else 1
        a, b = pairing
        team_a, team_b = idx[a], idx[b]
        start_home_a = (week_index + a + b) % 2 == 0

        for game_i in range(count):
            if count == 2:
                home_is_a = start_home_a if game_i == 0 else not start_home_a
            else:
                home_is_a = start_home_a

            if home_is_a:
                games.append((team_a, team_b))
                home_counts[team_a] += 1
                away_counts[team_b] += 1
            else:
                games.append((team_b, team_a))
                home_counts[team_b] += 1
                away_counts[team_a] += 1

    for _ in range(40):
        fixed = False
        for team in teams:
            if home_counts[team] > HOME_PER_TEAM_PER_WEEK:
                for i, (home, away) in enumerate(games):
                    if home == team:
                        games[i] = (away, home)
                        home_counts[home] -= 1
                        away_counts[away] -= 1
                        home_counts[away] += 1
                        away_counts[home] += 1
                        fixed = True
                        break
            elif away_counts[team] > HOME_PER_TEAM_PER_WEEK:
                for i, (home, away) in enumerate(games):
                    if away == team:
                        games[i] = (away, home)
                        home_counts[home] -= 1
                        away_counts[away] -= 1
                        home_counts[away] += 1
                        away_counts[home] += 1
                        fixed = True
                        break
        if not fixed:
            break

    return games


def pack_into_rounds(
    games: list[tuple[str, str]], teams: list[str]
) -> list[tuple[tuple[str, str], tuple[str, str]]]:
    """Pack 40 oriented games into 20 rounds of two vertex-disjoint matchups."""
    remaining: Counter[tuple[int, int]] = Counter()
    oriented: dict[tuple[int, int], list[tuple[str, str]]] = {}
    index = {t: i for i, t in enumerate(teams)}

    for home, away in games:
        edge = pair_key(index[home], index[away])
        remaining[edge] += 1
        oriented.setdefault(edge, []).append((home, away))

    rounds: list[tuple[tuple[int, int], tuple[int, int]]] = []
    play_count: Counter[int] = Counter()

    def candidates() -> list[tuple[tuple[int, int], tuple[int, int]]]:
        uniq = [e for e, c in remaining.items() if c > 0]
        out: list[tuple[tuple[int, int], tuple[int, int]]] = []
        for e1, e2 in combinations(uniq, 2):
            if len({e1[0], e1[1], e2[0], e2[1]}) == 4:
                out.append((e1, e2))
        return out

    def backtrack() -> bool:
        if sum(remaining.values()) == 0:
            return True
        cands = candidates()
        cands.sort(key=lambda p: sum(play_count[t] for t in p[0] + p[1]))
        for e1, e2 in cands[:180]:
            remaining[e1] -= 1
            remaining[e2] -= 1
            for t in e1 + e2:
                play_count[t] += 1
            rounds.append((e1, e2))
            if backtrack():
                return True
            rounds.pop()
            for t in e1 + e2:
                play_count[t] -= 1
            remaining[e1] += 1
            remaining[e2] += 1
        return False

    if not backtrack():
        raise RuntimeError("Could not pack games into 20 dual-court rounds")

    packed: list[tuple[tuple[str, str], tuple[str, str]]] = []
    for e1, e2 in rounds:
        g1 = oriented[e1].pop()
        g2 = oriented[e2].pop()
        packed.append((g1, g2))
    return packed


def assign_refs(
    rounds: list[tuple[tuple[str, str], tuple[str, str]]], teams: list[str]
) -> list[dict]:
    """Assign 2 refs per round from the 4 sitters; target 5 refs per team."""
    ref_count: Counter[str] = Counter()
    result: list[dict] = []

    for (h1, a1), (h2, a2) in rounds:
        playing = {h1, a1, h2, a2}
        sitters = [t for t in teams if t not in playing]
        under = [t for t in sitters if ref_count[t] < REFS_PER_TEAM_PER_WEEK]
        under.sort(key=lambda t: (ref_count[t], teams.index(t)))
        if len(under) >= 2:
            r1, r2 = under[0], under[1]
        elif len(under) == 1:
            others = sorted(
                [t for t in sitters if t != under[0]],
                key=lambda t: (ref_count[t], teams.index(t)),
            )
            r1, r2 = under[0], others[0]
        else:
            sitters.sort(key=lambda t: (ref_count[t], teams.index(t)))
            r1, r2 = sitters[0], sitters[1]

        ref_count[r1] += 1
        ref_count[r2] += 1
        offs = [t for t in sitters if t not in (r1, r2)]
        result.append(
            {
                "court1": (h1, a1),
                "court2": (h2, a2),
                "ref1": r1,
                "ref2": r2,
                "off": offs,
            }
        )

    return result


def build_week_rounds(teams: list[str], week_index: int) -> list[dict]:
    if len(teams) != NUM_TEAMS:
        raise ValueError(f"8-team schedule requires exactly {NUM_TEAMS} teams")
    heavy = HEAVY_SETS[week_index % len(HEAVY_SETS)]
    raw = assign_home_away(teams, heavy, week_index)
    packed = pack_into_rounds(raw, teams)
    return assign_refs(packed, teams)


def build_season(teams: list[str], num_weeks: int = NUM_WEEKS_DEFAULT) -> list[list[dict]]:
    return [build_week_rounds(teams, w) for w in range(num_weeks)]


def flatten_week_games(rounds: list[dict]) -> list[dict[str, str]]:
    """Flatten rounds to sequential games for pairing/home validation."""
    games: list[dict[str, str]] = []
    game_num = 1
    for rnd in rounds:
        for court_key, ref_key in (("court1", "ref1"), ("court2", "ref2")):
            home, away = rnd[court_key]
            games.append(
                {
                    "game": str(game_num),
                    "team1": home,
                    "team2": away,
                    "ref": rnd[ref_key],
                }
            )
            game_num += 1
    return games


def validate_week(teams: list[str], rounds: list[dict]) -> list[str]:
    errors: list[str] = []
    if len(rounds) != ROUNDS_PER_WEEK:
        errors.append(f"expected {ROUNDS_PER_WEEK} rounds, got {len(rounds)}")

    games = flatten_week_games(rounds)
    if len(games) != GAMES_PER_WEEK:
        errors.append(f"expected {GAMES_PER_WEEK} games, got {len(games)}")

    team_games: Counter[str] = Counter()
    home_games: Counter[str] = Counter()
    away_games: Counter[str] = Counter()
    ref_games: Counter[str] = Counter()
    off_games: Counter[str] = Counter()
    pairing_counts: Counter[tuple[int, int]] = Counter()

    for rnd in rounds:
        (h1, a1), (h2, a2) = rnd["court1"], rnd["court2"]
        playing = {h1, a1, h2, a2}
        if len(playing) != 4:
            errors.append(f"round has overlapping players: {playing}")
        refs = {rnd["ref1"], rnd["ref2"]}
        if refs & playing:
            errors.append(f"ref also playing: {refs & playing}")
        if len(refs) != 2:
            errors.append("duplicate refs in round")
        offs = set(rnd["off"])
        if offs & playing or offs & refs:
            errors.append("off overlaps play/ref")
        if playing | refs | offs != set(teams):
            errors.append("round does not partition all teams")

        for team in refs:
            ref_games[team] += 1
        for team in offs:
            off_games[team] += 1

    for game in games:
        home, away = game["team1"], game["team2"]
        team_games[home] += 1
        team_games[away] += 1
        home_games[home] += 1
        away_games[away] += 1
        pairing_counts[pair_key(teams.index(home), teams.index(away))] += 1

    for team in teams:
        if team_games[team] != GAMES_PER_TEAM_PER_WEEK:
            errors.append(
                f"{team}: expected {GAMES_PER_TEAM_PER_WEEK} games, got {team_games[team]}"
            )
        if home_games[team] != HOME_PER_TEAM_PER_WEEK:
            errors.append(
                f"{team}: expected {HOME_PER_TEAM_PER_WEEK} home, got {home_games[team]}"
            )
        if away_games[team] != HOME_PER_TEAM_PER_WEEK:
            errors.append(
                f"{team}: expected {HOME_PER_TEAM_PER_WEEK} away, got {away_games[team]}"
            )
        if ref_games[team] != REFS_PER_TEAM_PER_WEEK:
            errors.append(
                f"{team}: expected {REFS_PER_TEAM_PER_WEEK} refs, got {ref_games[team]}"
            )
        if off_games[team] != OFF_PER_TEAM_PER_WEEK:
            errors.append(
                f"{team}: expected {OFF_PER_TEAM_PER_WEEK} offs, got {off_games[team]}"
            )

    for pairing in all_pairings(len(teams)):
        count = pairing_counts[pairing]
        if count not in (1, 2):
            names = (teams[pairing[0]], teams[pairing[1]])
            errors.append(f"pairing {names}: expected 1 or 2 games, got {count}")

    heavy_count = sum(1 for p in all_pairings(len(teams)) if pairing_counts[p] == 2)
    if heavy_count != 12:
        errors.append(f"expected 12 heavy pairings, got {heavy_count}")

    return errors


def validate_season(teams: list[str], weeks: list[list[dict]]) -> list[str]:
    errors: list[str] = []
    season_pairings: Counter[tuple[int, int]] = Counter()
    season_home: Counter[str] = Counter()
    season_away: Counter[str] = Counter()
    season_ref: Counter[str] = Counter()

    for week_idx, week in enumerate(weeks):
        errors.extend(f"week {week_idx + 1}: {e}" for e in validate_week(teams, week))
        for game in flatten_week_games(week):
            pairing = pair_key(teams.index(game["team1"]), teams.index(game["team2"]))
            season_pairings[pairing] += 1
            season_home[game["team1"]] += 1
            season_away[game["team2"]] += 1
        for rnd in week:
            season_ref[rnd["ref1"]] += 1
            season_ref[rnd["ref2"]] += 1

    for pairing in all_pairings(len(teams)):
        count = season_pairings[pairing]
        if count not in SEASON_GAMES_PER_PAIRING:
            errors.append(
                f"season pairing {pairing}: expected 8 or 9 games, got {count}"
            )

    nines = sum(1 for p in all_pairings(len(teams)) if season_pairings[p] == 9)
    eights = sum(1 for p in all_pairings(len(teams)) if season_pairings[p] == 8)
    if nines != 16 or eights != 12:
        errors.append(f"season pairing split: expected 16×9 and 12×8, got {nines}×9 and {eights}×8")

    for team in teams:
        if season_home[team] != 30:
            errors.append(f"{team}: season home expected 30, got {season_home[team]}")
        if season_away[team] != 30:
            errors.append(f"{team}: season away expected 30, got {season_away[team]}")
        if season_ref[team] != 30:
            errors.append(f"{team}: season refs expected 30, got {season_ref[team]}")

    return errors


def week_sheet_names(num_weeks: int) -> list[str]:
    # Match Six / Seven Team League naming: "Week N Schedule"
    return [f"Week {week_num} Schedule" for week_num in range(1, num_weeks + 1)]


def _style_score_cell(cell) -> None:
    cell.font = _FONT_TEAM
    cell.fill = _FILL_WHITE
    cell.border = _THIN
    cell.alignment = _ALIGN_SCORE


def _style_team_cell(cell, value: str) -> None:
    cell.value = value
    cell.font = _FONT_TEAM
    cell.fill = _FILL_WHITE
    cell.border = _THIN


def _apply_gap_columns(ws, row: int) -> None:
    for col in (6, 11):  # F and K — gray separators like Six Team
        cell = ws.cell(row, col)
        cell.fill = _FILL_GAP
        cell.font = _FONT_STD


def write_week_sheet(ws, rounds: list[dict]) -> None:
    """Dual-court layout matching Inaugural/BYOT: Court labels, then Game NN + Refs."""
    for letter, width in _WEEK_COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    # Row 1: Court 1 / Court 2 labels (Inaugural Draft dual-court sheets)
    court1 = ws.cell(1, 2)
    court1.value = "Court 1"
    court1.font = _FONT_GAME
    court1.alignment = _ALIGN_GAME
    court2 = ws.cell(1, 7)
    court2.value = "Court 2"
    court2.font = _FONT_GAME
    court2.alignment = _ALIGN_GAME
    _apply_gap_columns(ws, 1)
    ws.row_dimensions[1].height = 15.75

    row = 2
    for game_num, rnd in enumerate(rounds, start=1):
        h1, a1 = rnd["court1"]
        h2, a2 = rnd["court2"]
        game_row = row
        ref_row = row + 1

        game_cell = ws.cell(game_row, 1)
        game_cell.value = f"Game {game_num:02d}"
        game_cell.font = _FONT_GAME
        game_cell.alignment = _ALIGN_GAME
        game_cell.border = _THIN
        ws.merge_cells(start_row=game_row, start_column=1, end_row=ref_row, end_column=1)
        ws.cell(ref_row, 1).border = _THIN

        _style_team_cell(ws.cell(game_row, 2), h1)
        _style_score_cell(ws.cell(game_row, 3))
        _style_team_cell(ws.cell(game_row, 4), a1)
        _style_score_cell(ws.cell(game_row, 5))
        _style_team_cell(ws.cell(game_row, 7), h2)
        _style_score_cell(ws.cell(game_row, 8))
        _style_team_cell(ws.cell(game_row, 9), a2)
        _style_score_cell(ws.cell(game_row, 10))
        _apply_gap_columns(ws, game_row)

        ref1 = ws.cell(ref_row, 2)
        ref1.value = f"Refs: {rnd['ref1']}"
        ref1.font = _FONT_REF
        ref1.fill = _FILL_REF
        ref1.border = _THIN
        ws.merge_cells(start_row=ref_row, start_column=2, end_row=ref_row, end_column=4)
        for col in (3, 4, 5):
            ws.cell(ref_row, col).fill = _FILL_REF
            ws.cell(ref_row, col).border = _THIN

        ref2 = ws.cell(ref_row, 7)
        ref2.value = f"Refs: {rnd['ref2']}"
        ref2.font = _FONT_REF
        ref2.fill = _FILL_REF
        ref2.border = _THIN
        ws.merge_cells(start_row=ref_row, start_column=7, end_row=ref_row, end_column=9)
        for col in (8, 9, 10):
            ws.cell(ref_row, col).fill = _FILL_REF
            ws.cell(ref_row, col).border = _THIN

        _apply_gap_columns(ws, ref_row)

        off = ws.cell(ref_row, 12)
        off.value = f"Off: {', '.join(rnd['off'])}"
        off.font = _FONT_REF

        ws.row_dimensions[game_row].height = 15.75
        ws.row_dimensions[ref_row].height = 15.75
        row += 2

    # Gray spacer rows before standings (Six Team leaves a short gap)
    for spacer in range(row, row + 2):
        for col in range(1, 12):
            cell = ws.cell(spacer, col)
            cell.fill = _FILL_GAP
            cell.font = _FONT_STD
        ws.row_dimensions[spacer].height = 15.75


def style_week_standings_block(ws, header_row: int, num_teams: int) -> None:
    """Apply Six Team standings-header fill after formulas are written."""
    section_header = header_row - 2
    col_header = header_row - 1
    for r in range(max(1, section_header - 1), section_header + 1):
        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.fill = _FILL_STANDINGS_HDR
            cell.font = _FONT_STD
    hdr = ws.cell(section_header, 1)
    if hdr.value:
        hdr.alignment = _ALIGN_STANDINGS_HDR
        hdr.font = _FONT_STD
        hdr.fill = _FILL_STANDINGS_HDR
    for c in range(1, 4):
        ws.cell(col_header, c).font = _FONT_STD
    for r in range(header_row, header_row + num_teams):
        for c in range(1, 4):
            ws.cell(r, c).font = _FONT_STD
        ws.row_dimensions[r].height = 15.75


def write_workbook(
    path: str | Path,
    teams: list[str],
    weeks: list[list[dict]],
    league_name: str = "Eight Team Team-Ref League",
) -> None:
    path = Path(path)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws_teams = wb.create_sheet("Teams")
    ws_teams.cell(1, 1).value = "Team Names"
    for col, team in enumerate(teams, start=3):
        ws_teams.cell(1, col).value = team
    write_format_to_teams_sheet(ws_teams, TEAM_REF)
    ws_teams.cell(3, 1).value = "League"
    ws_teams.cell(3, 2).value = league_name

    # Season pairing counts for Schedule Generator matrix
    season_pairings: Counter[tuple[int, int]] = Counter()
    for week in weeks:
        for game in flatten_week_games(week):
            season_pairings[
                pair_key(teams.index(game["team1"]), teams.index(game["team2"]))
            ] += 1

    ws_gen = wb.create_sheet("Schedule Generator")
    for col, team in enumerate(teams, start=2):
        ws_gen.cell(2, col).value = team
    for row_idx, team in enumerate(teams, start=3):
        ws_gen.cell(row_idx, 1).value = team
        for col_idx, opponent in enumerate(teams, start=2):
            if team == opponent:
                ws_gen.cell(row_idx, col_idx).value = "-"
            else:
                pairing = pair_key(teams.index(team), teams.index(opponent))
                ws_gen.cell(row_idx, col_idx).value = season_pairings[pairing]

    ws_standings = wb.create_sheet("League Standings")
    ws_standings.cell(1, 1).value = "LEAGUE STANDINGS"
    ws_standings.cell(2, 1).value = "Team Name"
    ws_standings.cell(2, 2).value = "Points For"
    ws_standings.cell(2, 3).value = "Points Against"
    ws_standings.cell(2, 4).value = "Point Differential"
    ws_standings.cell(11, 1).value = "Week #"
    ws_standings.cell(11, 2).value = 0
    ws_standings.cell(16, 1).value = "Teams"
    ws_standings.cell(16, 2).value = "Wins"
    ws_standings.cell(16, 5).value = "Losses"

    for week_name, week_rounds in zip(week_sheet_names(len(weeks)), weeks):
        ws = wb.create_sheet(week_name)
        write_week_sheet(ws, week_rounds)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def setup_dual_court_standings(path: str | Path, teams: list[str]) -> None:
    """Apply BYOT-style dual-court win/loss formulas (Court 1 + Court 2)."""
    from update_byot_standings import (
        apply_workbook_font_name,
        setup_league_standings,
        setup_week_sheet,
    )

    path = Path(path)
    wb = openpyxl.load_workbook(path)
    week_sheets = [n for n in wb.sheetnames if n.startswith("Week ")]
    # Court header + 20 games × 2 rows → rows 1–41; spacers 42–43; standings at 44+
    min_start_row = 44
    week_start_rows: dict[str, int] = {}
    for week_name in week_sheets:
        start = setup_week_sheet(wb[week_name], teams, week_name, min_start_row=min_start_row)
        style_week_standings_block(wb[week_name], start, len(teams))
        week_start_rows[week_name] = start
    week_start_row = next(iter(week_start_rows.values()))
    setup_league_standings(wb, teams, week_sheets, week_start_row)
    apply_workbook_font_name(wb, "Arial")
    wb.calculation.iterate = True
    wb.calculation.maxIter = 100
    wb.calculation.maxChange = 0.001
    wb.save(path)


def run_validation(teams: list[str] | None = None, num_weeks: int = NUM_WEEKS_DEFAULT) -> int:
    teams = teams or [f"Team {i}" for i in range(1, NUM_TEAMS + 1)]
    weeks = build_season(teams, num_weeks)
    errors = validate_season(teams, weeks)

    print(f"Validated {num_weeks}-week season for {', '.join(teams)}")
    print(f"Heavy-set rotations: {len(HEAVY_SETS)}")
    if errors:
        print(f"FAILED with {len(errors)} error(s):")
        for err in errors:
            print(f"  - {err}")
        return 1

    print("All checks passed.")
    for week_idx, week in enumerate(weeks, start=1):
        print(
            f"  Week {week_idx}: {len(week)} rounds, "
            f"{len(flatten_week_games(week))} games, "
            f"5 play/ref/off balance OK"
        )
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate 8-team / 2-court team-ref league schedule"
    )
    parser.add_argument("--validate", action="store_true", help="Run validation only")
    parser.add_argument("--output", type=str, help="Write workbook to this path")
    parser.add_argument(
        "--teams",
        nargs=NUM_TEAMS,
        metavar="TEAM",
        help=f"Exactly {NUM_TEAMS} team names",
    )
    parser.add_argument("--weeks", type=int, default=NUM_WEEKS_DEFAULT)
    parser.add_argument(
        "--standings",
        action="store_true",
        help="After writing, apply dual-court standings formulas",
    )
    args = parser.parse_args()

    teams = list(args.teams) if args.teams else [f"Team {i}" for i in range(1, NUM_TEAMS + 1)]

    if args.validate:
        return run_validation(teams, args.weeks)

    if args.output:
        weeks = build_season(teams, args.weeks)
        errors = validate_season(teams, weeks)
        if errors:
            print("Validation failed; not writing workbook:")
            for err in errors:
                print(f"  - {err}")
            return 1
        write_workbook(args.output, teams, weeks)
        print(f"Wrote {args.output}")
        if args.standings:
            setup_dual_court_standings(args.output, teams)
            print("Applied dual-court standings formulas")
        return 0

    return run_validation(teams, args.weeks)


if __name__ == "__main__":
    raise SystemExit(main())
