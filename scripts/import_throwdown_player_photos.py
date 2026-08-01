#!/usr/bin/env python3
"""
Import Throw Down captains-sheet headshots + skill/gender/home league into players.

Usage:
  python3 scripts/import_throwdown_player_photos.py [--dry-run] [--xlsx PATH]

Requires .env.local with DATABASE_URL and BLOB_READ_WRITE_TOKEN.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.error import HTTPError
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...", file=sys.stderr)
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

try:
    import psycopg
except ImportError:
    print("Installing psycopg...", file=sys.stderr)
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "psycopg[binary]", "-q"])
    import psycopg

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Downloads" / "CAPTAINS' EDITION_ THE THROW DOWN - 5TH EDITION.xlsx"

PHOTO_SHEETS = {
    "Male Identifying Players": "male",
    "SheThey Players": "female",
}

SKILL_COLUMNS = {
    1: 60,  # Male (Advanced)
    3: 60,  # Male (Advanced cont.)
    5: 40,  # Male (Intermediate)
    7: 60,  # She/They (Advanced)
    9: 40,  # She/They (Intermediate)
}

HOME_LEAGUE_ALIASES: dict[str, str] = {
    "PHILLY": "philly_dodgeball",
    "PHILADELPHIA": "philly_dodgeball",
    "BOSTON": "boston_dodgeball_league",
    "BDL": "boston_dodgeball_league",
    "DMV": "dmv_dodgeball",
    "BIG APPLE": "big_apple_dodgeball",
    "NYC": "big_apple_dodgeball",
    "NEW YORK": "new_york_dodgeball",
    "LI KICK": "li_kick",
    "LONG ISLAND": "li_kick",
    "CONNECTICUT": "connecticut_dodgeball",
    "CT": "connecticut_dodgeball",
    "NUTMEG": "nutmeg_dodgeball",
    "CACTUS": "cactus_dodgeball",
    "ARIZONA": "cactus_dodgeball",
    "MINNESOTA": "minnesota_dodgeball",
    "DALLAS": "dallas_dodgeball",
    "PITTSBURGH": "three_rivers_dodgeball_club",
    "DENVER": "summit_sports_league",
    "PDX": "dodgeball_pdx",
    "PORTLAND": "dodgeball_pdx",
    "SAN DIEGO": "dodgeball_san_diego",
    "SEATTLE": "seattle_dodgeball",
    "LA": "world_dodgeball_society",
    "LOS ANGELES": "world_dodgeball_society",
}


@dataclass
class SheetPlayer:
    full_name: str
    first_name: str
    last_name: str
    gender: str
    home_city_raw: str | None
    home_league: str | None
    skill_level: int | None
    image_bytes: bytes | None = None
    image_ext: str = "jpg"
    sheet: str = ""


@dataclass
class Report:
    matched: list[str] = field(default_factory=list)
    unmatched: list[str] = field(default_factory=list)
    ambiguous: list[str] = field(default_factory=list)
    unmapped_cities: list[str] = field(default_factory=list)
    updated_photo: int = 0
    updated_gender: int = 0
    updated_skill: int = 0
    updated_home_league: int = 0
    skipped_photo_exists: int = 0


def load_dotenv(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    raw = path.read_bytes()
    # Handle BASE64 / multiline carefully: only single-line KEY=VALUE entries here.
    for m in re.finditer(rb"^([A-Za-z_][A-Za-z0-9_]*)=(.*)$", raw, re.M):
        key = m.group(1).decode()
        val = m.group(2).decode()
        if (val.startswith('"') and val.endswith('"')) or (
            val.startswith("'") and val.endswith("'")
        ):
            val = val[1:-1]
        env[key] = val.replace("\\n", "\n")
    return env


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip())


def strip_parenthetical(name: str) -> str:
    return normalize_space(re.sub(r"\([^)]*\)", " ", name))


def split_name(full_name: str) -> tuple[str, str]:
    cleaned = strip_parenthetical(full_name)
    parts = cleaned.split(" ")
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def name_key(first: str, last: str) -> str:
    return f"{normalize_space(first).lower()}|{normalize_space(last).lower()}"


def full_name_key(full_name: str) -> str:
    first, last = split_name(full_name)
    return name_key(first, last)


def map_home_league(raw: str | None) -> tuple[str | None, str | None]:
    if raw is None:
        return None, None
    label = normalize_space(raw).upper()
    if not label:
        return None, None
    code = HOME_LEAGUE_ALIASES.get(label)
    if code:
        return code, None
    return None, label


def image_bytes_and_ext(img: Any) -> tuple[bytes, str]:
    data = img._data()
    if not isinstance(data, (bytes, bytearray)):
        data = bytes(data)
    fmt = (getattr(img, "format", None) or "jpeg").lower()
    if fmt == "jpeg":
        ext = "jpg"
    elif fmt in ("png", "gif", "webp"):
        ext = fmt
    else:
        ext = "jpg"
    return bytes(data), ext


def build_skill_map(wb: openpyxl.Workbook) -> dict[str, int]:
    ws = wb["Players"]
    skill: dict[str, int] = {}
    for col, level in SKILL_COLUMNS.items():
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, col).value
            if not val or not isinstance(val, str):
                continue
            name = normalize_space(val)
            if not name:
                continue
            skill[full_name_key(name)] = level
            # Also index without parenthetical for matching "Nicole Landro"
            skill[full_name_key(strip_parenthetical(name))] = level
    return skill


def extract_sheet_players(
    wb: openpyxl.Workbook, sheet_name: str, gender: str, skill_map: dict[str, int]
) -> list[SheetPlayer]:
    ws = wb[sheet_name]
    players: list[SheetPlayer] = []
    for img in list(getattr(ws, "_images", []) or []):
        anchor = img.anchor
        frm = getattr(anchor, "_from", None)
        if frm is None:
            continue
        excel_col = frm.col + 1
        excel_row = frm.row + 1
        name = None
        city = None
        for r in range(excel_row, min(excel_row + 14, ws.max_row + 1)):
            val = ws.cell(r, excel_col).value
            if not val or not isinstance(val, str):
                continue
            text = normalize_space(val)
            if not text:
                continue
            if name is None:
                name = text
            else:
                city = text
                break
        if not name:
            continue
        first, last = split_name(name)
        home_league, unmapped = map_home_league(city)
        key = full_name_key(name)
        skill = skill_map.get(key) or skill_map.get(full_name_key(strip_parenthetical(name)))
        img_bytes, ext = image_bytes_and_ext(img)
        players.append(
            SheetPlayer(
                full_name=name,
                first_name=first,
                last_name=last,
                gender=gender,
                home_city_raw=city,
                home_league=home_league,
                skill_level=skill,
                image_bytes=img_bytes,
                image_ext=ext,
                sheet=sheet_name,
            )
        )
        # stash unmapped on object via home_city_raw; report later
        if unmapped:
            players[-1].home_city_raw = unmapped  # keep raw for report
    return players


def upload_blob(token: str, pathname: str, data: bytes, content_type: str) -> dict[str, str]:
    # Vercel Blob REST put
    url = f"https://blob.vercel-storage.com/{quote(pathname, safe='/')}"
    req = Request(
        url,
        data=data,
        method="PUT",
        headers={
            "Authorization": f"Bearer {token}",
            "x-api-version": "7",
            "Content-Type": content_type,
        },
    )
    try:
        with urlopen(req) as res:
            body = json.loads(res.read().decode())
    except HTTPError as e:
        detail = e.read().decode()[:500]
        raise RuntimeError(f"Blob upload failed ({e.code}): {detail}") from e
    return {
        "url": body["url"],
        "pathname": body.get("pathname") or pathname,
    }


def content_type_for_ext(ext: str) -> str:
    return {
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "webp": "image/webp",
        "gif": "image/gif",
    }.get(ext, "image/jpeg")


def load_db_players(conn: psycopg.Connection) -> list[dict[str, Any]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, first_name, last_name, roster_name, gender, skill_level,
                   photo_url, photo_pathname, is_merged
            FROM players
            WHERE is_merged = false
            """
        )
        cols = [d.name for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


def index_db_players(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    by_key: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        keys = {
            name_key(row["first_name"], row["last_name"]),
            full_name_key(row["roster_name"] or ""),
        }
        for k in keys:
            if not k or k == "|":
                continue
            by_key.setdefault(k, []).append(row)
    return by_key


def resolve_player(
    sheet_player: SheetPlayer, by_key: dict[str, list[dict[str, Any]]]
) -> tuple[dict[str, Any] | None, str | None]:
    candidates_keys = [
        name_key(sheet_player.first_name, sheet_player.last_name),
        full_name_key(sheet_player.full_name),
        full_name_key(strip_parenthetical(sheet_player.full_name)),
        # Handle DB rows with swapped first/last (e.g. "Salamone Caysie")
        name_key(sheet_player.last_name, sheet_player.first_name),
    ]
    seen_ids: set[str] = set()
    matches: list[dict[str, Any]] = []
    for k in candidates_keys:
        for row in by_key.get(k, []):
            if row["id"] in seen_ids:
                continue
            seen_ids.add(row["id"])
            matches.append(row)
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        return None, "ambiguous"
    # Soft match: last name + first initial (either order)
    last = normalize_space(sheet_player.last_name).lower()
    first = normalize_space(sheet_player.first_name).lower()
    soft: list[dict[str, Any]] = []
    for rows in by_key.values():
        for row in rows:
            if row["id"] in seen_ids:
                continue
            db_first = normalize_space(row["first_name"]).lower()
            db_last = normalize_space(row["last_name"]).lower()
            if (db_last == last and db_first[:1] == first[:1]) or (
                db_last == first and db_first[:1] == last[:1]
            ):
                soft.append(row)
                seen_ids.add(row["id"])
    unique_soft = {r["id"]: r for r in soft}
    soft_list = list(unique_soft.values())
    if len(soft_list) == 1:
        return soft_list[0], None
    if len(soft_list) > 1:
        return None, "ambiguous"
    return None, "unmatched"


def ensure_home_league(conn: psycopg.Connection, player_id: str, home_league: str) -> bool:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT 1 FROM player_home_leagues
            WHERE player_id = %s AND home_league = %s
            LIMIT 1
            """,
            (player_id, home_league),
        )
        if cur.fetchone():
            return False
        cur.execute(
            """
            SELECT COALESCE(MAX(sort_order), -1) FROM player_home_leagues
            WHERE player_id = %s
            """,
            (player_id,),
        )
        next_sort = cur.fetchone()[0] + 1
        cur.execute(
            """
            INSERT INTO player_home_leagues (player_id, home_league, sort_order)
            VALUES (%s, %s, %s)
            """,
            (player_id, home_league, next_sort),
        )
        return True


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--overwrite-photos", action="store_true")
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"Missing xlsx: {args.xlsx}", file=sys.stderr)
        return 1

    env = {**os.environ, **load_dotenv(ROOT / ".env.local")}
    database_url = env.get("DATABASE_URL")
    blob_token = env.get("BLOB_READ_WRITE_TOKEN")
    if not database_url:
        print("DATABASE_URL is required", file=sys.stderr)
        return 1
    if not args.dry_run and not blob_token:
        print("BLOB_READ_WRITE_TOKEN is required unless --dry-run", file=sys.stderr)
        return 1

    print(f"Loading {args.xlsx} ...")
    wb = openpyxl.load_workbook(args.xlsx)
    skill_map = build_skill_map(wb)
    sheet_players: list[SheetPlayer] = []
    for sheet_name, gender in PHOTO_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"WARN: missing sheet {sheet_name}")
            continue
        extracted = extract_sheet_players(wb, sheet_name, gender, skill_map)
        print(f"  {sheet_name}: {len(extracted)} players with photos")
        sheet_players.extend(extracted)

    report = Report()
    for sp in sheet_players:
        if sp.home_city_raw and sp.home_league is None:
            # re-check: extract stored unmapped upper label
            code, unmapped = map_home_league(sp.home_city_raw)
            if code:
                sp.home_league = code
            elif unmapped and unmapped not in report.unmapped_cities:
                report.unmapped_cities.append(unmapped)

    with psycopg.connect(database_url) as conn:
        db_rows = load_db_players(conn)
        by_key = index_db_players(db_rows)
        print(f"DB players (active): {len(db_rows)}")

        for sp in sheet_players:
            player, status = resolve_player(sp, by_key)
            if status == "ambiguous":
                report.ambiguous.append(sp.full_name)
                continue
            if player is None:
                report.unmatched.append(sp.full_name)
                continue

            report.matched.append(f"{sp.full_name} -> {player['first_name']} {player['last_name']}")
            updates: dict[str, Any] = {}
            if sp.gender and not player["gender"]:
                updates["gender"] = sp.gender
            if sp.skill_level is not None and player["skill_level"] is None:
                updates["skill_level"] = sp.skill_level

            photo_url = None
            photo_pathname = None
            should_set_photo = sp.image_bytes and (
                args.overwrite_photos or not player["photo_url"]
            )
            if sp.image_bytes and player["photo_url"] and not args.overwrite_photos:
                report.skipped_photo_exists += 1

            if should_set_photo and not args.dry_run:
                pathname = f"player-photos/{player['id']}/{uuid.uuid4()}.{sp.image_ext}"
                blob = upload_blob(
                    blob_token or "",
                    pathname,
                    sp.image_bytes or b"",
                    content_type_for_ext(sp.image_ext),
                )
                photo_url = blob["url"]
                photo_pathname = blob["pathname"]
                updates["photo_url"] = photo_url
                updates["photo_pathname"] = photo_pathname
            elif should_set_photo and args.dry_run:
                updates["photo_url"] = "(dry-run)"
                updates["photo_pathname"] = "(dry-run)"

            if args.dry_run:
                if "photo_url" in updates:
                    report.updated_photo += 1
                if "gender" in updates:
                    report.updated_gender += 1
                if "skill_level" in updates:
                    report.updated_skill += 1
                if sp.home_league:
                    report.updated_home_league += 1
                continue

            if updates:
                sets = []
                vals: list[Any] = []
                for col, val in updates.items():
                    sets.append(f"{col} = %s")
                    vals.append(val)
                sets.append("updated_at = NOW()")
                vals.append(player["id"])
                with conn.cursor() as cur:
                    cur.execute(
                        f"UPDATE players SET {', '.join(sets)} WHERE id = %s",
                        vals,
                    )
                if "photo_url" in updates:
                    report.updated_photo += 1
                    player["photo_url"] = updates["photo_url"]
                if "gender" in updates:
                    report.updated_gender += 1
                    player["gender"] = updates["gender"]
                if "skill_level" in updates:
                    report.updated_skill += 1
                    player["skill_level"] = updates["skill_level"]

            if sp.home_league:
                if ensure_home_league(conn, player["id"], sp.home_league):
                    report.updated_home_league += 1

        if not args.dry_run:
            conn.commit()

    print("\n=== Import report ===")
    print(f"Matched: {len(report.matched)}")
    print(f"Unmatched: {len(report.unmatched)}")
    print(f"Ambiguous: {len(report.ambiguous)}")
    print(f"Photos set: {report.updated_photo} (skipped existing: {report.skipped_photo_exists})")
    print(f"Gender filled: {report.updated_gender}")
    print(f"Skill filled: {report.updated_skill}")
    print(f"Home leagues added: {report.updated_home_league}")
    if report.unmapped_cities:
        print(f"Unmapped cities: {', '.join(sorted(report.unmapped_cities))}")
    if report.unmatched:
        print("\nUnmatched names:")
        for n in report.unmatched:
            print(f"  - {n}")
    if report.ambiguous:
        print("\nAmbiguous names:")
        for n in report.ambiguous:
            print(f"  - {n}")
    if args.dry_run:
        print("\n(dry-run: no DB or Blob writes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
