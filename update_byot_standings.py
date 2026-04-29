"""
Update standings for Winter 2026 BYOT League.xlsx
Adds standings calculations to each of the 3 weekly schedule tabs
and updates the overall standings tab.
"""

import sys
import os
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font

# Import functions from setup_standings.py
from setup_standings import (
    detect_teams,
    detect_week_sheets,
    setup_head_to_head_matrix,
)

def is_cell_writable(ws, row, col):
    """Check if a cell can be written to (not a merged cell)."""
    try:
        cell = ws.cell(row, col)
        # MergedCell objects are read-only
        if isinstance(cell, MergedCell):
            return False
        return True
    except:
        return False

def find_win_loss_section(ws, num_teams=6, min_start_row=30):
    """Find where to place the win/loss section, avoiding merged cells.

    min_start_row: first row allowed for the header "Team Wins/Losses This Week"
    (e.g. 40 when game blocks run through row 39).

    We need to find a row where we can write:
    - start_row: header "Team Wins/Losses This Week"
    - start_row + 1: column headers
    - start_row + 2 to start_row + 2 + num_teams: team data
    """
    # Look for existing "Team Wins" or similar header (only at/after min_start_row)
    scan_top = max(1, min_start_row)
    for row in range(scan_top, scan_top + 35):
        try:
            cell = ws.cell(row, 1)
            # Skip if it's a merged cell (read-only)
            if isinstance(cell, MergedCell):
                continue
            cell_value = cell.value
            if cell_value and isinstance(cell_value, str):
                if 'win' in cell_value.lower() and 'loss' in cell_value.lower():
                    # Check if we can write to the rows we need
                    if (is_cell_writable(ws, row + 1, 1) and
                            is_cell_writable(ws, row + 2, 1)):
                        return row
        except Exception:
            continue

    # Find a safe row (not merged, preferably empty), starting no earlier than min_start_row
    # We need at least 3 + num_teams rows free
    for start_row in range(min_start_row, min_start_row + 80):
        # Check if all required cells are writable
        all_writable = True
        for offset in range(0, 3 + num_teams):
            if not is_cell_writable(ws, start_row + offset, 1):
                all_writable = False
                break
            if offset > 0:  # Also check columns B and C for data rows
                if not is_cell_writable(ws, start_row + offset, 2):
                    all_writable = False
                    break
                if not is_cell_writable(ws, start_row + offset, 3):
                    all_writable = False
                    break

        if all_writable:
            return start_row

    # Last resort
    return min_start_row

def setup_week_sheet(ws, teams, week_name, min_start_row=30):
    """Set up win/loss formulas for a week sheet.

    min_start_row: first row for the standings header (default 30; use 40+ if games extend lower).
    """
    start_row = find_win_loss_section(ws, len(teams), min_start_row=min_start_row)
    
    # Verify cells are writable before writing
    if not is_cell_writable(ws, start_row, 1):
        raise ValueError(f"Cannot write to row {start_row} in {week_name} - cell is merged")
    
    # Headers
    ws.cell(start_row, 1).value = 'Team Wins/Losses This Week'
    if is_cell_writable(ws, start_row + 1, 1):
        ws.cell(start_row + 1, 1).value = 'Team Name'
    if is_cell_writable(ws, start_row + 1, 2):
        ws.cell(start_row + 1, 2).value = 'Wins'
    if is_cell_writable(ws, start_row + 1, 3):
        ws.cell(start_row + 1, 3).value = 'Losses'
    
    # Add team names and formulas
    for i, team in enumerate(teams, start=start_row + 2):
        # Team name
        if is_cell_writable(ws, i, 1):
            ws.cell(i, 1).value = team
        
        # Wins formula: Count wins from both courts
        # Court 1: Team in B and score in C (Team 1 wins), or Team in D and score in E (Team 2 wins)
        # Court 2: Team in G and score in H (Team 1 wins), or Team in I and score in J (Team 2 wins)
        if ' ' in team or team.endswith('*'):
            # Handle teams with spaces or special characters
            team_pattern = f'"{team}*"' if not team.endswith('*') else f'"{team}"'
        else:
            team_pattern = f'"{team}"'
        
        # Wins: Court 1 (B/C and D/E) + Court 2 (G/H and I/J)
        wins_formula = (f'=SUMIFS(C:C,B:B,{team_pattern})+SUMIFS(E:E,D:D,{team_pattern})+'
                       f'SUMIFS(H:H,G:G,{team_pattern})+SUMIFS(J:J,I:I,{team_pattern})')
        if is_cell_writable(ws, i, 2):
            ws.cell(i, 2).value = wins_formula
        
        # Losses formula: Count losses from both courts
        # Court 1: Team in B and score in E (Team 2 wins, so Team 1 loses), or Team in D and score in C (Team 1 wins, so Team 2 loses)
        # Court 2: Team in G and score in J (Team 2 wins, so Team 1 loses), or Team in I and score in H (Team 1 wins, so Team 2 loses)
        losses_formula = (f'=SUMIFS(E:E,B:B,{team_pattern})+SUMIFS(C:C,D:D,{team_pattern})+'
                         f'SUMIFS(J:J,G:G,{team_pattern})+SUMIFS(H:H,I:I,{team_pattern})')
        if is_cell_writable(ws, i, 3):
            ws.cell(i, 3).value = losses_formula
    
    return start_row + 2  # Return the first data row

def setup_league_standings(wb, teams, week_sheets, week_start_row):
    """Set up the League Standings sheet, preserving existing content."""
    if 'League Standings' not in wb.sheetnames:
        # Create it if it doesn't exist
        ws = wb.create_sheet('League Standings')
        # Add title if creating new sheet
        ws.cell(1, 1).value = 'LEAGUE STANDINGS'
    else:
        ws = wb['League Standings']
        # Preserve title at row 1 if it exists, otherwise add it
        if not ws.cell(1, 1).value:
            ws.cell(1, 1).value = 'LEAGUE STANDINGS'
    
    # Only clear team data rows (17-30), preserve everything else
    for row in range(17, 31):
        for col in range(1, 6):
            ws.cell(row, col).value = None
    
    # Headers for team data section (only if not already present)
    if not ws.cell(16, 1).value:
        ws.cell(16, 1).value = 'Teams'
    if not ws.cell(16, 2).value:
        ws.cell(16, 2).value = 'Wins'
    if not ws.cell(16, 5).value:
        ws.cell(16, 5).value = 'Losses'  # Column E
    
    # Add team names and aggregate formulas
    for i, team in enumerate(teams, start=17):
        # Team name
        ws.cell(i, 1).value = team
        
        # Wins aggregation (from column B of week sheets)
        wins_row = week_start_row + i - 17  # B37 for row 17, B38 for row 18, etc.
        wins_parts = [f"'{week}'!B{wins_row}" for week in week_sheets]
        ws.cell(i, 2).value = '=' + '+'.join(wins_parts)
        
        # Tie-break for sort: wins + row/10000. Use literal row i (not ROW(Bi)) so merged
        # cells in column B don't make every ROW() identical (which breaks MATCH/LARGE).
        ws.cell(i, 3).value = f'=ROUND(B{i}+{i}/10000,8)'
        
        # Losses aggregation (from column C of week sheets)
        losses_parts = [f"'{week}'!C{wins_row}" for week in week_sheets]
        ws.cell(i, 5).value = '=' + '+'.join(losses_parts)
    
    # Set up standings display (rows 2-9, but only update formulas for the number of teams we have)
    # Preserve headers at row 2 if they exist
    if not ws.cell(2, 1).value:
        ws.cell(2, 1).value = 'Team Name'
    if not ws.cell(2, 2).value:
        ws.cell(2, 2).value = 'Points For'
    if not ws.cell(2, 3).value:
        ws.cell(2, 3).value = 'Points Against'
    if not ws.cell(2, 4).value:
        ws.cell(2, 4).value = 'Point Differential'
    
    # Update standings formulas (only for the number of teams we have)
    num_teams = len(teams)
    c_lo, c_hi = 17, 16 + num_teams
    c_rng = f'C{c_lo}:C{c_hi}'
    for i, rank in enumerate(range(1, num_teams + 1), start=3):
        # ROUND helps MATCH find LARGE(...) when Excel uses slightly different float bits
        large_k = f'ROUND(LARGE({c_rng},{rank}),8)'
        # Team name
        ws.cell(i, 1).value = (
            f'=INDEX(A{c_lo}:A{c_hi},MATCH({large_k},{c_rng},0))'
        )
        # Wins (Points For)
        ws.cell(i, 2).value = (
            f'=INDEX(B{c_lo}:B{c_hi},MATCH({large_k},{c_rng},0))'
        )
        # Points Against (Losses)
        ws.cell(i, 3).value = (
            f'=INDEX(E{c_lo}:E{c_hi},MATCH({large_k},{c_rng},0))'
        )
        
        # Point Differential
        ws.cell(i, 4).value = f'=B{i}-C{i}'
    
    # Clear any extra rows beyond the number of teams
    for row in range(3 + num_teams, 10):
        for col in range(1, 5):
            if ws.cell(row, col).value and 'INDEX' in str(ws.cell(row, col).value):
                ws.cell(row, col).value = None
    
    # Set week number to 0 if not set
    if ws.cell(11, 2).value is None:
        if not ws.cell(11, 1).value:
            ws.cell(11, 1).value = 'Week #'
        ws.cell(11, 2).value = 0

    setup_head_to_head_matrix(ws, teams, week_sheets, dual_court=True)


def setup_public_standings_sheet(wb, num_teams):
    """Add a clean 'Public Standings' tab (rank + team + W/L + diff) linked to League Standings.

    Safe to share or print: no raw team rows, week tabs, or schedule detail — only references
    to the ranked table on League Standings (rows 3..2+num_teams).
    """
    if "League Standings" not in wb.sheetnames:
        raise ValueError("League Standings sheet must exist before Public Standings")

    if "Public Standings" in wb.sheetnames:
        wb.remove(wb["Public Standings"])

    insert_at = wb.sheetnames.index("League Standings") + 1
    ws = wb.create_sheet("Public Standings", insert_at)

    # Title follows League Standings heading
    ws.cell(1, 1).value = "='League Standings'!A1"

    ws.cell(2, 1).value = "Rank"
    ws.cell(2, 2).value = "Team"
    ws.cell(2, 3).value = "Wins"
    ws.cell(2, 4).value = "Losses"
    ws.cell(2, 5).value = "Point Differential"

    max_rank_row = 9  # same upper bound as League Standings display block
    for r in range(3, max_rank_row + 1):
        for c in range(1, 6):
            ws.cell(r, c).value = None

    for offset, r in enumerate(range(3, 3 + num_teams)):
        ws.cell(r, 1).value = offset + 1
        ws.cell(r, 2).value = f"='League Standings'!A{r}"
        ws.cell(r, 3).value = f"='League Standings'!B{r}"
        ws.cell(r, 4).value = f"='League Standings'!C{r}"
        ws.cell(r, 5).value = f"='League Standings'!D{r}"

    note_row = 3 + num_teams + 1
    ws.cell(note_row, 1).value = (
        "Standings mirror the League Standings tab — update scores on week schedule sheets."
    )

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 18


def apply_workbook_font_name(wb, font_name="Commissioner"):
    """Set font family on every non-merged cell; keep size, bold, color, etc."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                old = cell.font
                sz = old.sz
                if sz is None and getattr(old, "size", None) is not None:
                    sz = old.size
                cell.font = Font(
                    name=font_name,
                    sz=sz,
                    b=old.bold,
                    i=old.italic,
                    u=old.u,
                    strike=old.strike,
                    color=old.color,
                    vertAlign=old.vertAlign,
                    charset=old.charset,
                    family=old.family,
                    outline=old.outline,
                    shadow=old.shadow,
                    condense=old.condense,
                    extend=old.extend,
                    scheme=old.scheme,
                )


def main():
    """Main function to update standings for BYOT League."""
    file_path = 'public/league_schedules/Winter 2026 BYOT League.xlsx'
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        sys.exit(1)
    
    print(f"Loading workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    
    print("\n=== Detecting Teams ===")
    all_teams = detect_teams(wb)
    # Filter out referee assignments (entries starting with "Refs:")
    teams = [team for team in all_teams if not team.startswith('Refs:')]
    print(f"Found {len(teams)} teams (filtered from {len(all_teams)} total entries): {', '.join(teams)}")
    
    if not teams:
        print("ERROR: No teams detected! Please check the Teams sheet or week sheets.")
        return
    
    print("\n=== Detecting Week Sheets ===")
    week_sheets = detect_week_sheets(wb)
    print(f"Found {len(week_sheets)} week sheets:")
    for week in week_sheets:
        print(f"  - {week}")
    
    if not week_sheets:
        print("ERROR: No week sheets detected!")
        return
    
    print("\n=== Setting Up Week Sheets ===")
    week_start_rows = {}
    for week_name in week_sheets:
        ws = wb[week_name]
        print(f"  Processing {week_name}...")
        start_row = setup_week_sheet(ws, teams, week_name)
        week_start_rows[week_name] = start_row
        print(f"    Added win/loss formulas starting at row {start_row}")
    
    # Use the first week's start row (they should all be the same)
    week_start_row = list(week_start_rows.values())[0] if week_start_rows else 32
    
    print("\n=== Setting Up League Standings ===")
    setup_league_standings(wb, teams, week_sheets, week_start_row)
    print("  League Standings sheet configured")

    print("\n=== Public Standings sheet ===")
    setup_public_standings_sheet(wb, len(teams))
    print("  Public Standings tab added (linked to League Standings)")
    
    print("\n=== Enabling Iterative Calculation ===")
    wb.calculation.iterate = True
    wb.calculation.maxIter = 100
    wb.calculation.maxChange = 0.001
    print("  Iterative calculation enabled")
    
    print("\n=== Applying Commissioner font ===")
    apply_workbook_font_name(wb)
    print("  Font set to Commissioner on all cells")

    print("\n=== Saving Workbook ===")
    wb.save(file_path)
    print(f"  Saved to {file_path}")
    
    print("\n✅ Setup complete!")
    print(f"\nSummary:")
    print(f"  - {len(teams)} teams configured")
    print(f"  - {len(week_sheets)} week sheets configured")
    print(f"  - League Standings sheet ready")
    print(f"  - All formulas applied")

if __name__ == '__main__':
    main()
