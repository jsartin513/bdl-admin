"""
Create a new league spreadsheet template with all necessary sheets and structure.
"""

import argparse
import openpyxl
from datetime import datetime, timedelta

from league_schedule_format import DEDICATED_REF, TEAM_REF, write_format_to_teams_sheet


def create_league_template(
    output_path,
    league_name,
    teams,
    num_weeks=6,
    start_date=None,
    schedule_format=TEAM_REF,
):
    """Create a new league spreadsheet template."""
    if not teams or len(teams) < 2:
        print("ERROR: Need at least 2 teams!")
        return
    """Create a new league spreadsheet template."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    print(f"Creating league template: {league_name}")
    print(f"Teams: {', '.join(teams)}")
    print(f"Weeks: {num_weeks}")
    print(f"Format: {schedule_format}")

    # 1. Teams Sheet
    ws_teams = wb.create_sheet('Teams')
    ws_teams.cell(1, 1).value = 'Team Names'
    for i, team in enumerate(teams, start=3):
        ws_teams.cell(1, i).value = team
    write_format_to_teams_sheet(ws_teams, schedule_format)
    print("  ✓ Created Teams sheet")
    
    # 2. Schedule Generator Sheet
    ws_gen = wb.create_sheet('Schedule Generator')
    
    # Create matchup matrix
    col = 2
    for team in teams:
        ws_gen.cell(2, col).value = team
        col += 1
    
    row = 3
    for team in teams:
        ws_gen.cell(row, 1).value = team
        col = 2
        for opponent in teams:
            if team == opponent:
                ws_gen.cell(row, col).value = '-'
            else:
                ws_gen.cell(row, col).value = 2  # Default: 2 games per matchup
            col += 1
        row += 1
    
    print("  ✓ Created Schedule Generator sheet")
    
    # 3. League Standings Sheet
    ws_standings = wb.create_sheet('League Standings')
    
    # Headers
    ws_standings.cell(1, 1).value = 'LEAGUE STANDINGS'
    ws_standings.cell(2, 1).value = 'Team Name'
    ws_standings.cell(2, 2).value = 'Points For'
    ws_standings.cell(2, 3).value = 'Points Against'
    ws_standings.cell(2, 4).value = 'Point Differential'
    
    # Week number
    ws_standings.cell(11, 1).value = 'Week #'
    ws_standings.cell(11, 2).value = 0
    
    # Team data section headers
    ws_standings.cell(16, 1).value = 'Teams'
    ws_standings.cell(16, 2).value = 'Wins'
    ws_standings.cell(16, 5).value = 'Losses'
    
    print("  ✓ Created League Standings sheet")
    
    # 4. Create Week Sheets
    if start_date is None:
        # Default to next Sunday
        today = datetime.now()
        days_until_sunday = (6 - today.weekday()) % 7
        if days_until_sunday == 0:
            days_until_sunday = 7
        start_date = today + timedelta(days=days_until_sunday)
    
    for week_num in range(1, num_weeks + 1):
        week_date = start_date + timedelta(weeks=week_num - 1)
        week_name = f'Week {week_num} ({week_date.month}.{week_date.day})'
        
        ws_week = wb.create_sheet(week_name)
        ws_week.cell(1, 2).value = 'Court 1'
        
        print(f"  ✓ Created {week_name} sheet")
    
    # Save
    wb.save(output_path)
    print(f"\n✅ Template created: {output_path}")
    print(f"\nNext steps:")
    print(f"  1. Fill in the Schedule Generator with desired matchups")
    print(f"  2. Run: python3 create_schedule_from_generator.py {output_path}")
    print(f"  3. Run: python3 setup_standings.py {output_path}")
    print(f"  4. Review and adjust the generated schedule")

if __name__ == '__main__':
    import sys

    parser = argparse.ArgumentParser(description='Create a new league spreadsheet template')
    parser.add_argument('output_path', nargs='?', help='Output .xlsx path')
    parser.add_argument('teams', nargs='*', help='Team names')
    parser.add_argument(
        '--format',
        choices=[TEAM_REF, DEDICATED_REF],
        default=TEAM_REF,
        help='Schedule format stored on Teams sheet',
    )
    parser.add_argument('--weeks', type=int, default=6, help='Number of week sheets')
    args = parser.parse_args()

    if args.output_path and args.teams:
        output_path = args.output_path
        teams = list(args.teams)
        league_name = "New League"
        num_weeks = args.weeks
        schedule_format = args.format
    else:
        # Interactive mode
        print("🏀 League Template Creator")
        print("=" * 50)
        print()
        
        league_name = input("League name: ").strip()
        if not league_name:
            league_name = "New League"
        
        output_path = input(f"Output filename (default: '{league_name}.xlsx'): ").strip()
        if not output_path:
            output_path = f"{league_name}.xlsx"
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
        
        print()
        print("Enter team names (press Enter on empty line when done):")
        teams = []
        while True:
            team = input(f"  Team {len(teams) + 1}: ").strip()
            if not team:
                if len(teams) < 2:
                    print("  ⚠️  Need at least 2 teams. Enter another team:")
                    continue
                break
            teams.append(team)
        
        num_weeks_input = input("\nNumber of weeks (default: 6): ").strip()
        num_weeks = int(num_weeks_input) if num_weeks_input.isdigit() else 6

        format_input = input("Schedule format (team-ref / dedicated-ref, default: team-ref): ").strip()
        schedule_format = format_input if format_input in (TEAM_REF, DEDICATED_REF) else TEAM_REF

        print()
        print(f"📋 Summary:")
        print(f"  League: {league_name}")
        print(f"  Teams: {', '.join(teams)}")
        print(f"  Weeks: {num_weeks}")
        print(f"  Format: {schedule_format}")
        print(f"  Output: {output_path}")
        
        confirm = input("\nCreate template? (Y/n): ").strip().lower()
        if confirm and confirm != 'y':
            print("Cancelled.")
            sys.exit(0)
        
        print()

    create_league_template(
        output_path,
        league_name,
        teams,
        num_weeks,
        schedule_format=schedule_format,
    )

