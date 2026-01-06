"""
Automatically set up standings formulas for a league schedule spreadsheet.
Detects teams and weeks, then applies all necessary formulas.
"""

import openpyxl
import re

def detect_teams(wb):
    """Detect team names from the Teams sheet, League Standings, or week sheets."""
    teams = []
    seen_teams = set()
    
    # First, try to get teams from the Teams sheet
    if 'Teams' in wb.sheetnames:
        ws = wb['Teams']
        # Teams are typically in row 1, starting from column 3 (C)
        for col in range(3, 10):  # Check columns C through I
            cell_value = ws.cell(1, col).value
            if cell_value and isinstance(cell_value, str) and cell_value.strip():
                # Skip if it's a header like "Team Names"
                if cell_value.lower() not in ['team names', 'team name', 'teams', '']:
                    team = cell_value.strip()
                    seen_teams.add(team)
    
    # Check League Standings sheet (rows 17-30 typically have teams)
    if 'League Standings' in wb.sheetnames:
        ws = wb['League Standings']
        for row in range(17, 31):
            cell_value = ws.cell(row, 1).value
            if cell_value and isinstance(cell_value, str) and cell_value.strip():
                team = cell_value.strip()
                if team.lower() not in ['teams', 'team', '']:
                    seen_teams.add(team)
    
    # Also check week sheets to find any additional teams
    week_sheets = [name for name in wb.sheetnames if re.match(r'Week\s+\d+', name, re.IGNORECASE)]
    if week_sheets:
        ws = wb[week_sheets[0]]
        # Look for team names in columns B and D (where teams are listed)
        for row in range(1, 30):
            for col in [2, 4]:  # Columns B and D
                cell_value = ws.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    team = cell_value.strip()
                    # Filter out common non-team values
                    if (team and 
                        team.lower() not in ['ref', 'refs', 'game', 'court', 'bye', '', 'team names'] and
                        not team.startswith('Game') and
                        not team.startswith('Court') and
                        len(team) > 2):
                        seen_teams.add(team)
    
    teams = sorted(list(seen_teams))
    return teams

def detect_week_sheets(wb):
    """Detect all week sheets."""
    week_sheets = []
    for sheet_name in wb.sheetnames:
        # Match patterns like "Week 1 (1.4)", "Week 2 (1.11)", etc.
        if re.match(r'Week\s+\d+', sheet_name, re.IGNORECASE):
            week_sheets.append(sheet_name)
    
    # Sort by week number if possible
    def get_week_num(name):
        match = re.search(r'Week\s+(\d+)', name, re.IGNORECASE)
        return int(match.group(1)) if match else 999
    
    week_sheets.sort(key=get_week_num)
    return week_sheets

def find_win_loss_section(ws):
    """Find where to place the win/loss section (around row 30)."""
    # Look for existing "Team Wins" or similar header
    for row in range(25, 40):
        cell_value = ws.cell(row, 1).value
        if cell_value and isinstance(cell_value, str):
            if 'win' in cell_value.lower() or 'team' in cell_value.lower():
                return row
    
    # Default to row 30
    return 30

def setup_week_sheet(ws, teams, week_name):
    """Set up win/loss formulas for a week sheet."""
    start_row = find_win_loss_section(ws)
    
    # Headers
    ws.cell(start_row, 1).value = 'Team Wins/Losses This Week'
    ws.cell(start_row + 1, 1).value = 'Team Name'
    ws.cell(start_row + 1, 2).value = 'Wins'
    ws.cell(start_row + 1, 3).value = 'Losses'
    
    # Add team names and formulas
    for i, team in enumerate(teams, start=start_row + 2):
        # Team name
        ws.cell(i, 1).value = team
        
        # Wins formula: Sum scores in C when team is in B (Team 1 wins) + Sum scores in E when team is in D (Team 2 wins)
        if ' ' in team or team.endswith('*'):
            # Handle teams with spaces or special characters
            team_pattern = f'"{team}*"' if not team.endswith('*') else f'"{team}"'
        else:
            team_pattern = f'"{team}"'
        
        wins_formula = f'=SUMIFS(C:C,B:B,{team_pattern})+SUMIFS(E:E,D:D,{team_pattern})'
        ws.cell(i, 2).value = wins_formula
        
        # Losses formula: Sum scores in E when team is in B (Team 1 loses, opponent's score) + Sum scores in C when team is in D (Team 2 loses, opponent's score)
        losses_formula = f'=SUMIFS(E:E,B:B,{team_pattern})+SUMIFS(C:C,D:D,{team_pattern})'
        ws.cell(i, 3).value = losses_formula
    
    return start_row + 2  # Return the first data row

def setup_league_standings(wb, teams, week_sheets, week_start_row):
    """Set up the League Standings sheet."""
    if 'League Standings' not in wb.sheetnames:
        # Create it if it doesn't exist
        ws = wb.create_sheet('League Standings')
    else:
        ws = wb['League Standings']
    
    # Clear old team data (rows 17-30)
    for row in range(17, 31):
        for col in range(1, 6):
            ws.cell(row, col).value = None
    
    # Headers
    ws.cell(16, 1).value = 'Teams'
    ws.cell(16, 2).value = 'Wins'
    ws.cell(16, 5).value = 'Losses'  # Column E
    
    # Add team names and aggregate formulas
    for i, team in enumerate(teams, start=17):
        # Team name
        ws.cell(i, 1).value = team
        
        # Wins aggregation (from column B of week sheets)
        wins_row = week_start_row + i - 17  # B32 for row 17, B33 for row 18, etc.
        wins_parts = [f"'{week}'!B{wins_row}" for week in week_sheets]
        ws.cell(i, 2).value = '=' + '+'.join(wins_parts)
        
        # Magic number for tie-breaking
        ws.cell(i, 3).value = f'=B{i}+ROW(B{i})/10000'
        
        # Losses aggregation (from column C of week sheets)
        losses_parts = [f"'{week}'!C{wins_row}" for week in week_sheets]
        ws.cell(i, 5).value = '=' + '+'.join(losses_parts)
    
    # Set up standings display (rows 3-9, but only show as many teams as we have)
    # Clear rows 3-9 first
    for row in range(3, 10):
        for col in range(1, 5):
            ws.cell(row, col).value = None
    
    # Headers for standings display
    ws.cell(2, 1).value = 'Team Name'
    ws.cell(2, 2).value = 'Points For'
    ws.cell(2, 3).value = 'Points Against'
    ws.cell(2, 4).value = 'Point Differential'
    
    # Add standings formulas (only for the number of teams we have)
    num_teams = len(teams)
    for i, rank in enumerate(range(1, num_teams + 1), start=3):
        # Team name
        ws.cell(i, 1).value = f'=INDEX(A17:A{16+num_teams},MATCH(LARGE(C17:C{16+num_teams},{rank}),C17:C{16+num_teams},0))'
        
        # Wins (Points For)
        ws.cell(i, 2).value = f'=INDEX(B17:B{16+num_teams},MATCH(LARGE(C17:C{16+num_teams},{rank}),C17:C{16+num_teams},0))'
        
        # Points Against (Losses)
        ws.cell(i, 3).value = f'=INDEX(E17:E{16+num_teams},MATCH(LARGE(C17:C{16+num_teams},{rank}),C17:C{16+num_teams},0))'
        
        # Point Differential
        ws.cell(i, 4).value = f'=B{i}-C{i}'
    
    # Set week number to 0 if not set
    if ws.cell(11, 2).value is None:
        ws.cell(11, 1).value = 'Week #'
        ws.cell(11, 2).value = 0

def main(file_path):
    """Main function to set up standings for a league spreadsheet."""
    print(f"Loading workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    
    print("\n=== Detecting Teams ===")
    teams = detect_teams(wb)
    print(f"Found {len(teams)} teams: {', '.join(teams)}")
    
    if not teams:
        print("ERROR: No teams detected! Please check the Teams sheet or week sheets.")
        return
    
    print("\n=== Detecting Week Sheets ===")
    week_sheets = detect_week_sheets(wb)
    print(f"Found {len(week_sheets)} week sheets:")
    for week in week_sheets:
        print(f"  - {week}")
    
    if not week_sheets:
        print("ERROR: No week sheets detected!")
        return
    
    print("\n=== Setting Up Week Sheets ===")
    week_start_rows = {}
    for week_name in week_sheets:
        ws = wb[week_name]
        print(f"  Processing {week_name}...")
        start_row = setup_week_sheet(ws, teams, week_name)
        week_start_rows[week_name] = start_row
        print(f"    Added win/loss formulas starting at row {start_row}")
    
    # Use the first week's start row (they should all be the same)
    week_start_row = list(week_start_rows.values())[0] if week_start_rows else 32
    
    print("\n=== Setting Up League Standings ===")
    setup_league_standings(wb, teams, week_sheets, week_start_row)
    print("  League Standings sheet configured")
    
    print("\n=== Enabling Iterative Calculation ===")
    wb.calculation.iterate = True
    wb.calculation.maxIter = 100
    wb.calculation.maxChange = 0.001
    print("  Iterative calculation enabled")
    
    print("\n=== Saving Workbook ===")
    wb.save(file_path)
    print(f"  Saved to {file_path}")
    
    print("\n✅ Setup complete!")
    print(f"\nSummary:")
    print(f"  - {len(teams)} teams configured")
    print(f"  - {len(week_sheets)} week sheets configured")
    print(f"  - League Standings sheet ready")
    print(f"  - All formulas applied")

if __name__ == '__main__':
    import sys
    import os
    
    # Check if arguments provided (non-interactive mode)
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # Interactive mode
        print("📊 Standings Setup")
        print("=" * 50)
        print()
        
        # Find existing spreadsheets
        default_dir = 'public/league_schedules'
        existing_files = []
        if os.path.exists(default_dir):
            existing_files = [f for f in os.listdir(default_dir) if f.endswith('.xlsx')]
        
        if existing_files:
            print("Existing league files:")
            for i, f in enumerate(existing_files, 1):
                print(f"  {i}. {f}")
            print()
        
        file_path = input("Spreadsheet file path: ").strip()
        if not file_path:
            if existing_files:
                file_path = os.path.join(default_dir, existing_files[0])
                print(f"Using: {file_path}")
            else:
                print("❌ No file specified")
                sys.exit(1)
        
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            sys.exit(1)
        
        print()
        print(f"📋 File: {file_path}")
        
        confirm = input("\nSet up standings formulas? (Y/n): ").strip().lower()
        if confirm and confirm != 'y':
            print("Cancelled.")
            sys.exit(0)
        
        print()
    
    main(file_path)

