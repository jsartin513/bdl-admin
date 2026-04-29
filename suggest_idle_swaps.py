"""
Analyze a week sheet for "consecutive two-court rounds without playing" (idle streaks).
Mirrors app/lib/scheduleParser.ts. Tries full-round swaps; optional 2-swap search and ref-flips.
Partial-swap search and --apply-partial reject a team appearing on court 1 and court 2 in the same round.

  python3 suggest_idle_swaps.py --file "public/league_schedules/Spring 2026 BYOT League.xlsx" --sheet "Week 3 Schedule"
  python3 suggest_idle_swaps.py ... --max-games 18   # optional cap
  python3 suggest_idle_swaps.py ... --deep
  python3 suggest_idle_swaps.py ... --deep-partial
  python3 suggest_idle_swaps.py ... --deep-partial --deep-partial-no-final-lock
  python3 suggest_idle_swaps.py ... --ref-flip
  python3 suggest_idle_swaps.py ... --apply-swap 4 14 --write
  python3 suggest_idle_swaps.py ... --apply-partial 10 1 18 2 --write
  python3 suggest_idle_swaps.py ... --move-round-to-front 17 --write
  # Phase D (default): swaps one round from each of two idle streaks (can be same team, two streaks).
  # Two-week flex quadruplets (balanced 2 matchups W5<->W6, then deep-partial idle search each week):
  python3 suggest_idle_swaps.py --file "public/league_schedules/Spring 2026 BYOT League.xlsx" \\
    --sheet "Week 5 Schedule" --sheet2 "Week 6 Schedule" --flex-quadruplets
"""

from __future__ import annotations

import argparse
import copy
from collections import Counter, defaultdict
from dataclasses import dataclass
from itertools import combinations
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from suggest_ref_swaps_week4 import (
    find_consecutive_refs,
    find_consecutive_same_matchup,
    games_after_round_swap,
    parse_week_schedule,
)


def valid_matchup_side(home: str, away: str) -> bool:
    return bool(
        home
        and away
        and home not in ("BYE", "TBD")
        and away not in ("BYE", "TBD")
    )


def round_has_two_courts(g: dict[str, Any]) -> bool:
    h1, a1 = g["court1_playing"]
    h2, a2 = g["court2_playing"]
    return valid_matchup_side(h1, a1) and valid_matchup_side(h2, a2)


def team_is_playing(team: str, g: dict[str, Any]) -> bool:
    slots = list(g["court1_playing"]) + list(g["court2_playing"])
    return any(
        s == team and s and s not in ("BYE", "TBD") for s in slots
    )


def teams_in_week(games: list[dict[str, Any]]) -> list[str]:
    """All teams that appear as players or refs (analogous to schedule stats keys for one week)."""
    seen: set[str] = set()
    for g in games:
        seen |= g["playing"]
        for r in (g["court1Ref"], g["court2Ref"]):
            if r:
                seen.add(r)
    return sorted(seen)


def find_idle_streak_issues(
    games: list[dict[str, Any]], teams: list[str] | None = None
) -> list[dict[str, Any]]:
    """
    Same logic as scheduleParser.ts: two-court rounds only; streak when team not playing;
    flush on play or non-two-court row; record when streak length >= 2.
    Each issue includes slot_indices (0-based positions in `games`) for swap search.
    """
    if teams is None:
        teams = teams_in_week(games)
    issues: list[dict[str, Any]] = []
    for team in teams:
        streak_rounds: list[dict[str, Any]] = []
        streak_slots: list[int] = []
        for idx, g in enumerate(games):
            if not round_has_two_courts(g):
                if len(streak_rounds) >= 2:
                    issues.append(_idle_issue(team, streak_rounds, streak_slots))
                streak_rounds = []
                streak_slots = []
                continue
            if team_is_playing(team, g):
                if len(streak_rounds) >= 2:
                    issues.append(_idle_issue(team, streak_rounds, streak_slots))
                streak_rounds = []
                streak_slots = []
            else:
                streak_rounds.append(g)
                streak_slots.append(idx)
        if len(streak_rounds) >= 2:
            issues.append(_idle_issue(team, streak_rounds, streak_slots))
    return issues


def _idle_issue(
    team: str,
    streak_rounds: list[dict[str, Any]],
    streak_slots: list[int],
) -> dict[str, Any]:
    n = len(streak_rounds)
    game_numbers = [g["gameNumber"] for g in streak_rounds]
    return {
        "team": team,
        "game_numbers": game_numbers,
        "slot_indices": streak_slots[:],
        "end_game": streak_rounds[-1]["gameNumber"],
        "severity": "error" if n >= 3 else "warning",
        "streak_len": n,
    }


def ref_play_conflicts(games: list[dict[str, Any]]) -> list[tuple[str, int, str]]:
    bad: list[tuple[str, int, str]] = []
    for i, g in enumerate(games):
        for court_label, ref in (
            ("court1", g["court1Ref"]),
            ("court2", g["court2Ref"]),
        ):
            if ref and ref in g["playing"]:
                bad.append((ref, i, court_label))
    return bad


def find_same_team_both_courts_issues(
    games: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Same team listed on court 1 and court 2 in one round (invalid for two-court play).
    Uses court cells, not `playing` (which is a set and hides duplicates).
    """
    skip = {"", "BYE", "TBD"}
    issues: list[dict[str, Any]] = []
    for g in games:
        on_c1 = {t for t in g["court1_playing"] if t and t not in skip}
        on_c2 = {t for t in g["court2_playing"] if t and t not in skip}
        for team in sorted(on_c1 & on_c2):
            issues.append({"gameNumber": g["gameNumber"], "team": team})
    return issues


@dataclass(frozen=True)
class Score:
    idle_count: int
    ref_pairs: int  # count of (team, consecutive ref pair) — same as len(find_consecutive_refs)
    same_match_adj: int
    ref_play: int

    def __lt__(self, other: Score) -> bool:
        return (self.idle_count, self.ref_pairs, self.same_match_adj, self.ref_play) < (
            other.idle_count,
            other.ref_pairs,
            other.same_match_adj,
            other.ref_play,
        )


def score_schedule(games: list[dict[str, Any]]) -> Score:
    teams = teams_in_week(games)
    return Score(
        idle_count=len(find_idle_streak_issues(games, teams)),
        ref_pairs=len(find_consecutive_refs(games)),
        same_match_adj=len(find_consecutive_same_matchup(games)),
        ref_play=len(ref_play_conflicts(games)),
    )


def format_score(s: Score) -> str:
    return (
        f"idle={s.idle_count} consecutive_ref_edges={s.ref_pairs} "
        f"same_matchup_adj={s.same_match_adj} ref_play={s.ref_play}"
    )


def games_after_partial_court_swap(
    games: list[dict[str, Any]],
    game_a_1based: int,
    court_a: int,
    game_b_1based: int,
    court_b: int,
) -> list[dict[str, Any]] | None:
    """
    In-memory partial swap: one court's matchup + that court's ref between two rounds.
    Slot labels (gameNumber, row, refRow) unchanged — same as apply_partial_court_swap_to_sheet.
    """
    g = copy.deepcopy(games)
    ia = game_a_1based - 1
    ib = game_b_1based - 1
    if ia == ib and court_a == court_b:
        return None
    if court_a not in (1, 2) or court_b not in (1, 2):
        return None
    if not (0 <= ia < len(g) and 0 <= ib < len(g)):
        return None
    A, B = g[ia], g[ib]
    pa, ra = (
        ("court1_playing", "court1Ref")
        if court_a == 1
        else ("court2_playing", "court2Ref")
    )
    pb, rb = (
        ("court1_playing", "court1Ref")
        if court_b == 1
        else ("court2_playing", "court2Ref")
    )
    ta, t_ra = list(A[pa]), A[ra]
    tb, t_rb = list(B[pb]), B[rb]
    A[pa], A[ra] = (tb[0], tb[1]), t_rb
    B[pb], B[rb] = (ta[0], ta[1]), t_ra
    for x in (A, B):
        h1, a1 = x["court1_playing"]
        h2, a2 = x["court2_playing"]
        x["playing"] = {t for t in [h1, a1, h2, a2] if t}
        x["court1_teams"] = frozenset({h1, a1} - {""})
        x["court2_teams"] = frozenset({h2, a2} - {""})
    return g


def final_round_single_matchup_ok(games: list[dict[str, Any]]) -> bool:
    """Exactly one court has a valid home/away matchup in the last parsed round."""
    if not games:
        return True
    slot = games[-1]
    h1, a1 = slot["court1_playing"]
    h2, a2 = slot["court2_playing"]
    v1 = valid_matchup_side(h1, a1)
    v2 = valid_matchup_side(h2, a2)
    return (v1 and not v2) or (v2 and not v1)


def partial_schedule_blockers(
    games: list[dict[str, Any]],
    teams: list[str],
    *,
    lock_final_single: bool,
) -> list[str]:
    """Human-readable reasons a schedule is rejected by --deep-partial / --apply-partial."""
    reasons: list[str] = []
    dups = find_same_team_both_courts_issues(games)
    if dups:
        reasons.append(
            "same team on both courts in one round ("
            + ", ".join(f"{x['team']} in {x['gameNumber']}" for x in dups)
            + ")"
        )
    if ref_play_conflicts(games):
        reasons.append("referee assigned while also playing that round")
    if any(
        x["severity"] == "error" for x in find_idle_streak_issues(games, teams)
    ):
        reasons.append("3+ round idle streak (error severity)")
    if lock_final_single and not final_round_single_matchup_ok(games):
        reasons.append("last round is not exactly one active matchup")
    return reasons


def schedule_ok_for_deep_partial(
    games: list[dict[str, Any]],
    teams: list[str],
    *,
    lock_final_single: bool,
) -> bool:
    return not partial_schedule_blockers(
        games, teams, lock_final_single=lock_final_single
    )


def partial_swap_move_key(move: tuple[int, int, int, int]) -> frozenset[tuple[int, int]]:
    ga, ca, gb, cb = move
    return frozenset({(ga, ca), (gb, cb)})


def iter_canonical_partial_moves(n: int) -> list[tuple[int, int, int, int]]:
    moves: list[tuple[int, int, int, int]] = []
    for ga in range(1, n + 1):
        for ca in (1, 2):
            for gb in range(1, n + 1):
                for cb in (1, 2):
                    if (ga, ca) >= (gb, cb):
                        continue
                    moves.append((ga, ca, gb, cb))
    return moves


def search_deep_partial_swap_sequences(
    games: list[dict[str, Any]],
    *,
    lock_final_single: bool = True,
    top: int = 30,
) -> tuple[
    list[tuple[Score, tuple[int, int, int, int]]],
    list[tuple[Score, tuple[int, int, int, int], tuple[int, int, int, int]]],
]:
    """
    Try one or two partial court-slot swaps (any order). Intermediate and final schedules must pass
    schedule_ok_for_deep_partial. Return strictly score-improving sequences vs baseline.
    """
    teams = teams_in_week(games)
    baseline = score_schedule(games)
    moves = iter_canonical_partial_moves(len(games))
    singles: list[tuple[Score, tuple[int, int, int, int]]] = []
    doubles: list[
        tuple[Score, tuple[int, int, int, int], tuple[int, int, int, int]]
    ] = []

    for m1 in moves:
        g1 = games_after_partial_court_swap(games, *m1)
        if g1 is None or not schedule_ok_for_deep_partial(
            g1, teams, lock_final_single=lock_final_single
        ):
            continue
        s1 = score_schedule(g1)
        if _better(s1, baseline):
            singles.append((s1, m1))

        k1 = partial_swap_move_key(m1)
        for m2 in moves:
            if partial_swap_move_key(m2) == k1:
                continue
            g2 = games_after_partial_court_swap(g1, *m2)
            if g2 is None or not schedule_ok_for_deep_partial(
                g2, teams, lock_final_single=lock_final_single
            ):
                continue
            s2 = score_schedule(g2)
            if _better(s2, baseline):
                doubles.append((s2, m1, m2))

    def sort_key_s(
        item: tuple[Score, ...],
    ) -> tuple[int, int, int, int]:
        s = item[0]
        return (s.idle_count, s.ref_pairs, s.same_match_adj, s.ref_play)

    singles.sort(key=sort_key_s)
    doubles.sort(key=sort_key_s)
    return singles[:top], doubles[:top]


def format_partial_move(m: tuple[int, int, int, int]) -> str:
    ga, ca, gb, cb = m
    return f"G{ga}c{ca}<->G{gb}c{cb}"


def _flex_pair_key(home: str, away: str) -> tuple[str, str]:
    return tuple(sorted((home, away)))


def apply_cross_week_flex_pairwise_swap(
    games_a: list[dict[str, Any]],
    games_b: list[dict[str, Any]],
    edge_a1: dict[str, Any],
    edge_a2: dict[str, Any],
    edge_b1: dict[str, Any],
    edge_b2: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Exchange two court matchups on `games_a` with two on `games_b` (pairwise).
    Ref assignments per round are unchanged; only home/away on the chosen courts swap across weeks.
    """
    ga = copy.deepcopy(games_a)
    gb = copy.deepcopy(games_b)

    def set_court(round_dict: dict[str, Any], court_key: str, h: str, a: str) -> None:
        round_dict[court_key] = (h, a)

    def refresh_round(gg: list[dict[str, Any]], slot: int) -> None:
        gslot = gg[slot]
        pset: set[str] = set()
        for ck in ("court1_playing", "court2_playing"):
            pset |= {t for t in gslot[ck] if t and t not in ("BYE", "TBD")}
        gslot["playing"] = pset
        h1, a1 = gslot["court1_playing"]
        h2, a2 = gslot["court2_playing"]
        gslot["court1_teams"] = frozenset({h1, a1} - {""})
        gslot["court2_teams"] = frozenset({h2, a2} - {""})

    for g_x, exa, g_y, exb in (
        (ga, edge_a1, gb, edge_b1),
        (ga, edge_a2, gb, edge_b2),
    ):
        rx = g_x[exa["slot"]]
        ry = g_y[exb["slot"]]
        hx, ax = rx[exa["court"]]
        hy, ay = ry[exb["court"]]
        set_court(rx, exa["court"], hy, ay)
        set_court(ry, exb["court"], hx, ax)
        refresh_round(g_x, exa["slot"])
        refresh_round(g_y, exb["slot"])
    return ga, gb


def _edge_dict_for_flat_tuple(
    games: list[dict[str, Any]],
    tup: tuple[Any, ...],
) -> dict[str, Any]:
    """Map movable edge tuple (pair, gameNumber, slot_idx, home, away) to slot + court keys."""
    pair, gn, slot, h, a = tup
    g = games[slot]
    for ck in ("court1_playing", "court2_playing"):
        hh, aa = g[ck]
        if {hh, aa} == {h, a}:
            return {
                "slot": slot,
                "court": ck,
                "pair": pair,
                "h": h,
                "a": a,
                "gameNumber": gn,
            }
    raise ValueError(f"No court for {tup!r} in week")


def best_idle_count_after_deep_partial(
    games: list[dict[str, Any]],
    *,
    lock_final_single: bool = True,
    top: int = 500,
) -> tuple[int, int]:
    """Return (best idle issue count, starting idle count) using up to two partial court swaps."""
    teams = teams_in_week(games)
    start = len(find_idle_streak_issues(games, teams))
    base_sc = score_schedule(games)
    singles, doubles = search_deep_partial_swap_sequences(
        games, lock_final_single=lock_final_single, top=top
    )
    best = base_sc.idle_count
    for s, _m in singles:
        best = min(best, s.idle_count)
    for s, _m1, _m2 in doubles:
        best = min(best, s.idle_count)
    return best, start


def run_flex_quadruplet_idle_report(
    file_path: str,
    sheet_week_a: str,
    sheet_week_b: str,
    *,
    max_games: int | None = None,
    lock_final_single: bool = True,
    deep_top: int = 500,
) -> None:
    """
    List balanced flex quadruplets (two movable H2Hs from week A <-> two from week B, same 4 teams),
    raw idle totals after swap, and best idle per week after search_deep_partial_swap_sequences.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_week_a not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Sheet {sheet_week_a!r} not in {wb.sheetnames}")
    if sheet_week_b not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Sheet {sheet_week_b!r} not in {wb.sheetnames}")

    games_a = parse_week_schedule(wb[sheet_week_a], max_games=max_games)
    games_b = parse_week_schedule(wb[sheet_week_b], max_games=max_games)
    wb.close()

    def collect_edges(games: list[dict[str, Any]]) -> list[tuple[Any, ...]]:
        edges: list[tuple[Any, ...]] = []
        for i, g in enumerate(games):
            for ck in ("court1_playing", "court2_playing"):
                h, a = g[ck]
                if valid_matchup_side(h, a):
                    edges.append((_flex_pair_key(h, a), g["gameNumber"], i, h, a))
        return edges

    e_a = collect_edges(games_a)
    e_b = collect_edges(games_b)
    c_a = Counter(p for p, *_ in e_a)
    c_b = Counter(p for p, *_ in e_b)
    all_pairs = set(c_a) | set(c_b)
    combined = {p: c_a.get(p, 0) + c_b.get(p, 0) for p in all_pairs}
    flexible = {p for p, t in combined.items() if t == 3}

    movable_a = [e for e in e_a if e[0] in flexible and c_a[e[0]] == 2]
    movable_b = [e for e in e_b if e[0] in flexible and c_b[e[0]] == 2]

    flat_a = [("A", e) for e in movable_a]
    flat_b = [("B", e) for e in movable_b]

    def team_incidence(o: tuple[str, tuple[Any, ...]]) -> Counter[str]:
        ctr: Counter[str] = Counter()
        _, e = o
        ctr[e[3]] += 1
        ctr[e[4]] += 1
        return ctr

    balanced: list[tuple[Any, ...]] = []
    for o_pair in combinations(flat_a, 2):
        vo = team_incidence(o_pair[0]) + team_incidence(o_pair[1])
        for i_pair in combinations(flat_b, 2):
            vi = team_incidence(i_pair[0]) + team_incidence(i_pair[1])
            if vo == vi:
                balanced.append((o_pair, i_pair))

    def idle_two_weeks(ga: list[dict[str, Any]], gb: list[dict[str, Any]]) -> int:
        ta = teams_in_week(ga)
        tb = teams_in_week(gb)
        return len(find_idle_streak_issues(ga, ta)) + len(find_idle_streak_issues(gb, tb))

    ba_start = idle_two_weeks(games_a, games_b)
    b_best_a, _ = best_idle_count_after_deep_partial(
        games_a, lock_final_single=lock_final_single, top=deep_top
    )
    b_best_b, _ = best_idle_count_after_deep_partial(
        games_b, lock_final_single=lock_final_single, top=deep_top
    )
    baseline_deep_total = b_best_a + b_best_b

    by_ts: defaultdict[frozenset[str], list[tuple[Any, ...]]] = defaultdict(list)
    for o_pair, i_pair in balanced:
        ts: set[str] = set()
        for _tag, e in o_pair:
            ts |= {e[3], e[4]}
        for _tag, e in i_pair:
            ts |= {e[3], e[4]}
        by_ts[frozenset(ts)].append((o_pair, i_pair))

    print(
        f"Flex quadruplets: {sheet_week_a!r} + {sheet_week_b!r}\n"
        f"  Flexible pairs (3 H2Hs across both weeks): {len(flexible)}\n"
        f"  Balanced quadruplets (2 movable from each week, same 4 teams): {len(balanced)} "
        f"({len(by_ts)} distinct 4-team patterns)"
    )
    print(
        f"\nBaseline idle issues (raw): {ba_start} "
        f"({sheet_week_a}={len(find_idle_streak_issues(games_a, teams_in_week(games_a)))}, "
        f"{sheet_week_b}={len(find_idle_streak_issues(games_b, teams_in_week(games_b)))})"
    )
    print(
        f"Baseline best idle (each week: up to 2 partial court swaps, top={deep_top}): "
        f"combined={baseline_deep_total} ({sheet_week_a}={b_best_a}, {sheet_week_b}={b_best_b})"
    )

    def fmt_edge(tag_e: tuple[str, tuple[Any, ...]]) -> str:
        _, e = tag_e
        return f"{e[1]} {e[3]} vs {e[4]}"

    # Post-swap deep idle is identical for all quadruplets in the same 4-team pattern; one rep per pattern.
    post_swap_best: int | None = None
    print(f"\n--- One line per distinct 4-team pattern ({len(by_ts)} patterns) ---")
    for ts in sorted(by_ts.keys(), key=lambda s: sorted(s)):
        o_pair, i_pair = by_ts[ts][0]
        ea1 = _edge_dict_for_flat_tuple(games_a, o_pair[0][1])
        ea2 = _edge_dict_for_flat_tuple(games_a, o_pair[1][1])
        eb1 = _edge_dict_for_flat_tuple(games_b, i_pair[0][1])
        eb2 = _edge_dict_for_flat_tuple(games_b, i_pair[1][1])
        ng_a, ng_b = apply_cross_week_flex_pairwise_swap(
            games_a, games_b, ea1, ea2, eb1, eb2
        )
        raw_tot = idle_two_weeks(ng_a, ng_b)
        bn_a, sna = best_idle_count_after_deep_partial(
            ng_a, lock_final_single=lock_final_single, top=deep_top
        )
        bn_b, snb = best_idle_count_after_deep_partial(
            ng_b, lock_final_single=lock_final_single, top=deep_top
        )
        comb = bn_a + bn_b
        post_swap_best = comb if post_swap_best is None else min(post_swap_best, comb)
        teams_s = ", ".join(sorted(ts))
        print(f"  Teams: {teams_s}")
        print(
            f"    Out {sheet_week_a}: {fmt_edge(o_pair[0])}; {fmt_edge(o_pair[1])}\n"
            f"    Out {sheet_week_b}: {fmt_edge(i_pair[0])}; {fmt_edge(i_pair[1])}"
        )
        print(
            f"    Raw idle after swap: {raw_tot} | "
            f"best after reorder: {bn_a + bn_b} ({sheet_week_a} {sna}->{bn_a}, {sheet_week_b} {snb}->{bn_b})"
        )

    if post_swap_best is not None:
        print(
            f"\nBest combined idle after quadruplet + deep-partial (min over {len(by_ts)} patterns): "
            f"{post_swap_best} | baseline (no quadruplet) best: {baseline_deep_total}"
        )
        if post_swap_best < baseline_deep_total:
            print("  Cross-week flex improves on baseline deep-partial for this workbook.")
        else:
            print(
                "  Cross-week flex does not beat baseline deep-partial under this search (same or worse)."
            )


def apply_ref_flip_mask(
    games: list[dict[str, Any]], mask: int
) -> list[dict[str, Any]]:
    """If bit i set, swap court1Ref and court2Ref for game i. Game numbers preserved."""
    out = copy.deepcopy(games)
    n = len(out)
    for i in range(n):
        if mask & (1 << i):
            g = out[i]
            g["court1Ref"], g["court2Ref"] = g["court2Ref"], g["court1Ref"]
    return out


def search_ref_flips(games: list[dict[str, Any]]) -> list[tuple[Score, int]]:
    """Try all 2^n within-round ref swaps; same playing set — idle streak count never changes."""
    n = len(games)
    results: list[tuple[Score, int]] = []
    for mask in range(1 << n):
        ng = apply_ref_flip_mask(games, mask)
        s = score_schedule(ng)
        results.append((s, mask))
    results.sort(key=lambda x: (x[0].idle_count, x[0].ref_pairs, x[0].same_match_adj, x[0].ref_play))
    return results


def search_two_round_swaps(games: list[dict[str, Any]], top: int = 25) -> list[tuple[Score, list[tuple[int, int]]]]:
    """Apply up to two pairwise round swaps (composition on original order)."""
    n = len(games)
    baseline = score_schedule(games)
    best: list[tuple[Score, list[tuple[int, int]]]] = []

    pairs: list[tuple[int, int]] = []
    for i in range(n):
        for j in range(i + 1, n):
            pairs.append((i, j))

    for a, b in pairs:
        g1 = games_after_round_swap(games, a, b)
        s1 = score_schedule(g1)
        if _better(s1, baseline):
            best.append((s1, [(a, b)]))

        for c, d in pairs:
            g2 = games_after_round_swap(g1, c, d)
            s2 = score_schedule(g2)
            if _better(s2, baseline):
                seq = [(a, b), (c, d)]
                best.append((s2, seq))

    best.sort(key=lambda x: (x[0].idle_count, x[0].ref_pairs, x[0].same_match_adj, x[0].ref_play))
    return best[:top]


def _better(s: Score, baseline: Score) -> bool:
    return s < baseline


def search_cross_streak_swaps(
    games: list[dict[str, Any]],
    baseline: Score,
    top: int = 40,
) -> list[tuple[Score, str, str, int, int, str, str]]:
    """
    Try swapping one full round from each of two double-break streaks (possibly same team twice).
    If team A sits rounds at slots {i,i+1} and team B sits {j,j+...}, a swap (i,j) moves
    content between those time slots (labels fixed) — can break both streak patterns.
    """
    teams = teams_in_week(games)
    issues = find_idle_streak_issues(games, teams)
    seen: set[tuple[int, int]] = set()
    candidates: list[tuple[Score, str, str, int, int, str, str]] = []
    for p in range(len(issues)):
        for q in range(p + 1, len(issues)):
            ta, tb = issues[p]["team"], issues[q]["team"]
            gna = "/".join(issues[p]["game_numbers"])
            gnb = "/".join(issues[q]["game_numbers"])
            for i in issues[p]["slot_indices"]:
                for j in issues[q]["slot_indices"]:
                    if i == j:
                        continue
                    a, b = (i, j) if i < j else (j, i)
                    if (a, b) in seen:
                        continue
                    seen.add((a, b))
                    ng = games_after_round_swap(games, a, b)
                    s = score_schedule(ng)
                    candidates.append((s, ta, tb, a + 1, b + 1, gna, gnb))
    candidates.sort(
        key=lambda x: (
            x[0].idle_count,
            x[0].ref_pairs,
            x[0].same_match_adj,
            x[0].ref_play,
        )
    )
    improving = [c for c in candidates if c[0] < baseline]
    if improving:
        return improving[:top]
    return candidates[:top]


MAX_COL = 12


def _read_block(ws: Worksheet, game_row: int, ref_row: int) -> list[list[Any]]:
    out = []
    for r in (game_row, ref_row):
        out.append([ws.cell(r, c).value for c in range(1, MAX_COL + 1)])
    return out


def _write_block_to_rows(
    ws: Worksheet,
    game_row: int,
    ref_row: int,
    data: list[list[Any]],
    game_label: str,
) -> None:
    data = copy.deepcopy(data)
    data[0][0] = game_label
    for col in range(1, MAX_COL + 1):
        c_game = ws.cell(game_row, col)
        c_ref = ws.cell(ref_row, col)
        if not isinstance(c_game, MergedCell):
            c_game.value = data[0][col - 1]
        if not isinstance(c_ref, MergedCell):
            c_ref.value = data[1][col - 1]


def shift_following_rounds_up_one_slot(
    ws: Worksheet,
    games: list[dict[str, Any]],
    first_empty_slot_1based: int,
) -> None:
    """
    If slot `first_empty_slot_1based` should receive the content from the next slot, shift each
    following block up by one physical row-pair: slot N+1 content -> slot N, etc. Clears the
    final slot except column A (keeps Game label). Use when e.g. Game 18 is blank but 19–20 have games.
    """
    ie = first_empty_slot_1based - 1
    n = len(games)
    if not (0 <= ie < n):
        raise SystemExit(f"first_empty_slot must be 1..{n}")
    if ie >= n - 1:
        raise SystemExit("Nothing to shift into last slot")
    blocks = [(g["row"], g["refRow"]) for g in games]
    block_data = [_read_block(ws, gr, rr) for gr, rr in blocks]
    labels = [g["gameNumber"] for g in games]
    for dest in range(ie, n - 1):
        src = dest + 1
        gr, rr = blocks[dest]
        _write_block_to_rows(ws, gr, rr, block_data[src], labels[dest])
    gr_last, rr_last = blocks[n - 1]
    for col in range(2, MAX_COL + 1):
        c1 = ws.cell(gr_last, col)
        c2 = ws.cell(rr_last, col)
        if not isinstance(c1, MergedCell):
            c1.value = None
        if not isinstance(c2, MergedCell):
            c2.value = None
    c_a = ws.cell(gr_last, 1)
    if not isinstance(c_a, MergedCell):
        c_a.value = labels[n - 1]


def apply_round_permutation_to_sheet(
    ws: Worksheet,
    games: list[dict[str, Any]],
    perm: list[int],
) -> None:
    """
    Reorder round blocks. perm[dest] = 0-based index of the source slot whose content
    moves into destination `dest`. Sheet row pairs stay fixed; Game 01.. labels unchanged.
    """
    n = len(games)
    if len(perm) != n or set(perm) != set(range(n)):
        raise SystemExit("perm must be a permutation of 0..n-1 for all parsed games")
    blocks = [(g["row"], g["refRow"]) for g in games]
    block_data = [_read_block(ws, gr, rr) for gr, rr in blocks]
    labels = [g["gameNumber"] for g in games]
    for dest, (game_row, ref_row) in enumerate(blocks):
        src = perm[dest]
        _write_block_to_rows(ws, game_row, ref_row, block_data[src], labels[dest])


def move_round_to_front_of_sheet(
    ws: Worksheet,
    games: list[dict[str, Any]],
    round_1based: int,
) -> None:
    """Put `round_1based`'s content in slot 1; former 1..round-1 shift to 2..round; rest unchanged."""
    idx = round_1based - 1
    n = len(games)
    if not 0 <= idx < n:
        raise SystemExit(f"round must be 1..{n}, got {round_1based}")
    perm = [idx] + list(range(0, idx)) + list(range(idx + 1, n))
    apply_round_permutation_to_sheet(ws, games, perm)


def apply_round_swap_to_sheet(
    ws: Worksheet,
    games: list[dict[str, Any]],
    slot_a_1based: int,
    slot_b_1based: int,
) -> None:
    """
    Swap full round content (game + ref rows) between two slots. Keeps each row's Game NN
    label in column A to match in-memory games_after_round_swap semantics.
    """
    i = slot_a_1based - 1
    j = slot_b_1based - 1
    n = len(games)
    if not (0 <= i < n and 0 <= j < n):
        raise SystemExit(f"Swap slots must be 1..{n}, got {slot_a_1based} and {slot_b_1based}")
    if i == j:
        return
    blocks = [(g["row"], g["refRow"]) for g in games]
    block_data = [_read_block(ws, gr, rr) for gr, rr in blocks]
    block_data[i], block_data[j] = block_data[j], block_data[i]
    labels = [g["gameNumber"] for g in games]
    for k, (game_row, ref_row) in enumerate(blocks):
        data = block_data[k]
        data[0][0] = labels[k]
        for col in range(1, MAX_COL + 1):
            c_game = ws.cell(game_row, col)
            c_ref = ws.cell(ref_row, col)
            if not isinstance(c_game, MergedCell):
                c_game.value = data[0][col - 1]
            if not isinstance(c_ref, MergedCell):
                c_ref.value = data[1][col - 1]


def _court_cols(court: int) -> tuple[int, int, int]:
    """Return (home_col, away_col, ref_col) for court 1 or 2."""
    if court == 1:
        return (2, 4, 2)
    if court == 2:
        return (7, 9, 7)
    raise SystemExit("court must be 1 or 2")


def apply_partial_court_swap_to_sheet(
    ws: Worksheet,
    games: list[dict[str, Any]],
    game_a_1based: int,
    court_a: int,
    game_b_1based: int,
    court_b: int,
) -> None:
    """
    Swap one court's matchup (home/away on game row) and that court's ref (on ref row)
    between two rounds. court is 1 or 2 (matches sheet layout from parse_week_schedule).
    """
    ia = game_a_1based - 1
    ib = game_b_1based - 1
    n = len(games)
    if not (0 <= ia < n and 0 <= ib < n):
        raise SystemExit(f"Game numbers must be 1..{n}")
    if ia == ib and court_a == court_b:
        raise SystemExit("Need two different court slots to swap")

    gr_a, rr_a = games[ia]["row"], games[ia]["refRow"]
    gr_b, rr_b = games[ib]["row"], games[ib]["refRow"]
    h_a, aa_a, r_a = _court_cols(court_a)
    h_b, aa_b, r_b = _court_cols(court_b)

    def write_cell(row: int, col: int, val: Any) -> None:
        c = ws.cell(row, col)
        if not isinstance(c, MergedCell):
            c.value = val

    vals_a = [
        ws.cell(gr_a, h_a).value,
        ws.cell(gr_a, aa_a).value,
        ws.cell(rr_a, r_a).value,
    ]
    vals_b = [
        ws.cell(gr_b, h_b).value,
        ws.cell(gr_b, aa_b).value,
        ws.cell(rr_b, r_b).value,
    ]

    write_cell(gr_a, h_a, vals_b[0])
    write_cell(gr_a, aa_a, vals_b[1])
    write_cell(rr_a, r_a, vals_b[2])
    write_cell(gr_b, h_b, vals_a[0])
    write_cell(gr_b, aa_b, vals_a[1])
    write_cell(rr_b, r_b, vals_a[2])


def apply_cross_sheet_court_swap_to_sheets(
    ws_a: Worksheet,
    games_a: list[dict[str, Any]],
    game_a_1based: int,
    court_a: int,
    ws_b: Worksheet,
    games_b: list[dict[str, Any]],
    game_b_1based: int,
    court_b: int,
) -> None:
    """
    Swap one court's matchup + ref on sheet A with one court on sheet B (different weeks).
    `court` is 1 or 2; game numbers are 1-based slot indices (Game 01 = 1).
    """
    ia = game_a_1based - 1
    ib = game_b_1based - 1
    na, nb = len(games_a), len(games_b)
    if not (0 <= ia < na and 0 <= ib < nb):
        raise SystemExit(
            f"Cross-sheet swap: games out of range (A has {na} rounds, B has {nb})"
        )
    if court_a not in (1, 2) or court_b not in (1, 2):
        raise SystemExit("court must be 1 or 2")

    gr_a, rr_a = games_a[ia]["row"], games_a[ia]["refRow"]
    gr_b, rr_b = games_b[ib]["row"], games_b[ib]["refRow"]
    h_a, aa_a, r_a = _court_cols(court_a)
    h_b, aa_b, r_b = _court_cols(court_b)

    def write_cell(ws: Worksheet, row: int, col: int, val: Any) -> None:
        c = ws.cell(row, col)
        if not isinstance(c, MergedCell):
            c.value = val

    vals_a = [
        ws_a.cell(gr_a, h_a).value,
        ws_a.cell(gr_a, aa_a).value,
        ws_a.cell(rr_a, r_a).value,
    ]
    vals_b = [
        ws_b.cell(gr_b, h_b).value,
        ws_b.cell(gr_b, aa_b).value,
        ws_b.cell(rr_b, r_b).value,
    ]

    write_cell(ws_a, gr_a, h_a, vals_b[0])
    write_cell(ws_a, gr_a, aa_a, vals_b[1])
    write_cell(ws_a, rr_a, r_a, vals_b[2])
    write_cell(ws_b, gr_b, h_b, vals_a[0])
    write_cell(ws_b, gr_b, aa_b, vals_a[1])
    write_cell(ws_b, rr_b, r_b, vals_a[2])


def main() -> None:
    parser = argparse.ArgumentParser(description="Idle streak analysis and swap exploration.")
    parser.add_argument("--file", required=True, help="Path to .xlsx")
    parser.add_argument("--sheet", required=True, help='Sheet name, e.g. "Week 3 Schedule"')
    parser.add_argument(
        "--deep",
        action="store_true",
        help="Search sequences of up to two round-swaps that improve the score vs baseline.",
    )
    parser.add_argument(
        "--deep-partial",
        action="store_true",
        help=(
            "Search one- or two-step partial court-slot swaps (matchup+ref per court) that strictly "
            "improve the score; by default the last round must stay a single-court matchup."
        ),
    )
    parser.add_argument(
        "--deep-partial-no-final-lock",
        action="store_true",
        help="With --deep-partial, allow any last-round court pattern (not only one active matchup).",
    )
    parser.add_argument(
        "--ref-flip",
        action="store_true",
        help="Exhaustive ref swap per round (court1Ref↔court2Ref); idle count unchanged; may reduce ref_pairs.",
    )
    parser.add_argument(
        "--max-games",
        type=int,
        default=None,
        metavar="N",
        help="Parse at most N rounds (default: all games on sheet up to row 200).",
    )
    parser.add_argument(
        "--no-cross-streak",
        action="store_true",
        help="Skip cross double-break swap search (different teams' idle streaks).",
    )
    parser.add_argument(
        "--apply-swap",
        nargs=2,
        type=int,
        metavar=("SLOT_A", "SLOT_B"),
        help="1-based round slots to swap (full game+ref blocks). Column A keeps Game 01.. labels.",
    )
    parser.add_argument(
        "--write",
        action="store_true",
        help="Save after any sheet mutation flag (no data_only load).",
    )
    parser.add_argument(
        "--apply-partial",
        nargs=4,
        type=int,
        metavar=("GAME_A", "COURT_A", "GAME_B", "COURT_B"),
        help="Swap court COURT_A of GAME_A with court COURT_B of GAME_B (1-based game nums, courts 1|2).",
    )
    parser.add_argument(
        "--shift-up-from",
        type=int,
        metavar="SLOT",
        help="1-based first empty slot: copy slot N+1..end each up one row-pair; clear last slot (keep label).",
    )
    parser.add_argument(
        "--move-round-to-front",
        type=int,
        metavar="N",
        help="Put round N's content in slot 1; former 1..N-1 shift to slots 2..N; N+1..end unchanged.",
    )
    parser.add_argument(
        "--flex-quadruplets",
        action="store_true",
        help=(
            "With --sheet2: score balanced flex quadruplets (two movable H2Hs from each week, same 4 teams) "
            "and run the same deep-partial idle search as --deep-partial on each week before/after (no write)."
        ),
    )
    parser.add_argument(
        "--sheet2",
        default=None,
        metavar="NAME",
        help='Second week sheet (required with --flex-quadruplets), e.g. "Week 6 Schedule".',
    )
    parser.add_argument(
        "--flex-deep-top",
        type=int,
        default=500,
        metavar="N",
        help="With --flex-quadruplets: top-N improving moves to keep in search_deep_partial (default 500).",
    )
    args = parser.parse_args()

    if args.flex_quadruplets:
        if not args.sheet2:
            raise SystemExit("--flex-quadruplets requires --sheet2")
        if args.write or args.apply_swap or args.apply_partial or args.shift_up_from is not None or args.move_round_to_front is not None:
            raise SystemExit("--flex-quadruplets cannot be combined with sheet mutation flags")
        run_flex_quadruplet_idle_report(
            args.file,
            args.sheet,
            args.sheet2,
            max_games=args.max_games,
            lock_final_single=not args.deep_partial_no_final_lock,
            deep_top=args.flex_deep_top,
        )
        return

    has_apply = bool(
        args.apply_swap
        or args.apply_partial
        or args.shift_up_from is not None
        or args.move_round_to_front is not None
    )
    if args.write and not has_apply:
        raise SystemExit(
            "--write requires --apply-swap, --apply-partial, --shift-up-from, or --move-round-to-front"
        )

    mode_count = sum(
        1
        for flag in (
            args.apply_swap,
            args.apply_partial,
            args.shift_up_from is not None,
            args.move_round_to_front is not None,
        )
        if flag
    )
    if mode_count > 1:
        raise SystemExit(
            "Use only one of --apply-swap, --apply-partial, --shift-up-from, --move-round-to-front per run"
        )

    if args.apply_swap:
        if not args.write:
            raise SystemExit(
                "Refusing to modify the workbook without --write. "
                "Example: --apply-swap 4 14 --write"
            )
        wb = openpyxl.load_workbook(args.file, data_only=False)
        if args.sheet not in wb.sheetnames:
            raise SystemExit(f"Sheet {args.sheet!r} not in {wb.sheetnames}")
        ws = wb[args.sheet]
        games = parse_week_schedule(ws, max_games=args.max_games)
        a, b = args.apply_swap
        print(f"Applying swap: slot {a} <-> slot {b}")
        apply_round_swap_to_sheet(ws, games, a, b)
        wb.save(args.file)
        wb.close()
        print(f"Saved {args.file}")

        wb = openpyxl.load_workbook(args.file, data_only=True)
        ws = wb[args.sheet]
        games = parse_week_schedule(ws, max_games=args.max_games)
        wb.close()
        teams = teams_in_week(games)
        after = score_schedule(games)
        idle_after = find_idle_streak_issues(games, teams)
        print(f"Rounds parsed: {len(games)}\nAfter swap: {format_score(after)}")
        print(f"\nIdle streak issues ({len(idle_after)}):")
        for issue in idle_after:
            gn = ", ".join(issue["game_numbers"])
            print(f"  {issue['team']}: {issue['streak_len']} rounds ({gn}) [{issue['severity']}]")
        ref_after = find_consecutive_refs(games)
        if ref_after:
            print("\nConsecutive referee assignments:")
            for team, i, j in ref_after:
                print(f"  {team}: {games[i]['gameNumber']} → {games[j]['gameNumber']}")
        else:
            print("\nNo consecutive referee assignments.")
        return

    if args.apply_partial:
        if not args.write:
            raise SystemExit(
                "Refusing to modify without --write. "
                "Example: --apply-partial 10 1 18 2 --write"
            )
        ga, ca, gb, cb = args.apply_partial
        if ca not in (1, 2) or cb not in (1, 2):
            raise SystemExit("Courts must be 1 or 2")
        wb = openpyxl.load_workbook(args.file, data_only=False)
        if args.sheet not in wb.sheetnames:
            raise SystemExit(f"Sheet {args.sheet!r} not in {wb.sheetnames}")
        ws = wb[args.sheet]
        games = parse_week_schedule(ws, max_games=args.max_games)
        teams = teams_in_week(games)
        ng = games_after_partial_court_swap(games, ga, ca, gb, cb)
        if ng is None:
            raise SystemExit("Invalid partial swap (need two different court slots).")
        blockers = partial_schedule_blockers(
            ng, teams, lock_final_single=True
        )
        if blockers:
            raise SystemExit(
                "Refusing partial swap — " + "; ".join(blockers) + "."
            )
        print(f"Applying partial swap: Game {ga} court {ca} <-> Game {gb} court {cb}")
        apply_partial_court_swap_to_sheet(ws, games, ga, ca, gb, cb)
        wb.save(args.file)
        wb.close()
        print(f"Saved {args.file}")

        wb = openpyxl.load_workbook(args.file, data_only=True)
        ws = wb[args.sheet]
        games = parse_week_schedule(ws, max_games=args.max_games)
        wb.close()
        teams = teams_in_week(games)
        after = score_schedule(games)
        idle_after = find_idle_streak_issues(games, teams)
        print(f"Rounds parsed: {len(games)}\nAfter partial swap: {format_score(after)}")
        print(f"\nIdle streak issues ({len(idle_after)}):")
        for issue in idle_after:
            gn = ", ".join(issue["game_numbers"])
            print(f"  {issue['team']}: {issue['streak_len']} rounds ({gn}) [{issue['severity']}]")
        ref_after = find_consecutive_refs(games)
        if ref_after:
            print("\nConsecutive referee assignments:")
            for team, i, j in ref_after:
                print(f"  {team}: {games[i]['gameNumber']} → {games[j]['gameNumber']}")
        else:
            print("\nNo consecutive referee assignments.")
        return

    if args.move_round_to_front is not None:
        if not args.write:
            raise SystemExit(
                "Refusing to modify without --write. "
                "Example: --move-round-to-front 17 --write"
            )
        wb = openpyxl.load_workbook(args.file, data_only=False)
        if args.sheet not in wb.sheetnames:
            raise SystemExit(f"Sheet {args.sheet!r} not in {wb.sheetnames}")
        ws = wb[args.sheet]
        games = parse_week_schedule(ws, max_games=args.max_games)
        r = args.move_round_to_front
        print(
            f"Moving round {r} to front: new Game 01 gets former Game {r:02d} content; "
            f"former Games 01..{r - 1:02d} shift to 02..{r:02d}; Games {r + 1}.. unchanged in slots."
        )
        move_round_to_front_of_sheet(ws, games, r)
        wb.save(args.file)
        wb.close()
        print(f"Saved {args.file}")

        wb = openpyxl.load_workbook(args.file, data_only=True)
        ws = wb[args.sheet]
        games = parse_week_schedule(ws, max_games=args.max_games)
        wb.close()
        teams = teams_in_week(games)
        after = score_schedule(games)
        idle_after = find_idle_streak_issues(games, teams)
        print(f"Rounds parsed: {len(games)}\nAfter move-to-front: {format_score(after)}")
        print(f"\nIdle streak issues ({len(idle_after)}):")
        for issue in idle_after:
            gn = ", ".join(issue["game_numbers"])
            print(f"  {issue['team']}: {issue['streak_len']} rounds ({gn}) [{issue['severity']}]")
        ref_after = find_consecutive_refs(games)
        if ref_after:
            print("\nConsecutive referee assignments:")
            for team, i, j in ref_after:
                print(f"  {team}: {games[i]['gameNumber']} → {games[j]['gameNumber']}")
        else:
            print("\nNo consecutive referee assignments.")
        return

    if args.shift_up_from is not None:
        if not args.write:
            raise SystemExit(
                "Refusing to modify without --write. "
                "Example: --shift-up-from 18 --write"
            )
        wb = openpyxl.load_workbook(args.file, data_only=False)
        if args.sheet not in wb.sheetnames:
            raise SystemExit(f"Sheet {args.sheet!r} not in {wb.sheetnames}")
        ws = wb[args.sheet]
        games = parse_week_schedule(ws, max_games=args.max_games)
        print(
            f"Shifting rounds up from empty slot {args.shift_up_from}: "
            f"content moves to fill gap; last slot cleared."
        )
        shift_following_rounds_up_one_slot(ws, games, args.shift_up_from)
        wb.save(args.file)
        wb.close()
        print(f"Saved {args.file}")

        wb = openpyxl.load_workbook(args.file, data_only=True)
        ws = wb[args.sheet]
        games = parse_week_schedule(ws, max_games=args.max_games)
        wb.close()
        teams = teams_in_week(games)
        after = score_schedule(games)
        idle_after = find_idle_streak_issues(games, teams)
        print(f"Rounds parsed: {len(games)}\nAfter shift: {format_score(after)}")
        print(f"\nIdle streak issues ({len(idle_after)}):")
        for issue in idle_after:
            gn = ", ".join(issue["game_numbers"])
            print(f"  {issue['team']}: {issue['streak_len']} rounds ({gn}) [{issue['severity']}]")
        ref_after = find_consecutive_refs(games)
        if ref_after:
            print("\nConsecutive referee assignments:")
            for team, i, j in ref_after:
                print(f"  {team}: {games[i]['gameNumber']} → {games[j]['gameNumber']}")
        else:
            print("\nNo consecutive referee assignments.")
        return

    wb = openpyxl.load_workbook(args.file, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet {args.sheet!r} not in {wb.sheetnames}")
    ws = wb[args.sheet]
    games = parse_week_schedule(ws, max_games=args.max_games)
    wb.close()

    print(f"Rounds parsed: {len(games)}")

    teams = teams_in_week(games)
    baseline = score_schedule(games)
    idle_issues = find_idle_streak_issues(games, teams)

    print(f"File: {args.file}\nSheet: {args.sheet}\nBaseline: {format_score(baseline)}")
    print(f"\nIdle streak issues ({len(idle_issues)}):")
    for issue in idle_issues:
        gn = ", ".join(issue["game_numbers"])
        slots = issue["slot_indices"]
        print(
            f"  {issue['team']}: {issue['streak_len']} rounds ({gn}) "
            f"slots {[s + 1 for s in slots]} [{issue['severity']}]"
        )

    ref_issues = find_consecutive_refs(games)
    if ref_issues:
        print("\nConsecutive referee assignments:")
        for team, i, j in ref_issues:
            print(
                f"  {team}: {games[i]['gameNumber']} → {games[j]['gameNumber']}"
            )

    dup_court = find_same_team_both_courts_issues(games)
    if dup_court:
        print("\nSame team on both courts in one round:")
        for x in dup_court:
            print(f"  {x['team']}: {x['gameNumber']}")

    sm = find_consecutive_same_matchup(games)
    if sm:
        print("\nSame matchup in consecutive games:")
        for i, j, m in sm:
            print(f"  games {i + 1}/{j + 1}: {tuple(m)}")

    rp = ref_play_conflicts(games)
    if rp:
        print("\nRef also playing:")
        for ref, i, court in rp:
            print(f"  {ref} in {games[i]['gameNumber']} ({court})")

    print("\n--- Phase A: adjacent round swaps (swap slots i and i+1) ---")
    for i in range(len(games) - 1):
        ng = games_after_round_swap(games, i, i + 1)
        s = score_schedule(ng)
        delta = _delta_str(baseline, s)
        print(
            f"  Swap {games[i]['gameNumber']} <-> {games[i + 1]['gameNumber']}: "
            f"{format_score(s)}  ({delta})"
        )

    print("\n--- Phase B: all pairwise round swaps (improving or equal-best idle) ---")
    improved: list[tuple[Score, int, int]] = []
    n = len(games)
    for i in range(n):
        for j in range(i + 1, n):
            ng = games_after_round_swap(games, i, j)
            s = score_schedule(ng)
            if s < baseline:
                improved.append((s, i + 1, j + 1))
    improved.sort(key=lambda x: (x[0].idle_count, x[0].ref_pairs, x[0].same_match_adj, x[0].ref_play))
    if not improved:
        print("  None strictly better than baseline.")
        neutral = []
        for i in range(n):
            for j in range(i + 1, n):
                ng = games_after_round_swap(games, i, j)
                s = score_schedule(ng)
                if s.idle_count == baseline.idle_count and s == baseline:
                    neutral.append((i + 1, j + 1))
        if neutral:
            print(f"  {len(neutral)} swap(s) tied baseline exactly (not listed).")
    else:
        for s, a, b in improved[:30]:
            print(f"  Swap Game slots {a:02d} & {b:02d}: {format_score(s)}")

    if not args.no_cross_streak and len(idle_issues) >= 2:
        print(
            "\n--- Phase D: cross double-break swaps "
            "(swap one round from each of two idle streaks; teams may match if streaks differ) ---"
        )
        cross = search_cross_streak_swaps(games, baseline, top=35)
        if not cross:
            print("  No pairs of distinct-team streaks to combine.")
        else:
            for s, ta, tb, sa, sb, gna, gnb in cross:
                mark = " *" if s < baseline else ""
                print(
                    f"  slots {sa:02d} & {sb:02d} ({ta} [{gna}] vs {tb} [{gnb}]): "
                    f"{format_score(s)}{mark}"
                )
            if any(c[0] < baseline for c in cross):
                print("  (* = strictly better than baseline)")
            else:
                print(
                    "  No cross-streak single swap beats baseline; options above are best by score order."
                )

    if args.deep:
        print("\n--- Phase C (--deep): up to two round-swaps, strictly better than baseline ---")
        top = search_two_round_swaps(games, top=30)
        if not top:
            print("  No improving sequence found.")
        else:
            for s, seq in top:
                desc = " then ".join(
                    f"swap({a + 1},{b + 1})" for a, b in seq
                )
                print(f"  {desc} -> {format_score(s)}")

    if args.deep_partial:
        lock_final = not args.deep_partial_no_final_lock
        print(
            "\n--- --deep-partial: partial court-slot swaps, strictly better than baseline ---"
        )
        print(
            "  Filters: no same team on both courts, no ref-play, no error idle streaks; "
            f"last round single-court matchup: {'required' if lock_final else 'off'}"
        )
        singles, doubles = search_deep_partial_swap_sequences(
            games, lock_final_single=lock_final, top=30
        )
        if singles:
            print(f"  One partial swap ({len(singles)} improving move(s), showing up to 30):")
            for s, m in singles:
                print(f"    {format_partial_move(m)} -> {format_score(s)}")
        else:
            print("  No single partial swap improves the score under these filters.")
        if doubles:
            print(
                f"  Two partial swaps ({len(doubles)} improving sequence(s), showing up to 30):"
            )
            for s, m1, m2 in doubles:
                print(
                    f"    {format_partial_move(m1)} then {format_partial_move(m2)} -> "
                    f"{format_score(s)}"
                )
        else:
            print(
                "  No two-step partial swap improves the score under these filters."
            )

    if args.ref_flip:
        print("\n--- Ref flips (court1Ref ↔ court2Ref per selected rounds; idle unchanged) ---")
        if len(games) > 22:
            print(
                f"  Skipped: {len(games)} games → 2^n exhaustive ref-flip search too large "
                f"(max 22)."
            )
        else:
            ranked = search_ref_flips(games)
            best = ranked[0]
            print(f"  Best mask={best[1]:#x}: {format_score(best[0])}")
            if best[0] < baseline:
                print("  Improves non-idle metrics vs baseline.")
            else:
                print("  No improvement vs baseline on ref_pairs/same_match/ref_play.")
            base_rp = baseline.ref_pairs
            better_rp = [x for x in ranked if x[0].ref_pairs < base_rp][:10]
            if better_rp:
                print(f"  Top masks lowering consecutive_ref_edges below {base_rp}:")
                for s, mask in better_rp:
                    print(f"    mask={mask:#x} {format_score(s)}")


def _delta_str(before: Score, after: Score) -> str:
    parts = []
    if after.idle_count != before.idle_count:
        parts.append(f"idle {before.idle_count}->{after.idle_count}")
    if after.ref_pairs != before.ref_pairs:
        parts.append(f"refs {before.ref_pairs}->{after.ref_pairs}")
    if after.same_match_adj != before.same_match_adj:
        parts.append(f"sm {before.same_match_adj}->{after.same_match_adj}")
    if after.ref_play != before.ref_play:
        parts.append(f"rp {before.ref_play}->{after.ref_play}")
    return ", ".join(parts) if parts else "no change"


if __name__ == "__main__":
    main()
