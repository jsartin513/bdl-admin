#!/usr/bin/env python3
"""
Generate print-ready Open Gym schedule variants from source xlsx.

Reads:  Open Gym Schedules.xlsx  (repo root by default)
Writes: Open Gym Schedules (Print).xlsx  (repo root by default)

Sheets produced:
  2 team          – pass-through
  3 team          – pass-through
  4 team - Long   – 3 sections × 8 games; section 3 home/away corrected
  4 team - Short  – 3 sections × 4 games
  5 team - Long   – 3 sections × 10 games
  5 team - Short  – 3 sections × 5 games
  6 team - waves     – same data interleaved by wave (Rd across all 5 sections)
  6 team - by section – contiguous sections (original ordering)

Non–6-team sheets use columns: Overall, Round, Home, Away, Winner.

6-team **waves** sheet: Overall, Section, Round, Home, Away, Winner.

6-team **by section** sheet: Overall, Round, Home, Away, Winner (same as other rotations).

Usage:
  python3 open_gym_schedules.py
  python3 open_gym_schedules.py --input path/to/in.xlsx --output path/to/out.xlsx
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

HERE = Path(__file__).resolve().parent
DEFAULT_IN = HERE / "Open Gym Schedules.xlsx"
DEFAULT_OUT = HERE / "Open Gym Schedules (Print).xlsx"

# ---------------------------------------------------------------------------
# Styles — colors and fonts consistent with other BDL admin Excel outputs
# ---------------------------------------------------------------------------

FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_SECTION = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
FILL_WINNER_COL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_SECTION = Font(bold=True, size=10)
FONT_BODY = Font(size=10)

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

NUM_COLS = 5

COL_WIDTHS = {
    1: 9,   # Overall round #
    2: 8,   # Section round label (Rd n)
    3: 14,
    4: 14,
    5: 12,  # Winner
}

NUM_COLS_SIX_TEAM = 6

COL_WIDTHS_SIX_TEAM = {
    1: 9,   # Overall game #
    2: 9,   # Rotation section (1–5)
    3: 8,   # Rd within section (Rd 1 … Rd 6)
    4: 14,
    5: 14,
    6: 12,  # Winner
}

NUM_SECTIONS_SIX_TEAM = 5
ROUNDS_PER_SECTION_SIX_TEAM = 6


def style_merged_banner_row(
    ws,
    row: int,
    start_col: int,
    end_col: int,
    *,
    title: str,
    font: Font,
    fill: PatternFill,
) -> None:
    """Merge a title row and draw a perimeter border across the span."""
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    empty_side = Side()
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if col == start_col:
            cell.value = title
            cell.font = font
            cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = Border(
            left=THIN if col == start_col else empty_side,
            right=THIN if col == end_col else empty_side,
            top=THIN,
            bottom=THIN,
        )


@dataclass
class Game:
    round_label: str
    home: str
    away: str


@dataclass
class Section:
    games: list[Game]


def section_take_first_n(section: Section, n: int) -> Section:
    return Section(games=list(section.games[:n]))


def _is_column_header_row(round_val: object, home_val: object) -> bool:
    if round_val is not None:
        return False
    label = (home_val or "").strip()
    return label == "Home"


def parse_sections(ws) -> list[Section]:
    """Split a worksheet into sections using repeated column-header rows.

    A new section begins whenever we see the standard header row (blank round cell,
    ``Home`` label in column B). Blank spacer rows are ignored and do not end a section.
    """
    sections: list[Section] = []
    current: list[Game] = []

    for row in ws.iter_rows(values_only=True):
        round_val = row[0]
        home_val = row[1] if len(row) > 1 else None
        away_val = row[2] if len(row) > 2 else None

        if _is_column_header_row(round_val, home_val):
            if current:
                sections.append(Section(games=current))
                current = []
            continue

        if round_val is None or home_val is None or away_val is None:
            continue

        rnd = str(round_val).strip()
        home = str(home_val).strip()
        away = str(away_val).strip()
        if not rnd or not home or not away:
            continue

        current.append(Game(round_label=rnd, home=home, away=away))

    if current:
        sections.append(Section(games=current))

    return sections


def fix_4team_section3(section: Section) -> Section:
    """
    Original section 3 repeated Team 4 vs Team 2 / Team 3 vs Team 1 without
    swapping home/away. Alternate: 4v2, 3v1, 2v4, 1v3 (repeating).
    """
    pattern = [
        ("Team 4", "Team 2"),
        ("Team 3", "Team 1"),
        ("Team 2", "Team 4"),
        ("Team 1", "Team 3"),
    ]
    fixed = []
    for i, game in enumerate(section.games):
        h, a = pattern[i % len(pattern)]
        fixed.append(Game(round_label=game.round_label, home=h, away=a))
    return Section(games=fixed)


def fix_6team_sections(sections: list[Section]) -> list[Section]:
    """Fix known typos in the 6-team sheet."""
    out: list[Section] = []
    last_idx = len(sections) - 1
    for s_idx, section in enumerate(sections):
        games: list[Game] = []
        for game in section.games:
            home = game.home
            away = game.away
            if home.lower() == "team 3":
                home = "Team 3"
            if away.lower() == "team 3":
                away = "Team 3"
            if s_idx == last_idx and home == "Team 3" and away == "Team 3":
                away = "Team 2"
            games.append(Game(round_label=game.round_label, home=home, away=away))
        out.append(Section(games=games))
    return out


def six_team_sections_to_waves(sections: list[Section]) -> list[list[tuple[int, Game]]]:
    """
    Interleave the five rotation sections so each global “wave” runs one in-section
    round across all sections (Wave 1 = all Rd 1, …, Wave 6 = all Rd 6).

    Each inner list has five entries: (section_number 1–5, Game).
    """
    if len(sections) != NUM_SECTIONS_SIX_TEAM:
        raise ValueError(
            f"Expected {NUM_SECTIONS_SIX_TEAM} sections for 6-team rotation, got {len(sections)}"
        )
    for i, sec in enumerate(sections, start=1):
        if len(sec.games) != ROUNDS_PER_SECTION_SIX_TEAM:
            raise ValueError(
                f"Section {i}: expected {ROUNDS_PER_SECTION_SIX_TEAM} games, got {len(sec.games)}"
            )

    waves: list[list[tuple[int, Game]]] = []
    for round_idx in range(ROUNDS_PER_SECTION_SIX_TEAM):
        wave: list[tuple[int, Game]] = []
        for sec_idx, sec in enumerate(sections, start=1):
            wave.append((sec_idx, sec.games[round_idx]))
        waves.append(wave)
    return waves


def write_six_team_wave_sheet(wb: Workbook, sheet_name: str, sections: list[Section]) -> None:
    """Print layout for 6-team: waves interleave sections so opponents meet sooner."""
    ws = wb.create_sheet(title=sheet_name)
    waves = six_team_sections_to_waves(sections)

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    for col_idx, width in COL_WIDTHS_SIX_TEAM.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    current_row = 1
    overall_round = 1

    for wave_idx, wave in enumerate(waves, start=1):
        style_merged_banner_row(
            ws,
            current_row,
            1,
            NUM_COLS_SIX_TEAM,
            title=f"Wave {wave_idx}",
            font=FONT_SECTION,
            fill=FILL_SECTION,
        )
        current_row += 1

        headers = ["Overall", "Section", "Round", "Home", "Away", "Winner"]
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=label)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER
        current_row += 1

        for g_idx, (section_num, game) in enumerate(wave):
            row_fill = FILL_ROW_ALT if g_idx % 2 else None
            row_vals = [overall_round, section_num, game.round_label, game.home, game.away, None]
            overall_round += 1
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = FONT_BODY
                cell.border = BORDER
                if col_idx == NUM_COLS_SIX_TEAM:
                    cell.fill = FILL_WINNER_COL
                elif row_fill:
                    cell.fill = row_fill
                if col_idx in (4, 5):
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_CENTER
            current_row += 1

        current_row += 1

    last_data_row = max(1, current_row - 2)
    last_col = get_column_letter(NUM_COLS_SIX_TEAM)
    ws.print_area = f"A1:{last_col}{last_data_row}"


def write_schedule_sheet(wb: Workbook, sheet_name: str, sections: list[Section]) -> None:
    ws = wb.create_sheet(title=sheet_name)

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    current_row = 1
    overall_round = 1

    for s_idx, section in enumerate(sections):
        style_merged_banner_row(
            ws,
            current_row,
            1,
            NUM_COLS,
            title=f"Section {s_idx + 1}",
            font=FONT_SECTION,
            fill=FILL_SECTION,
        )
        current_row += 1

        headers = ["Overall", "Round", "Home", "Away", "Winner"]
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=label)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER
        current_row += 1

        for g_idx, game in enumerate(section.games):
            row_fill = FILL_ROW_ALT if g_idx % 2 else None
            row_vals = [overall_round, game.round_label, game.home, game.away, None]
            overall_round += 1
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = FONT_BODY
                cell.border = BORDER
                if col_idx == NUM_COLS:
                    cell.fill = FILL_WINNER_COL
                elif row_fill:
                    cell.fill = row_fill
                if col_idx in (3, 4):
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_CENTER
            current_row += 1

        current_row += 1

    last_data_row = max(1, current_row - 2)
    last_col = get_column_letter(NUM_COLS)
    ws.print_area = f"A1:{last_col}{last_data_row}"


def build(input_path: Path, output_path: Path) -> None:
    src = openpyxl.load_workbook(input_path, read_only=False, data_only=True)
    out_wb: Workbook | None = None
    try:
        out_wb = Workbook()
        del_sheet = out_wb.active
        out_wb.remove(del_sheet)

        sections_2 = parse_sections(src["2 team"])
        write_schedule_sheet(out_wb, "2 team", sections_2)

        sections_3 = parse_sections(src["3 team"])
        write_schedule_sheet(out_wb, "3 team", sections_3)

        sections_4 = parse_sections(src["4 team"])
        if len(sections_4) != 3:
            raise ValueError(f"Expected 3 sections in '4 team', got {len(sections_4)}")
        sections_4_fixed = sections_4[:2] + [fix_4team_section3(sections_4[2])]
        write_schedule_sheet(out_wb, "4 team - Long", sections_4_fixed)
        sections_4_short = [section_take_first_n(s, 4) for s in sections_4_fixed]
        write_schedule_sheet(out_wb, "4 team - Short", sections_4_short)

        sections_5 = parse_sections(src["5 team"])
        write_schedule_sheet(out_wb, "5 team - Long", sections_5)
        sections_5_short = [section_take_first_n(s, 5) for s in sections_5]
        write_schedule_sheet(out_wb, "5 team - Short", sections_5_short)

        sections_6 = fix_6team_sections(parse_sections(src["6 team"]))
        write_six_team_wave_sheet(out_wb, "6 team - waves", sections_6)
        write_schedule_sheet(out_wb, "6 team - by section", sections_6)

        out_wb.save(output_path)
    finally:
        src.close()
        if out_wb is not None:
            out_wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate print-ready Open Gym schedules.")
    parser.add_argument("--input", type=Path, default=DEFAULT_IN)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()

    if not input_path.exists():
        raise SystemExit(f"Error: input file not found: {input_path}")
    if not output_path.parent.exists():
        raise SystemExit(f"Error: output directory does not exist: {output_path.parent}")

    build(input_path, output_path)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
