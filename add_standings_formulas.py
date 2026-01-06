import openpyxl
from openpyxl.utils import get_column_letter

# Load the workbook
wb = openpyxl.load_workbook('public/league_schedules/2026 Winter She_They League.xlsx')

# Week sheet names
week_sheets = ['Week 1 (1.4)', 'Week 2 (1.11)', 'Week 3 (1.25)', 'Week 4 (2.1)', 'Week 5 (2.8)', 'Week 6 (2.15)']

# Team names
teams = ['Persephone', 'Artemis', 'Dionysus', 'Athena']

print("Adding formulas to week sheets...")

# Add formulas to each week sheet
for sheet_name in week_sheets:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Processing {sheet_name}...")
        
        # Add header
        ws['A30'] = 'Team Wins This Week'
        ws['A31'] = 'Team Name'
        ws['B31'] = 'Wins'
        
        # Add team names and formulas
        for i, team in enumerate(teams, start=32):
            # Team name in column A
            ws[f'A{i}'] = team
            
            # Formula in column B
            if team == 'Athena':
                # Use wildcard for Athena to handle trailing spaces
                formula = f'=COUNTIFS(B:B,"Athena*",C:C,"<>")+COUNTIFS(D:D,"Athena*",E:E,"<>")'
            else:
                formula = f'=COUNTIFS(B:B,"{team}",C:C,"<>")+COUNTIFS(D:D,"{team}",E:E,"<>")'
            
            ws[f'B{i}'] = formula
            print(f"    Added formula for {team} in B{i}")

print("\nUpdating League Standings sheet...")

# Update League Standings sheet
if 'League Standings' in wb.sheetnames:
    ws = wb['League Standings']
    
    # Find where to add the new teams - look for empty rows or add after existing teams
    # Check rows 17-30 for existing teams or empty space
    start_row = None
    for row in range(17, 31):
        if ws[f'A{row}'].value is None or ws[f'A{row}'].value == '':
            start_row = row
            break
    
    # If no empty row found, add after row 23
    if start_row is None:
        start_row = 24
    
    # Add header if not present
    if ws['A16'].value != 'Teams':
        ws['A16'] = 'Teams'
        ws['B16'] = 'Wins'
    
    # Add team names and formulas
    for i, team in enumerate(teams):
        row = start_row + i
        # Team name in column A
        ws[f'A{row}'] = team
        
        # Create formula that sums wins from all week sheets
        # Each team's wins are in B32, B33, B34, B35 for Persephone, Artemis, Dionysus, Athena
        row_num = 32 + i
        formula_parts = []
        for week_sheet in week_sheets:
            formula_parts.append(f"'{week_sheet}'!B{row_num}")
        
        formula = '=' + '+'.join(formula_parts)
        ws[f'B{row}'] = formula
        print(f"  Added {team} in row {row} with formula in B{row}")

# Save the workbook
print("\nSaving workbook...")
wb.save('public/league_schedules/2026 Winter She_They League.xlsx')
print("Done! Formulas have been added to the spreadsheet.")

